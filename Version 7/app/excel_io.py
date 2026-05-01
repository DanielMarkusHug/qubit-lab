"""Excel upload and workbook-introspection helpers."""

from __future__ import annotations

import tempfile
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename

from app.config import Config
from app.schemas import ApiError, json_safe


def validate_excel_upload(file_storage: FileStorage | None) -> str:
    if file_storage is None:
        raise ApiError(400, "missing_file", "Multipart form field 'file' is required.")

    original_name = file_storage.filename or ""
    safe_name = secure_filename(original_name)
    if not safe_name:
        raise ApiError(400, "missing_filename", "Uploaded file must have a filename.")

    suffix = Path(safe_name).suffix.lower()
    if suffix not in Config.ALLOWED_UPLOAD_SUFFIXES:
        raise ApiError(400, "invalid_file_type", "Only .xlsx workbooks are supported.")

    return safe_name


def save_upload_to_temp(file_storage: FileStorage | None) -> tuple[Path, str]:
    safe_name = validate_excel_upload(file_storage)
    suffix = Path(safe_name).suffix.lower()
    tmp = tempfile.NamedTemporaryFile(prefix="qaoa_rqp_", suffix=suffix, delete=False)
    tmp_path = Path(tmp.name)
    tmp.close()

    try:
        file_storage.save(tmp_path)
    except Exception:
        cleanup_temp_file(tmp_path)
        raise

    return tmp_path, safe_name


def cleanup_temp_file(path: Path | None) -> None:
    if path is None:
        return
    try:
        Path(path).unlink(missing_ok=True)
    except Exception:
        pass


def workbook_structure(path: Path) -> dict:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ApiError(400, "invalid_workbook", "Uploaded file is not a readable .xlsx workbook.") from exc

    try:
        sheet_names = list(wb.sheetnames)
        ignored = [name for name in sheet_names if name in Config.IGNORED_OUTPUT_SHEETS]
        input_sheet_names = [name for name in sheet_names if name not in Config.IGNORED_OUTPUT_SHEETS]
        sheets = []

        for sheet_name in input_sheet_names:
            ws = wb[sheet_name]
            first_row = _row_values(ws, 1)
            second_row = _row_values(ws, 2)
            detected_header_row = 2 if _non_empty_count(second_row) > _non_empty_count(first_row) else 1
            detected_headers = second_row if detected_header_row == 2 else first_row
            sheets.append(
                {
                    "name": sheet_name,
                    "row_count": int(ws.max_row or 0),
                    "column_count": int(ws.max_column or 0),
                    "first_row_headers": json_safe(first_row),
                    "detected_header_row": detected_header_row,
                    "detected_headers": json_safe(detected_headers),
                }
            )

        return {
            "sheet_names": sheet_names,
            "input_sheet_names": input_sheet_names,
            "ignored_output_sheets": ignored,
            "sheets": sheets,
        }
    finally:
        wb.close()


def validate_required_input_sheets(path: Path, required_sheets: Iterable[str] | None = None) -> None:
    required = tuple(required_sheets or Config.REQUIRED_CLASSICAL_INPUT_SHEETS)
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ApiError(400, "invalid_workbook", "Uploaded file is not a readable .xlsx workbook.") from exc

    try:
        present = set(wb.sheetnames)
    finally:
        wb.close()

    missing = [sheet for sheet in required if sheet not in present]
    if missing:
        raise ApiError(
            400,
            "missing_required_sheets",
            "Workbook is missing required input sheets for classical_only mode.",
            {"missing_sheets": missing, "required_sheets": list(required)},
        )


def _row_values(ws, row_number: int) -> list:
    if ws.max_row is None or ws.max_row < row_number:
        return []
    values = next(ws.iter_rows(min_row=row_number, max_row=row_number, values_only=True), ())
    return _trim_trailing_empty(list(values))


def _trim_trailing_empty(values: list) -> list:
    trimmed = list(values)
    while trimmed and trimmed[-1] is None:
        trimmed.pop()
    return trimmed


def _non_empty_count(values: list) -> int:
    return sum(1 for value in values if value not in (None, ""))

