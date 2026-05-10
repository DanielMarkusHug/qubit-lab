from __future__ import annotations

import io
import json
import os
import sys
import datetime as dt
import importlib.util
import types
from types import SimpleNamespace
from pathlib import Path

import pytest
import numpy as np
import pandas as pd
from openpyxl import load_workbook


os.environ.setdefault("QAOA_RQP_LOCAL_DEV", "1")

REPO_ROOT = Path(__file__).resolve().parents[2]
VERSION_DIR = REPO_ROOT / "Version 9"
SAMPLE_WORKBOOK = REPO_ROOT / "Version 3" / "parametric_assets_only_input_small.xlsx"
EXTENDED_WORKBOOK = REPO_ROOT / "Version 3" / "parametric_assets_only_input_extended.xlsx"

sys.path.insert(0, str(VERSION_DIR))

import app.main as main_module  # noqa: E402
import app.cloud_run_jobs as cloud_run_jobs_module  # noqa: E402
import app.job_worker as job_worker_module  # noqa: E402
import app.memory_monitor as memory_monitor_module  # noqa: E402
import app.qaoa_engine as qaoa_engine_module  # noqa: E402
from app.main import app, create_app  # noqa: E402
from app.qubo_builder import build_qubo_from_workbook  # noqa: E402
from app.key_store import FirestoreApiKeyStore  # noqa: E402
from app.job_storage import GcsJobStorage, get_job_storage  # noqa: E402
from app.job_store import get_job_store, initial_job_document  # noqa: E402
from app.run_ledger import FirestoreRunLedger, RunLedger, _LOCAL_PUBLIC_RUNS  # noqa: E402
from app.schemas import ApiError  # noqa: E402
from app.usage_policy import (  # noqa: E402
    RuntimeInputs,
    UsageContext,
    generate_key_hash,
    estimate_policy_result,
    load_usage_config,
    validate_problem_policy,
    validate_secret_configuration,
)
from app.workbook_diagnostics import candidate_export_diagnostics, workbook_warnings  # noqa: E402

KEY_ADMIN_PATH = VERSION_DIR / "scripts" / "key_admin_firestore.py"
_key_admin_spec = importlib.util.spec_from_file_location("key_admin_firestore", KEY_ADMIN_PATH)
key_admin_firestore = importlib.util.module_from_spec(_key_admin_spec)
assert _key_admin_spec.loader is not None
_key_admin_spec.loader.exec_module(key_admin_firestore)


DEMO_KEY_ID = "demo-qualified-001"
INTERNAL_POWER_KEY_ID = "demo-internal-power-001"
INTERNAL_POWER_KEY = "INTERNAL-POWER-123"


@pytest.fixture(autouse=True)
def _default_env(monkeypatch):
    monkeypatch.setenv("QAOA_RQP_LOCAL_DEV", "1")
    monkeypatch.delenv("QAOA_KEY_STORE", raising=False)
    monkeypatch.delenv("QAOA_RQP_KEY_STORE", raising=False)
    monkeypatch.delenv("QAOA_LEDGER_STORE", raising=False)
    monkeypatch.delenv("QAOA_RQP_LEDGER_STORE", raising=False)
    monkeypatch.delenv("QAOA_RQP_ENABLE_LOCAL_LEDGER", raising=False)
    monkeypatch.delenv("QAOA_RQP_LEDGER_PATH", raising=False)
    monkeypatch.delenv("QAOA_JOB_STORE", raising=False)
    monkeypatch.delenv("QAOA_JOB_STORAGE", raising=False)
    monkeypatch.delenv("QAOA_JOB_BUCKET", raising=False)
    monkeypatch.delenv("RUN_JOBS_INLINE_FOR_LOCAL", raising=False)


def _xlsx_upload_data():
    return {
        "file": (
            io.BytesIO(SAMPLE_WORKBOOK.read_bytes()),
            "parametric_assets_only_input_small.xlsx",
        )
    }


def _post_run(headers=None, **fields):
    data = _xlsx_upload_data()
    data.update(fields)
    return app.test_client().post(
        "/run-qaoa",
        headers=headers or {},
        data=data,
        content_type="multipart/form-data",
    )


def _post_async(headers=None, **fields):
    data = _xlsx_upload_data()
    data.update(fields)
    return app.test_client().post(
        "/run-qaoa-async",
        headers=headers or {},
        data=data,
        content_type="multipart/form-data",
    )


def _post_run_file(path: Path, headers=None, **fields):
    data = {
        "file": (
            io.BytesIO(path.read_bytes()),
            path.name,
        )
    }
    data.update(fields)
    return app.test_client().post(
        "/run-qaoa",
        headers=headers or {},
        data=data,
        content_type="multipart/form-data",
    )


def _post_inspect(path: Path = SAMPLE_WORKBOOK, headers=None, **fields):
    data = {
        "file": (
            io.BytesIO(path.read_bytes()),
            path.name,
        )
    }
    data.update(fields)
    return app.test_client().post(
        "/inspect-workbook",
        headers=headers or {},
        data=data,
        content_type="multipart/form-data",
    )


def _limited_workbook(tmp_path, qubits: int = 4, source_workbook: Path = SAMPLE_WORKBOOK) -> Path:
    workbook_path = tmp_path / f"qaoa_limited_{qubits}_qubits.xlsx"
    workbook = load_workbook(source_workbook)
    assets = workbook["Assets"]
    headers = [cell.value for cell in assets[2]]
    allowed_col = headers.index("Allowed") + 1
    ticker_col = headers.index("Ticker") + 1
    seen = 0
    for row_idx in range(3, assets.max_row + 1):
        if assets.cell(row_idx, ticker_col).value in (None, ""):
            continue
        seen += 1
        assets.cell(row_idx, allowed_col).value = 1 if seen <= qubits else 0
    workbook.save(workbook_path)
    return workbook_path


def _indicative_cost_workbook(
    tmp_path,
    *,
    keep_legacy: bool = False,
    add_usd: float = 0.0,
    source_workbook: Path = SAMPLE_WORKBOOK,
) -> Path:
    workbook_path = tmp_path / "indicative_market_cost_input.xlsx"
    workbook = load_workbook(source_workbook)
    assets = workbook["Assets"]
    headers = [cell.value for cell in assets[2]]
    approx_col = headers.index("Approx Cost USD") + 1
    existing_indicative_col = (
        headers.index("Indicative Market Cost USD") + 1
        if "Indicative Market Cost USD" in headers
        else None
    )
    if keep_legacy:
        indicative_col = existing_indicative_col or assets.max_column + 1
        assets.cell(2, indicative_col).value = "Indicative Market Cost USD"
    else:
        indicative_col = existing_indicative_col or approx_col
        assets.cell(2, indicative_col).value = "Indicative Market Cost USD"
    for row_idx in range(3, assets.max_row + 1):
        value = assets.cell(row_idx, approx_col).value
        if value not in (None, ""):
            assets.cell(row_idx, indicative_col).value = float(value) + float(add_usd)
    workbook.save(workbook_path)
    return workbook_path


def _type_constraint_workbook(
    tmp_path,
    *,
    qubits: int = 3,
    constraint_count: int | str | float = 1,
    include_type_a_size: bool = True,
    include_type_b_size: bool = False,
    type_a_budget=100.0,
    type_a_penalty=30.0,
    type_b_budget=50.0,
    type_b_penalty=12.0,
    type_a_values: list[float] | None = None,
    type_b_values: list[float] | None = None,
    decision_roles: list[str] | None = None,
) -> Path:
    source_path = _limited_workbook(tmp_path, qubits=qubits)
    workbook_path = tmp_path / f"type_constraints_{len(list(tmp_path.glob('type_constraints_*.xlsx')))}.xlsx"
    workbook = load_workbook(source_path)
    assets = workbook["Assets"]
    settings = workbook["Settings"]

    if include_type_a_size:
        _set_asset_column(assets, "Type A Size", type_a_values or [20.0, 50.0, 80.0, 10.0, 5.0])
    if include_type_b_size:
        _set_asset_column(assets, "Type B Size", type_b_values or [5.0, 15.0, 30.0, 45.0, 60.0])
    if decision_roles is not None:
        _set_asset_roles(assets, decision_roles)

    _set_setting(settings, "Additional Type Constraints", constraint_count)
    count_as_float = float(constraint_count)
    if count_as_float >= 1:
        _set_setting(settings, "Type A Name", "Bond")
        _set_setting(settings, "Type A Budget", type_a_budget)
        _set_setting(settings, "Type A Budget Penalty", type_a_penalty)
    if count_as_float >= 2:
        _set_setting(settings, "Type B Name", "Equity")
        _set_setting(settings, "Type B Budget", type_b_budget)
        _set_setting(settings, "Type B Budget Penalty", type_b_penalty)

    workbook.save(workbook_path)
    return workbook_path


def _set_asset_column(assets, header: str, values: list[float]) -> None:
    headers = [cell.value for cell in assets[2]]
    if header in headers:
        col_idx = headers.index(header) + 1
    else:
        col_idx = assets.max_column + 1
        assets.cell(2, col_idx).value = header

    ticker_col = headers.index("Ticker") + 1
    value_idx = 0
    for row_idx in range(3, assets.max_row + 1):
        if assets.cell(row_idx, ticker_col).value in (None, ""):
            continue
        assets.cell(row_idx, col_idx).value = values[value_idx % len(values)]
        value_idx += 1


def _set_asset_roles(assets, roles: list[str]) -> None:
    headers = [cell.value for cell in assets[2]]
    ticker_col = headers.index("Ticker") + 1
    if "Decision Role" in headers:
        role_col = headers.index("Decision Role") + 1
    else:
        role_col = assets.max_column + 1
        assets.cell(2, role_col).value = "Decision Role"

    value_idx = 0
    for row_idx in range(3, assets.max_row + 1):
        if assets.cell(row_idx, ticker_col).value in (None, ""):
            continue
        assets.cell(row_idx, role_col).value = roles[value_idx % len(roles)]
        value_idx += 1


def _set_setting(settings, key: str, value) -> None:
    headers = [cell.value for cell in settings[2]]
    key_col = headers.index("Key") + 1
    value_col = headers.index("Value") + 1
    for row_idx in range(3, settings.max_row + 1):
        if settings.cell(row_idx, key_col).value == key:
            settings.cell(row_idx, value_col).value = value
            return
    row_idx = settings.max_row + 1
    settings.cell(row_idx, key_col).value = key
    settings.cell(row_idx, value_col).value = value


def _enable_temp_ledger(monkeypatch, tmp_path):
    ledger_path = tmp_path / "run_ledger.json"
    monkeypatch.setenv("QAOA_RQP_ENABLE_LOCAL_LEDGER", "1")
    monkeypatch.setenv("QAOA_RQP_LEDGER_PATH", str(ledger_path))
    return ledger_path


class _FakeFirestoreSnapshot:
    def __init__(self, doc_id, data):
        self.id = doc_id
        self.exists = data is not None
        self._data = dict(data or {})

    def to_dict(self):
        return dict(self._data)


class _FakeFirestoreDocument:
    def __init__(self, store, collection, doc_id):
        self.store = store
        self.collection = collection
        self.id = doc_id

    def get(self, transaction=None):
        return _FakeFirestoreSnapshot(self.id, self.store.data.get(self.collection, {}).get(self.id))

    def set(self, payload, merge=False):
        collection = self.store.data.setdefault(self.collection, {})
        if merge and self.id in collection:
            collection[self.id].update(payload)
        else:
            collection[self.id] = dict(payload)


class _FakeFirestoreQuery:
    def __init__(self, store, collection, field=None, value=None):
        self.store = store
        self.collection = collection
        self.field = field
        self.value = value
        self.limit_count = None

    def limit(self, count):
        self.limit_count = count
        return self

    def order_by(self, *_args, **_kwargs):
        return self

    def stream(self):
        matches = []
        for doc_id, data in self.store.data.get(self.collection, {}).items():
            if self.field is None or data.get(self.field) == self.value:
                matches.append(_FakeFirestoreSnapshot(doc_id, data))
            if self.limit_count is not None and len(matches) >= self.limit_count:
                break
        return matches


class _FakeFirestoreCollection:
    def __init__(self, store, name):
        self.store = store
        self.name = name

    def document(self, doc_id):
        return _FakeFirestoreDocument(self.store, self.name, doc_id)

    def limit(self, count):
        return _FakeFirestoreQuery(self.store, self.name).limit(count)

    def order_by(self, *_args, **_kwargs):
        return _FakeFirestoreQuery(self.store, self.name).order_by(*_args, **_kwargs)

    def where(self, field=None, operator=None, value=None, *, filter=None):
        if filter is not None:
            self.store.keyword_where_calls += 1
            field = getattr(filter, "field_path", None)
            operator = getattr(filter, "op_string", None)
            value = getattr(filter, "value", None)
        else:
            self.store.positional_where_calls += 1
        assert operator == "=="
        return _FakeFirestoreQuery(self.store, self.name, field, value)

    def stream(self):
        return _FakeFirestoreQuery(self.store, self.name).stream()


class _FakeFirestoreTransaction:
    def update(self, doc_ref, updates):
        collection = doc_ref.store.data.setdefault(doc_ref.collection, {})
        collection.setdefault(doc_ref.id, {}).update(updates)

    def set(self, doc_ref, payload, merge=False):
        collection = doc_ref.store.data.setdefault(doc_ref.collection, {})
        if merge and doc_ref.id in collection:
            collection[doc_ref.id].update(payload)
        else:
            collection[doc_ref.id] = dict(payload)


class _FakeFirestoreClient:
    def __init__(self):
        self.data = {
            "qaoa_keys": {},
            "qaoa_usage_events": {},
            "qaoa_public_run_state_v9": {},
            "qaoa_public_run_locks_v9": {},
            "qaoa_key_run_state_v9": {},
            "qaoa_key_run_locks_v9": {},
            "qaoa_jobs_v9": {},
        }
        self.positional_where_calls = 0
        self.keyword_where_calls = 0

    def collection(self, name):
        return _FakeFirestoreCollection(self, name)

    def transaction(self):
        return _FakeFirestoreTransaction()


class _FakeGcsBlob:
    def __init__(self, store: dict[str, object], name: str):
        self.store = store
        self.name = name

    def upload_from_filename(self, filename):
        self.store[self.name] = Path(filename).read_bytes()

    def download_to_filename(self, filename):
        Path(filename).write_bytes(self.store[self.name])

    def upload_from_string(self, data, content_type=None):
        self.store[self.name] = data
        self.store[f"{self.name}:content_type"] = content_type

    def download_as_text(self):
        value = self.store[self.name]
        return value.decode("utf-8") if isinstance(value, bytes) else str(value)


class _FakeGcsBucket:
    def __init__(self, store: dict[str, object]):
        self.store = store

    def blob(self, name: str):
        return _FakeGcsBlob(self.store, name)


class _FakeGcsClient:
    def __init__(self):
        self.objects: dict[str, object] = {}

    def bucket(self, _name: str):
        return _FakeGcsBucket(self.objects)


def _identity_transactional(callback):
    return callback


def _firestore_usage_context(key_record):
    return SimpleNamespace(
        authenticated=True,
        key_record=key_record,
        usage_level_name=key_record["level"],
        usage_level={"level_id": 1, "show_identity": True},
    )


def _policy_result():
    return SimpleNamespace(
        estimated_runtime_sec=2.0,
        raw_estimated_runtime_sec=2.0,
        max_estimated_runtime_sec=60.0,
        within_limit=True,
        n_qubits=3,
        candidate_count=8,
        runtime_limit_source="usage_level",
        runtime_inputs=SimpleNamespace(
            layers=1,
            iterations=60,
            restarts=1,
            warm_start=False,
            qaoa_shots=None,
            restart_perturbation=None,
            random_seed=None,
        ),
        effective_settings={
            "qaoa_shots_display": "not_applicable",
            "shots_mode": "disabled",
            "random_seed": None,
        },
    )


def _fake_budget_warning_optimizer():
    return SimpleNamespace(
        n=2,
        qaoa_p=1,
        qaoa_maxiter=10,
        qaoa_multistart_restarts=1,
        qaoa_layerwise_warm_start=False,
        qaoa_shots=None,
        qaoa_restart_perturbation=None,
        lambda_budget=50.0,
        lambda_variance=6.0,
        risk_free=0.04,
        settings={},
        budget_usd=250_000.0,
        fixed_cost=np.array([1_600_000.0]),
        opt_cost=np.array([100_000.0, 250_000.0]),
        variable_options_df=pd.DataFrame({"Ticker": ["AAA", "BBB"]}),
        fixed_options_df=pd.DataFrame({"Ticker": ["FIXED"]}),
        asset_universe=["AAA", "BBB", "FIXED"],
        assets_df=pd.DataFrame({"Ticker": ["AAA", "BBB", "FIXED"]}),
        Q=np.zeros((2, 2)),
        classical_results=None,
    )


class _RouteLockLedger:
    def __init__(self, *, fail_lock: ApiError | None = None, fail_public_slot: ApiError | None = None):
        self.fail_lock = fail_lock
        self.fail_public_slot = fail_public_slot
        self.acquired = False
        self.released = False
        self.public_acquired = False
        self.public_released = False
        self.completed = False
        self.failed = False
        self.rejected = False
        self.consumed = False

    def can_consume_run(self, _key_record):
        return True

    def get_remaining_runs(self, _key_record):
        return 10

    def acquire_run_lock(self, _key_record, _run_id, **_kwargs):
        if self.fail_lock is not None:
            raise self.fail_lock
        self.acquired = True
        return {"acquired": True}

    def release_run_lock(self, _key_record, _run_id):
        self.released = True
        return True

    def acquire_public_run_slot(self, *_args, **_kwargs):
        if self.fail_public_slot is not None:
            raise self.fail_public_slot
        self.public_acquired = True
        return {"acquired": True}

    def release_public_run_slot(self, _run_id):
        self.public_released = True
        return True

    def record_run_started(self, **_kwargs):
        return None

    def consume_run(self, _key_record, _run_id):
        self.consumed = bool(_key_record)
        return bool(_key_record)

    def record_run_completed(self, **_kwargs):
        self.completed = True

    def record_run_rejected(self, **_kwargs):
        self.rejected = True

    def record_run_failed(self, **_kwargs):
        self.failed = True

    def license_status(self, usage_context):
        return {
            "authenticated": bool(usage_context.authenticated),
            "key_id": (usage_context.key_record or {}).get("key_id"),
            "usage_level": usage_context.usage_level_name,
            "remaining_runs": 10,
        }

    def safe_license_summary(self, usage_context):
        return {
            "authenticated": bool(usage_context.authenticated),
            "key_id": (usage_context.key_record or {}).get("key_id"),
            "usage_level": usage_context.usage_level_name,
            "remaining_runs": 10,
        }


def _patch_fast_run(monkeypatch, ledger: _RouteLockLedger, *, fail_execution: bool = False):
    optimizer = SimpleNamespace(
        n=3,
        qaoa_p=1,
        qaoa_maxiter=10,
        qaoa_multistart_restarts=1,
        qaoa_layerwise_warm_start=False,
        qaoa_shots=None,
        classical_results=[{"bitstring": "101"}],
    )
    monkeypatch.setattr(main_module, "get_run_ledger", lambda: ledger)
    monkeypatch.setattr(main_module, "validate_required_input_sheets", lambda _path: None)
    monkeypatch.setattr(main_module, "workbook_structure", lambda _path: {})
    monkeypatch.setattr(main_module, "build_qubo_from_workbook", lambda _path, _log, _form=None: optimizer)
    monkeypatch.setattr(main_module, "validate_problem_policy", lambda *_args, **_kwargs: _policy_result())
    if fail_execution:
        def fail_optimizer(_optimizer, _logs):
            raise RuntimeError("simulated optimizer failure")

        monkeypatch.setattr(main_module, "run_classical_optimizer", fail_optimizer)
    else:
        monkeypatch.setattr(main_module, "run_classical_optimizer", lambda current, logs: (current, logs))
    monkeypatch.setattr(
        main_module,
        "build_classical_response",
        lambda run_id, *_args, **_kwargs: {
            "status": "completed",
            "run_id": run_id,
            "model_version": "9.1.0",
            "mode": "classical_only",
        },
    )


def test_health():
    response = app.test_client().get("/health")

    assert response.status_code == 200
    assert response.get_json() == {"status": "ok"}


def test_root_reports_current_version():
    response = app.test_client().get("/")
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["service"] == "qaoa-rqp-api-v9"
    assert payload["version"] == "9.1.0"


def test_capabilities():
    response = app.test_client().get("/capabilities")
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["service"] == "qaoa-rqp-api-v9"
    assert payload["version"] == "9.1.0"
    assert payload["supported_modes"] == ["classical_only", "qaoa_lightning_sim", "qaoa_tensor_sim"]
    assert payload["disabled_modes"] == []
    assert payload["mode_aliases"]["qaoa_limited"] == "qaoa_lightning_sim"
    assert payload["default_response_level"] == "full"
    assert "qaoa_lightning_sim" in payload["usage_levels"]["public_demo"]["allowed_modes"]
    assert "qaoa_full" not in payload["usage_levels"]["tester"]["allowed_modes"]
    assert "qaoa_tensor_sim" in payload["usage_levels"]["tester"]["allowed_modes"]
    assert payload["usage_levels"]["public_demo"]["allowed_response_levels"] == ["compact", "standard", "full"]
    assert payload["usage_levels"]["public_demo"]["max_parallel_runs"] == 5
    assert payload["usage_levels"]["qualified_demo"]["allowed_response_levels"] == ["compact", "standard", "full"]
    assert payload["qaoa_effective_limits"]["public_demo"]["qaoa_lightning_sim"]["max_qubits"] == 8
    assert payload["qaoa_effective_limits"]["public_demo"]["qaoa_lightning_sim"]["max_layers"] == 1
    assert payload["usage_levels"]["internal_power"]["qaoa_lightning_sim_limits"]["max_qubits"] == 24
    assert payload["qaoa_effective_limits"]["tester"]["qaoa_lightning_sim"]["max_qubits"] == 16
    assert payload["qaoa_effective_limits"]["tester"]["qaoa_tensor_sim"]["max_qubits"] == 16
    assert payload["qaoa_effective_limits"]["tester"]["qaoa_lightning_sim"]["max_layers"] == 6
    assert payload["qaoa_effective_limits"]["tester"]["qaoa_lightning_sim"]["max_iterations"] == 200
    assert payload["qaoa_effective_limits"]["tester"]["qaoa_lightning_sim"]["max_restarts"] == 3
    assert payload["qaoa_effective_limits"]["internal_power"]["qaoa_lightning_sim"]["max_layers"] == 8
    assert payload["default_worker_profile"] == "small"
    assert payload["worker_profiles"]["large"]["cpu"] == 4
    assert payload["worker_profiles"]["large"]["memory_gib"] == 8
    assert "demo_keys" not in payload
    assert "KEY_HASH_SECRET" not in str(payload)
    assert "key_hash" not in str(payload)


def test_async_submission_creates_job_and_returns_job_id(monkeypatch, tmp_path):
    monkeypatch.setattr(main_module.Config, "LOCAL_JOB_DIR", tmp_path / "jobs")
    ledger = _RouteLockLedger()
    _patch_fast_run(monkeypatch, ledger)
    monkeypatch.setattr(main_module, "trigger_cloud_run_job", lambda job_id: {"triggered": False, "mode": "test"})

    response = _post_async(mode="classical_only")
    payload = response.get_json()

    assert response.status_code == 202
    assert payload["ok"] is True
    assert payload["job_id"].startswith("job_")
    assert payload["status"] == "queued"
    assert payload["status_url"] == f"/jobs/{payload['job_id']}/status"
    assert payload["result_url"] == f"/jobs/{payload['job_id']}/result"
    job = get_job_store().get_job(payload["job_id"])
    assert job["status"] == "queued"
    assert job["input"]["original_filename"] == "parametric_assets_only_input_small.xlsx"
    assert job["result"]["available"] is False
    assert "key_hash" in job
    assert ledger.public_acquired is True
    assert ledger.public_released is False


def test_async_submission_stores_effective_random_seed(monkeypatch, tmp_path):
    monkeypatch.setattr(main_module.Config, "LOCAL_JOB_DIR", tmp_path / "jobs")
    ledger = _RouteLockLedger()
    _patch_fast_run(monkeypatch, ledger)
    monkeypatch.setattr(main_module, "trigger_cloud_run_job", lambda job_id: {"triggered": False, "mode": "test"})
    def _seeded_policy_result(*_args, **_kwargs):
        base = _policy_result()
        base.runtime_inputs.random_seed = 987
        base.effective_settings["random_seed"] = 987
        base.effective_settings["rng_seed"] = 987
        return base

    monkeypatch.setattr(main_module, "validate_problem_policy", _seeded_policy_result)

    response = _post_async(headers={"X-API-Key": "DEMO-123"}, mode="classical_only", random_seed="987")
    payload = response.get_json()
    job = get_job_store().get_job(payload["job_id"])

    assert response.status_code == 202
    assert job["settings"]["random_seed"] == "987"
    assert job["settings"]["rng_seed"] == "987"
    assert job["settings"]["effective_settings"]["random_seed"] == 987
    assert job["settings"]["runtime_inputs"]["random_seed"] == 987
    assert any("Random seed: 987" in line for line in job["logs_tail"])


def test_async_status_and_result_before_completion(monkeypatch, tmp_path):
    monkeypatch.setattr(main_module.Config, "LOCAL_JOB_DIR", tmp_path / "jobs")
    ledger = _RouteLockLedger()
    _patch_fast_run(monkeypatch, ledger)
    monkeypatch.setattr(main_module, "trigger_cloud_run_job", lambda job_id: {"triggered": False, "mode": "test"})

    submit_response = _post_async(mode="classical_only")
    job_id = submit_response.get_json()["job_id"]

    status_response = app.test_client().get(f"/jobs/{job_id}/status")
    status_payload = status_response.get_json()

    assert status_response.status_code == 200
    assert status_payload["job_id"] == job_id
    assert status_payload["status"] == "queued"
    assert status_payload["result_available"] is False
    assert {
        "job_id",
        "status",
        "phase",
        "progress",
        "latest_log",
        "logs_tail",
        "created_at",
        "started_at",
        "heartbeat_at",
        "finished_at",
        "result_available",
        "error",
    }.issubset(status_payload)

    result_response = app.test_client().get(f"/jobs/{job_id}/result")
    result_payload = result_response.get_json()

    assert result_response.status_code == 409
    assert result_payload["error"]["code"] == "job_not_completed"
    assert result_payload["error"]["details"]["job_id"] == job_id


def test_async_worker_profile_defaults_to_small(monkeypatch, tmp_path):
    monkeypatch.setattr(main_module.Config, "LOCAL_JOB_DIR", tmp_path / "jobs")
    ledger = _RouteLockLedger()
    _patch_fast_run(monkeypatch, ledger)
    monkeypatch.setattr(main_module, "trigger_cloud_run_job", lambda job_id: {"triggered": False, "mode": "test"})

    response = _post_async(headers={"X-API-Key": "DEMO-123"}, mode="classical_only")
    payload = response.get_json()
    job = get_job_store().get_job(payload["job_id"])
    status = app.test_client().get(f"/jobs/{payload['job_id']}/status").get_json()

    assert response.status_code == 202
    assert payload["worker_profile"] == "small"
    assert payload["configured_cpu"] == 2
    assert payload["configured_memory_gib"] == 2
    assert job["worker_profile"] == "small"
    assert job["settings"]["worker_profile"] == "small"
    assert status["worker_profile"] == "small"
    assert status["configured_cpu"] == 2
    assert status["configured_memory_gib"] == 2


def test_async_invalid_worker_profile_is_rejected(monkeypatch, tmp_path):
    monkeypatch.setattr(main_module.Config, "LOCAL_JOB_DIR", tmp_path / "jobs")
    ledger = _RouteLockLedger()
    _patch_fast_run(monkeypatch, ledger)
    monkeypatch.setattr(
        main_module,
        "trigger_cloud_run_job",
        lambda _job_id: pytest.fail("invalid worker profile should not trigger a job"),
    )

    response = _post_async(headers={"X-API-Key": "DEMO-123"}, mode="classical_only", worker_profile="huge")
    payload = response.get_json()

    assert response.status_code == 400
    assert payload["error"]["code"] == "invalid_worker_profile"
    assert ledger.acquired is False
    assert ledger.public_acquired is False


@pytest.mark.parametrize(
    ("api_key", "profile", "expected_status"),
    [
        ("DEMO-123", "small", 202),
        ("DEMO-123", "medium", 403),
        ("DEMO-123", "large", 403),
        ("TESTER-123", "small", 202),
        ("TESTER-123", "medium", 202),
        ("TESTER-123", "large", 403),
        (INTERNAL_POWER_KEY, "small", 202),
        (INTERNAL_POWER_KEY, "medium", 202),
        (INTERNAL_POWER_KEY, "large", 202),
    ],
)
def test_async_worker_profile_permissions(monkeypatch, tmp_path, api_key, profile, expected_status):
    monkeypatch.setattr(main_module.Config, "LOCAL_JOB_DIR", tmp_path / f"jobs-{api_key}-{profile}")
    ledger = _RouteLockLedger()
    _patch_fast_run(monkeypatch, ledger)
    monkeypatch.setattr(main_module, "trigger_cloud_run_job", lambda job_id: {"triggered": False, "mode": "test"})

    response = _post_async(headers={"X-API-Key": api_key}, mode="classical_only", worker_profile=profile)
    payload = response.get_json()

    assert response.status_code == expected_status
    if expected_status == 202:
        assert payload["worker_profile"] == profile
        assert get_job_store().get_job(payload["job_id"])["worker_profile"] == profile
    else:
        assert payload["error"]["code"] == "worker_profile_not_allowed"
        assert profile in payload["error"]["message"]


def test_sync_result_includes_worker_profile_metadata():
    response = _post_run(headers={"X-API-Key": "DEMO-123"}, mode="classical_only", worker_profile="small")
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["worker_profile"] == "small"
    assert payload["worker_profile_label"] == "Small"
    assert payload["configured_cpu"] == 2
    assert payload["configured_memory_gib"] == 2
    assert payload["memory_used_gib"] is None
    assert payload["memory_history"] == []
    assert payload["diagnostics"]["worker_profile"] == "small"
    assert payload["reporting"]["summary"]["configured_memory_gib"] == 2


def test_cloud_run_trigger_uses_profile_job_name(monkeypatch, tmp_path):
    monkeypatch.setattr(main_module.Config, "LOCAL_JOB_DIR", tmp_path / "jobs")
    assert cloud_run_jobs_module.trigger_cloud_run_job("job_profile_test", "large") == {
        "triggered": False,
        "mode": "local_manual",
        "worker_profile": "large",
        "worker_job_name": "qaoa-rqp-worker-large",
    }


def test_job_result_reads_completed_gcs_result(monkeypatch, tmp_path):
    monkeypatch.setattr(main_module.Config, "LOCAL_JOB_DIR", tmp_path / "jobs")
    job_id = "job_result_gcs"
    storage_path = "gs://unit-test-job-bucket/jobs/job_result_gcs/result.json"
    get_job_store().create_job(
        {
            "job_id": job_id,
            "status": "completed",
            "phase": "completed",
            "result": {"available": True, "storage_path": storage_path, "summary": {"status": "completed"}},
        }
    )
    fake_storage = SimpleNamespace(
        read_result_json=lambda path: {
            "status": "completed",
            "run_id": job_id,
            "storage_path_seen": path,
        }
    )
    monkeypatch.setattr(main_module, "get_job_storage", lambda: fake_storage)

    response = app.test_client().get(f"/jobs/{job_id}/result")
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["run_id"] == job_id
    assert payload["storage_path_seen"] == storage_path


def test_async_qaoa_full_disabled_before_lock(monkeypatch, tmp_path):
    monkeypatch.setattr(main_module.Config, "LOCAL_JOB_DIR", tmp_path / "jobs")
    ledger = _RouteLockLedger()
    _patch_fast_run(monkeypatch, ledger)
    monkeypatch.setattr(
        main_module,
        "trigger_cloud_run_job",
        lambda _job_id: pytest.fail("qaoa_full should be rejected before job trigger"),
    )

    response = _post_async(headers={"X-API-Key": INTERNAL_POWER_KEY}, mode="qaoa_full")
    payload = response.get_json()

    assert response.status_code == 400
    assert payload["error"]["code"] == "unsupported_mode"
    assert ledger.acquired is False
    assert ledger.public_acquired is False
    assert ledger.consumed is False


def test_async_trigger_failure_releases_authenticated_lock(monkeypatch, tmp_path):
    monkeypatch.setattr(main_module.Config, "LOCAL_JOB_DIR", tmp_path / "jobs")
    ledger = _RouteLockLedger()
    _patch_fast_run(monkeypatch, ledger)

    def fail_trigger(_job_id):
        raise ApiError(500, "cloud_run_job_trigger_failed", "Cloud Run Job trigger failed.")

    monkeypatch.setattr(main_module, "trigger_cloud_run_job", fail_trigger)

    response = _post_async(headers={"X-API-Key": INTERNAL_POWER_KEY}, mode="classical_only")
    payload = response.get_json()

    assert response.status_code == 500
    assert payload["error"]["code"] == "cloud_run_job_trigger_failed"
    assert ledger.acquired is True
    assert ledger.released is True
    assert ledger.consumed is False

    job_files = list((tmp_path / "jobs").glob("job_*/job.json"))
    assert len(job_files) == 1
    job = json.loads(job_files[0].read_text(encoding="utf-8"))
    assert job["status"] == "failed"
    assert job["phase"] == "submission_failed"


def test_async_trigger_failure_releases_public_slot(monkeypatch, tmp_path):
    monkeypatch.setattr(main_module.Config, "LOCAL_JOB_DIR", tmp_path / "jobs")
    ledger = _RouteLockLedger()
    _patch_fast_run(monkeypatch, ledger)
    monkeypatch.setattr(
        main_module,
        "trigger_cloud_run_job",
        lambda _job_id: (_ for _ in ()).throw(RuntimeError("trigger transport failed")),
    )

    response = _post_async(mode="classical_only")
    payload = response.get_json()

    assert response.status_code == 500
    assert payload["error"]["code"] == "async_job_submission_failed"
    assert ledger.public_acquired is True
    assert ledger.public_released is True
    assert ledger.consumed is False


def test_async_authenticated_active_lock_rejection(monkeypatch, tmp_path):
    monkeypatch.setattr(main_module.Config, "LOCAL_JOB_DIR", tmp_path / "jobs")
    error = ApiError(
        409,
        "active_run_exists",
        "This license key already has an active run. Please wait until it has finished.",
        {"active_run_id": "job_existing"},
    )
    ledger = _RouteLockLedger(fail_lock=error)
    _patch_fast_run(monkeypatch, ledger)

    response = _post_async(headers={"X-API-Key": INTERNAL_POWER_KEY}, mode="classical_only")
    payload = response.get_json()

    assert response.status_code == 409
    assert payload["error"]["code"] == "active_run_exists"
    assert ledger.rejected is True
    assert ledger.consumed is False
    assert ledger.released is False


def _create_local_worker_job(
    tmp_path,
    monkeypatch,
    *,
    job_id: str = "job_test_worker_001",
    settings: dict | None = None,
    worker_profile: str = "small",
):
    monkeypatch.setattr(main_module.Config, "LOCAL_JOB_DIR", tmp_path / "jobs")
    usage_level = load_usage_config()["usage_levels"]["public_demo"]
    usage_context = UsageContext(
        usage_level_name="public_demo",
        usage_level=usage_level,
        key_record=None,
        authenticated=False,
    )
    storage = get_job_storage()
    input_info = storage.save_input_from_path(job_id, SAMPLE_WORKBOOK, SAMPLE_WORKBOOK.name)
    get_job_store().create_job(
        initial_job_document(
            job_id=job_id,
            mode="classical_only",
            response_level="full",
            settings=settings or {"mode": "classical_only", "response_level": "full"},
            input_info=input_info,
            usage_context=usage_context,
            policy_result=_policy_result(),
            lock_type="public",
            key_hash=None,
            worker_profile=worker_profile,
        )
    )
    return job_id


def test_worker_marks_completed_and_releases_public_lock(monkeypatch, tmp_path):
    job_id = _create_local_worker_job(tmp_path, monkeypatch)
    ledger = _RouteLockLedger()
    optimizer = SimpleNamespace(
        n=3,
        qaoa_p=1,
        qaoa_maxiter=10,
        qaoa_multistart_restarts=1,
        qaoa_layerwise_warm_start=False,
        qaoa_shots=None,
        classical_results=[{"bitstring": "101"}],
        progress_callback=lambda _message, _progress=None: None,
    )
    monkeypatch.setattr(job_worker_module, "get_run_ledger", lambda: ledger)
    monkeypatch.setattr(job_worker_module, "validate_required_input_sheets", lambda _path: None)
    monkeypatch.setattr(job_worker_module, "workbook_structure", lambda _path: {})
    monkeypatch.setattr(job_worker_module, "build_qubo_from_workbook", lambda _path, _log, _settings=None: optimizer)
    monkeypatch.setattr(job_worker_module, "validate_problem_policy", lambda *_args, **_kwargs: _policy_result())
    monkeypatch.setattr(job_worker_module, "run_classical_optimizer", lambda current, logs: (current, logs))
    monkeypatch.setattr(
        job_worker_module,
        "build_classical_response",
        lambda run_id, *_args, **_kwargs: {
            "status": "completed",
            "run_id": run_id,
            "model_version": "9.1.0",
            "mode": "classical_only",
            "binary_variables": 3,
            "objective": 1.23,
        },
    )

    job_worker_module.run_job(job_id)
    job = get_job_store().get_job(job_id)

    assert job["status"] == "completed"
    assert job["result"]["available"] is True
    assert job["progress"]["progress_pct"] == 100.0
    assert ledger.completed is True
    assert ledger.public_released is True
    result_response = app.test_client().get(f"/jobs/{job_id}/result")
    assert result_response.status_code == 200
    assert result_response.get_json()["run_id"] == job_id


def test_worker_final_result_includes_worker_profile_and_memory(monkeypatch, tmp_path):
    settings = {"mode": "classical_only", "response_level": "full", "worker_profile": "medium"}
    job_id = _create_local_worker_job(
        tmp_path,
        monkeypatch,
        job_id="job_test_worker_memory",
        settings=settings,
        worker_profile="medium",
    )
    ledger = _RouteLockLedger()
    optimizer = SimpleNamespace(
        n=3,
        qaoa_p=1,
        qaoa_maxiter=10,
        qaoa_multistart_restarts=1,
        qaoa_layerwise_warm_start=False,
        qaoa_shots=None,
        classical_results=[{"bitstring": "101"}],
        progress_callback=lambda _message, _progress=None: None,
    )

    class FakeMemoryTracker:
        def snapshot(self, *, elapsed_sec=None, force=False):
            return {
                "memory_used_gib": 2.5,
                "memory_limit_gib": 4.0,
                "memory_remaining_gib": 1.5,
                "memory_used_pct": 62.5,
                "peak_memory_used_gib": 3.0,
                "memory_history": [
                    {
                        "timestamp": "2026-05-09T08:24:11Z",
                        "elapsed_sec": 1.0 if elapsed_sec is None else elapsed_sec,
                        "memory_used_gib": 2.5,
                        "memory_limit_gib": 4.0,
                        "memory_remaining_gib": 1.5,
                        "memory_used_pct": 62.5,
                    }
                ],
            }

    def _response(run_id, *_args, **kwargs):
        metadata = kwargs["run_metadata"]
        return {
            "status": "completed",
            "run_id": run_id,
            "model_version": "9.1.0",
            "mode": "classical_only",
            **metadata,
            "diagnostics": dict(metadata),
        }

    monkeypatch.setattr(job_worker_module, "MemoryTracker", FakeMemoryTracker)
    monkeypatch.setattr(job_worker_module, "get_run_ledger", lambda: ledger)
    monkeypatch.setattr(job_worker_module, "validate_required_input_sheets", lambda _path: None)
    monkeypatch.setattr(job_worker_module, "workbook_structure", lambda _path: {})
    monkeypatch.setattr(job_worker_module, "build_qubo_from_workbook", lambda _path, _log, _settings=None: optimizer)
    monkeypatch.setattr(job_worker_module, "validate_problem_policy", lambda *_args, **_kwargs: _policy_result())
    monkeypatch.setattr(job_worker_module, "run_classical_optimizer", lambda current, logs: (current, logs))
    monkeypatch.setattr(job_worker_module, "build_classical_response", _response)

    job_worker_module.run_job(job_id)
    job = get_job_store().get_job(job_id)
    result = app.test_client().get(f"/jobs/{job_id}/result").get_json()

    assert job["worker_profile"] == "medium"
    assert job["configured_cpu"] == 4
    assert job["configured_memory_gib"] == 4
    assert job["memory_used_gib"] == 2.5
    assert job["peak_memory_used_gib"] == 3.0
    assert result["worker_profile"] == "medium"
    assert result["worker_profile_label"] == "Medium"
    assert result["configured_memory_gib"] == 4
    assert result["memory_history"][0]["memory_used_gib"] == 2.5


def test_job_status_returns_memory_history(monkeypatch, tmp_path):
    job_id = _create_local_worker_job(tmp_path, monkeypatch, job_id="job_test_status_memory")
    get_job_store().update_job(
        job_id,
        {
            "memory_used_gib": 1.25,
            "memory_limit_gib": 4.0,
            "memory_remaining_gib": 2.75,
            "memory_used_pct": 31.25,
            "peak_memory_used_gib": 1.5,
            "memory_history": [
                {
                    "timestamp": "2026-05-09T08:24:11Z",
                    "elapsed_sec": 12.4,
                    "memory_used_gib": 1.25,
                    "memory_limit_gib": 4.0,
                    "memory_remaining_gib": 2.75,
                    "memory_used_pct": 31.25,
                }
            ],
        },
    )

    payload = app.test_client().get(f"/jobs/{job_id}/status").get_json()

    assert payload["memory_used_gib"] == 1.25
    assert payload["memory_limit_gib"] == 4.0
    assert payload["memory_remaining_gib"] == 2.75
    assert payload["memory_used_pct"] == 31.25
    assert payload["peak_memory_used_gib"] == 1.5
    assert payload["memory_history"][0]["elapsed_sec"] == 12.4


def test_memory_snapshot_handles_missing_cgroup_files(tmp_path):
    snapshot = memory_monitor_module.get_memory_snapshot(cgroup_root=tmp_path)

    assert snapshot["memory_used_gib"] is None
    assert snapshot["memory_limit_gib"] is None
    assert snapshot["memory_remaining_gib"] is None
    assert snapshot["memory_used_pct"] is None
    assert snapshot["peak_memory_used_gib"] is None


def test_memory_snapshot_reads_cgroup_v2_values(tmp_path):
    (tmp_path / "memory.current").write_text(str(2 * 1024**3), encoding="utf-8")
    (tmp_path / "memory.max").write_text(str(8 * 1024**3), encoding="utf-8")

    snapshot = memory_monitor_module.get_memory_snapshot(cgroup_root=tmp_path)

    assert snapshot["memory_used_gib"] == pytest.approx(2.0)
    assert snapshot["memory_limit_gib"] == pytest.approx(8.0)
    assert snapshot["memory_remaining_gib"] == pytest.approx(6.0)
    assert snapshot["memory_used_pct"] == pytest.approx(25.0)
    assert snapshot["peak_memory_used_gib"] == pytest.approx(2.0)


def test_memory_snapshot_handles_unbounded_cgroup_v2_limit(tmp_path):
    (tmp_path / "memory.current").write_text(str(1024**3), encoding="utf-8")
    (tmp_path / "memory.max").write_text("max", encoding="utf-8")

    snapshot = memory_monitor_module.get_memory_snapshot(cgroup_root=tmp_path)

    assert snapshot["memory_used_gib"] == pytest.approx(1.0)
    assert snapshot["memory_limit_gib"] is None
    assert snapshot["memory_remaining_gib"] is None
    assert snapshot["memory_used_pct"] is None


def test_memory_tracker_persists_history_and_tracks_peak(tmp_path):
    (tmp_path / "memory.current").write_text(str(1024**3), encoding="utf-8")
    (tmp_path / "memory.max").write_text(str(4 * 1024**3), encoding="utf-8")
    tracker = memory_monitor_module.MemoryTracker(cgroup_root=tmp_path, sample_interval_sec=0.0)

    first = tracker.snapshot(elapsed_sec=1.0, force=True)
    (tmp_path / "memory.current").write_text(str(3 * 1024**3), encoding="utf-8")
    second = tracker.snapshot(elapsed_sec=2.0, force=True)
    (tmp_path / "memory.current").write_text(str(2 * 1024**3), encoding="utf-8")
    third = tracker.snapshot(elapsed_sec=3.0, force=True)

    assert first["memory_used_gib"] == pytest.approx(1.0)
    assert second["memory_used_gib"] == pytest.approx(3.0)
    assert third["memory_used_gib"] == pytest.approx(2.0)
    assert third["peak_memory_used_gib"] == pytest.approx(3.0)
    assert [point["memory_used_gib"] for point in third["memory_history"]] == [1.0, 3.0, 2.0]


def test_worker_preserves_random_seed_in_final_result(monkeypatch, tmp_path):
    settings = {"mode": "classical_only", "response_level": "full", "random_seed": "321", "rng_seed": "321"}
    job_id = _create_local_worker_job(tmp_path, monkeypatch, job_id="job_test_worker_seed", settings=settings)
    ledger = _RouteLockLedger()
    optimizer = SimpleNamespace(
        n=3,
        qaoa_p=1,
        qaoa_maxiter=10,
        qaoa_multistart_restarts=1,
        qaoa_layerwise_warm_start=False,
        qaoa_shots=None,
        classical_results=[{"bitstring": "101"}],
        progress_callback=lambda _message, _progress=None: None,
    )

    def _seeded_policy_result(*_args, **_kwargs):
        base = _policy_result()
        base.runtime_inputs.random_seed = 321
        base.effective_settings["random_seed"] = 321
        base.effective_settings["rng_seed"] = 321
        return base

    def _seeded_response(run_id, *_args, **kwargs):
        policy = kwargs["policy_result"]
        return {
            "status": "completed",
            "run_id": run_id,
            "model_version": "9.1.0",
            "mode": "classical_only",
            "diagnostics": {
                "random_seed": policy.runtime_inputs.random_seed,
                "effective_settings": policy.effective_settings,
            },
        }

    monkeypatch.setattr(job_worker_module, "get_run_ledger", lambda: ledger)
    monkeypatch.setattr(job_worker_module, "validate_required_input_sheets", lambda _path: None)
    monkeypatch.setattr(job_worker_module, "workbook_structure", lambda _path: {})
    monkeypatch.setattr(job_worker_module, "build_qubo_from_workbook", lambda _path, _log, _settings=None: optimizer)
    monkeypatch.setattr(job_worker_module, "validate_problem_policy", _seeded_policy_result)
    monkeypatch.setattr(job_worker_module, "run_classical_optimizer", lambda current, logs: (current, logs))
    monkeypatch.setattr(job_worker_module, "build_classical_response", _seeded_response)

    job_worker_module.run_job(job_id)
    result = app.test_client().get(f"/jobs/{job_id}/result").get_json()
    job = get_job_store().get_job(job_id)

    assert result["diagnostics"]["random_seed"] == 321
    assert result["diagnostics"]["effective_settings"]["random_seed"] == 321
    assert any("Random seed: 321" in line for line in job["logs_tail"])


def test_worker_marks_failed_and_releases_public_lock(monkeypatch, tmp_path):
    job_id = _create_local_worker_job(tmp_path, monkeypatch, job_id="job_test_worker_failed")
    ledger = _RouteLockLedger()
    monkeypatch.setattr(job_worker_module, "get_run_ledger", lambda: ledger)
    monkeypatch.setattr(job_worker_module, "validate_required_input_sheets", lambda _path: None)
    monkeypatch.setattr(job_worker_module, "workbook_structure", lambda _path: {})

    def fail_build(*_args, **_kwargs):
        raise RuntimeError("simulated worker failure")

    monkeypatch.setattr(job_worker_module, "build_qubo_from_workbook", fail_build)

    job_worker_module.run_job(job_id)
    job = get_job_store().get_job(job_id)

    assert job["status"] == "failed"
    assert job["error"]["type"] == "RuntimeError"
    assert "simulated worker failure" in job["error"]["message"]
    assert ledger.rejected is True
    assert ledger.public_released is True


def test_worker_cancelled_job_releases_public_lock(monkeypatch, tmp_path):
    job_id = _create_local_worker_job(tmp_path, monkeypatch, job_id="job_test_worker_cancelled")
    get_job_store().update_job(job_id, {"cancel_requested": True})
    ledger = _RouteLockLedger()
    monkeypatch.setattr(job_worker_module, "get_run_ledger", lambda: ledger)

    job_worker_module.run_job(job_id)
    job = get_job_store().get_job(job_id)

    assert job["status"] == "cancelled"
    assert job["phase"] == "cancelled"
    assert job["result"]["available"] is False
    assert ledger.public_released is True


def test_worker_releases_key_lock_when_usage_context_rebuild_fails(monkeypatch, tmp_path):
    monkeypatch.setattr(main_module.Config, "LOCAL_JOB_DIR", tmp_path / "jobs")
    usage_level = load_usage_config()["usage_levels"]["internal_power"]
    usage_context = UsageContext(
        usage_level_name="internal_power",
        usage_level=usage_level,
        key_record={"key_id": "missing-key"},
        authenticated=True,
    )
    job_id = "job_test_missing_key_release"
    get_job_store().create_job(
        initial_job_document(
            job_id=job_id,
            mode="classical_only",
            response_level="full",
            settings={"mode": "classical_only", "response_level": "full"},
            input_info={"storage_path": "unused.xlsx", "original_filename": "unused.xlsx", "storage_mode": "local"},
            usage_context=usage_context,
            policy_result=_policy_result(),
            lock_type="key",
            key_hash="unit-test-key-hash",
        )
    )
    ledger = _RouteLockLedger()
    monkeypatch.setattr(job_worker_module, "get_run_ledger", lambda: ledger)
    monkeypatch.setattr(job_worker_module, "get_key_store", lambda: SimpleNamespace(find_key_by_id=lambda _key_id: None))

    job_worker_module.run_job(job_id)
    job = get_job_store().get_job(job_id)

    assert job["status"] == "failed"
    assert job["error"]["type"] == "ApiError"
    assert ledger.released is True


def test_license_status_no_key_returns_public_demo():
    response = app.test_client().get("/license-status")
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["authenticated"] is False
    assert payload["key_id"] == "anonymous"
    assert payload["usage_level"] == "public_demo"
    assert payload["limits"]["max_qubits"] == 8
    assert payload["allowed_worker_profiles"] == ["small"]
    assert payload["worker_profiles"]["medium"]["enabled"] is False


def test_license_status_demo_key_returns_remaining_runs():
    response = app.test_client().get("/license-status", headers={"X-API-Key": "DEMO-123"})
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["authenticated"] is True
    assert payload["key_id"] == DEMO_KEY_ID
    assert payload["usage_level"] == "qualified_demo"
    assert payload["max_runs"] == 1000
    assert payload["used_runs"] == 0
    assert payload["remaining_runs"] == 1000
    assert payload["allowed_worker_profiles"] == ["small"]


def test_license_status_internal_power_shows_qaoa_sim_effective_limits():
    response = app.test_client().get("/license-status", headers={"X-API-Key": INTERNAL_POWER_KEY})
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["authenticated"] is True
    assert payload["key_id"] == INTERNAL_POWER_KEY_ID
    assert payload["usage_level"] == "internal_power"
    assert payload["display_name"] == "Internal Power"
    assert payload["general_limits"]["max_qubits"] == 24
    assert payload["general_limits"]["max_estimated_runtime_sec"] == 7200
    assert payload["qaoa_lightning_sim_limits"]["max_qubits"] == 24
    assert payload["qaoa_lightning_sim_limits"]["max_layers"] == 8
    assert payload["qaoa_lightning_sim_limits"]["max_iterations"] == 300
    assert payload["qaoa_lightning_sim_limits"]["max_restarts"] == 5
    assert payload["qaoa_lightning_sim_limits"]["max_estimated_runtime_sec"] == 7200
    assert payload["limits"]["qaoa_lightning_sim"]["max_qubits"] == 24
    assert payload["limits"]["qaoa_tensor_sim"]["max_qubits"] == 24
    assert payload["allowed_worker_profiles"] == ["small", "medium", "large"]
    assert payload["worker_profiles"]["large"]["enabled"] is True


def test_license_status_invalid_key_rejected():
    response = app.test_client().get("/license-status", headers={"X-API-Key": "NOT-A-REAL-KEY"})
    payload = response.get_json()

    assert response.status_code == 401
    assert payload["error"]["code"] == "invalid_api_key"


def test_firestore_key_store_matches_hmac_hash_without_exposing_hash():
    client = _FakeFirestoreClient()
    key_hash = generate_key_hash("FIRESTORE-RAW-KEY", secret="unit-test-secret")
    client.data["qaoa_keys"]["firestore-demo-001"] = {
        "key_id": "firestore-demo-001",
        "key_hash": key_hash,
        "usage_level": "qualified_demo",
        "status": "active",
        "max_runs": 3,
        "used_runs": 0,
        "remaining_runs": 3,
        "expires_at": "2099-01-01T00:00:00Z",
    }

    record = FirestoreApiKeyStore(client).find_key_by_hash(key_hash)
    missing = FirestoreApiKeyStore(client).find_key_by_hash(generate_key_hash("WRONG", secret="unit-test-secret"))

    assert record["key_id"] == "firestore-demo-001"
    assert record["level"] == "qualified_demo"
    assert record["usage_level"] == "qualified_demo"
    assert record["_key_store"] == "firestore"
    assert record["_firestore_doc_id"] == "firestore-demo-001"
    assert "key_hash" not in record
    assert missing is None


def test_firestore_ledger_consumes_success_atomically_and_records_run():
    client = _FakeFirestoreClient()
    client.data["qaoa_keys"]["firestore-demo-001"] = {
        "key_id": "firestore-demo-001",
        "usage_level": "qualified_demo",
        "status": "active",
        "max_runs": 2,
        "used_runs": 0,
        "remaining_runs": 2,
        "expires_at": "2099-01-01T00:00:00Z",
    }
    key_record = {
        **client.data["qaoa_keys"]["firestore-demo-001"],
        "level": "qualified_demo",
        "_firestore_doc_id": "firestore-demo-001",
    }
    ledger = FirestoreRunLedger(client=client, transactional=_identity_transactional)

    assert ledger.can_consume_run(key_record) is True
    assert ledger.consume_run(key_record, "run-firestore-001") is True
    ledger.record_run_completed(
        run_id="run-firestore-001",
        usage_context=_firestore_usage_context(key_record),
        mode="classical_only",
        response_level="compact",
        filename="input.xlsx",
        optimizer=None,
        policy_result=_policy_result(),
        timestamp_start_utc="2026-01-01T00:00:00Z",
        actual_runtime_sec=1.0,
        solver="classical_heuristic",
        consumed_run=True,
    )

    assert client.data["qaoa_keys"]["firestore-demo-001"]["used_runs"] == 1
    assert client.data["qaoa_keys"]["firestore-demo-001"]["remaining_runs"] == 1
    run_record = client.data["qaoa_usage_events"]["run-firestore-001"]
    assert run_record["status"] == "completed"
    assert run_record["consumed_run"] is True
    assert run_record["key_id"] == "firestore-demo-001"
    assert run_record["n_qubits"] == 3
    assert run_record["candidate_count"] == 8
    assert run_record["runtime_ratio"] == 0.5
    assert run_record["service_version"] == "9.1.0"
    assert "timestamp" in run_record


def test_firestore_ledger_rejected_run_does_not_increment_usage():
    client = _FakeFirestoreClient()
    client.data["qaoa_keys"]["firestore-demo-001"] = {
        "key_id": "firestore-demo-001",
        "usage_level": "qualified_demo",
        "status": "active",
        "max_runs": 2,
        "used_runs": 0,
        "remaining_runs": 2,
        "expires_at": "2099-01-01T00:00:00Z",
    }
    key_record = {
        **client.data["qaoa_keys"]["firestore-demo-001"],
        "level": "qualified_demo",
        "_firestore_doc_id": "firestore-demo-001",
    }
    ledger = FirestoreRunLedger(client=client, transactional=_identity_transactional)

    ledger.record_run_rejected(
        run_id="run-firestore-rejected",
        usage_context=_firestore_usage_context(key_record),
        mode="classical_only",
        response_level="full",
        filename="input.xlsx",
        optimizer=None,
        policy_result=_policy_result(),
        timestamp_start_utc="2026-01-01T00:00:00Z",
        actual_runtime_sec=0.1,
        rejection_code="response_level_not_allowed",
        consumed_run=False,
    )

    assert client.data["qaoa_keys"]["firestore-demo-001"]["used_runs"] == 0
    assert client.data["qaoa_keys"]["firestore-demo-001"]["remaining_runs"] == 2
    run_record = client.data["qaoa_usage_events"]["run-firestore-rejected"]
    assert run_record["status"] == "rejected"
    assert run_record["consumed_run"] is False
    assert run_record["error_code"] == "response_level_not_allowed"


def test_firestore_ledger_exhausted_key_rejected_before_consumption():
    client = _FakeFirestoreClient()
    client.data["qaoa_keys"]["firestore-demo-001"] = {
        "key_id": "firestore-demo-001",
        "usage_level": "qualified_demo",
        "status": "active",
        "max_runs": 1,
        "used_runs": 1,
        "remaining_runs": 0,
        "expires_at": "2099-01-01T00:00:00Z",
    }
    key_record = {
        **client.data["qaoa_keys"]["firestore-demo-001"],
        "level": "qualified_demo",
        "_firestore_doc_id": "firestore-demo-001",
    }
    ledger = FirestoreRunLedger(client=client, transactional=_identity_transactional)

    assert ledger.can_consume_run(key_record) is False
    assert ledger.consume_run(key_record, "run-firestore-exhausted") is False
    assert client.data["qaoa_keys"]["firestore-demo-001"]["used_runs"] == 1


def test_firestore_run_lock_first_run_acquires_and_releases():
    client = _FakeFirestoreClient()
    client.data["qaoa_keys"]["firestore-demo-001"] = {
        "key_id": "firestore-demo-001",
        "usage_level": "qualified_demo",
        "status": "active",
        "max_runs": 2,
        "used_runs": 0,
        "remaining_runs": 2,
        "expires_at": "2099-01-01T00:00:00Z",
    }
    key_record = {
        **client.data["qaoa_keys"]["firestore-demo-001"],
        "level": "qualified_demo",
        "_firestore_doc_id": "firestore-demo-001",
    }
    ledger = FirestoreRunLedger(client=client, transactional=_identity_transactional)

    lock_info = ledger.acquire_run_lock(
        key_record,
        "run-lock-001",
        policy_result=_policy_result(),
        user_agent="unit-test",
        ip_address="127.0.0.1",
    )

    stored = client.data["qaoa_keys"]["firestore-demo-001"]
    assert lock_info["acquired"] is True
    assert lock_info["stale_lock_cleared"] is False
    assert stored["active_run_id"] == "run-lock-001"
    assert stored["active_run_status"] == "running"
    assert stored["active_run_user_agent"] == "unit-test"
    assert stored["active_run_ip"] == "127.0.0.1"
    assert stored["max_parallel_runs"] == 1

    assert ledger.release_run_lock(key_record, "run-lock-001") is True
    assert stored["active_run_id"] is None
    assert stored["active_run_started_at"] is None
    assert stored["active_run_status"] is None


def test_firestore_run_lock_second_same_key_rejected_and_not_consumed():
    client = _FakeFirestoreClient()
    client.data["qaoa_keys"]["firestore-demo-001"] = {
        "key_id": "firestore-demo-001",
        "usage_level": "qualified_demo",
        "status": "active",
        "max_runs": 2,
        "used_runs": 0,
        "remaining_runs": 2,
        "expires_at": "2099-01-01T00:00:00Z",
        "active_run_id": "run-existing",
        "active_run_started_at": dt.datetime.now(dt.timezone.utc),
        "active_run_status": "running",
    }
    key_record = {
        **client.data["qaoa_keys"]["firestore-demo-001"],
        "level": "qualified_demo",
        "_firestore_doc_id": "firestore-demo-001",
    }
    ledger = FirestoreRunLedger(client=client, transactional=_identity_transactional)

    with pytest.raises(ApiError) as exc_info:
        ledger.acquire_run_lock(key_record, "run-second", policy_result=_policy_result())

    assert exc_info.value.status_code == 409
    assert exc_info.value.code == "active_run_exists"
    assert client.data["qaoa_keys"]["firestore-demo-001"]["used_runs"] == 0
    assert client.data["qaoa_keys"]["firestore-demo-001"]["remaining_runs"] == 2
    assert client.data["qaoa_keys"]["firestore-demo-001"]["active_run_id"] == "run-existing"

    ledger.record_run_rejected(
        run_id="run-second",
        usage_context=_firestore_usage_context(key_record),
        mode="classical_only",
        response_level="compact",
        filename="input.xlsx",
        optimizer=None,
        policy_result=_policy_result(),
        timestamp_start_utc="2026-01-01T00:00:00Z",
        actual_runtime_sec=0.0,
        rejection_code="active_run_exists",
        consumed_run=False,
    )
    run_record = client.data["qaoa_usage_events"]["run-second"]
    assert run_record["status"] == "rejected"
    assert run_record["consumed_run"] is False
    assert run_record["error_code"] == "active_run_exists"
    assert client.data["qaoa_keys"]["firestore-demo-001"]["used_runs"] == 0


def test_firestore_run_lock_different_keys_can_run_independently():
    client = _FakeFirestoreClient()
    for key_id in ("firestore-demo-001", "firestore-demo-002"):
        client.data["qaoa_keys"][key_id] = {
            "key_id": key_id,
            "usage_level": "qualified_demo",
            "status": "active",
            "max_runs": 2,
            "used_runs": 0,
            "remaining_runs": 2,
            "expires_at": "2099-01-01T00:00:00Z",
        }
    ledger = FirestoreRunLedger(client=client, transactional=_identity_transactional)
    first_record = {**client.data["qaoa_keys"]["firestore-demo-001"], "level": "qualified_demo", "_firestore_doc_id": "firestore-demo-001"}
    second_record = {**client.data["qaoa_keys"]["firestore-demo-002"], "level": "qualified_demo", "_firestore_doc_id": "firestore-demo-002"}

    first = ledger.acquire_run_lock(first_record, "run-first", policy_result=_policy_result())
    second = ledger.acquire_run_lock(second_record, "run-second", policy_result=_policy_result())

    assert first["acquired"] is True
    assert second["acquired"] is True
    assert client.data["qaoa_keys"]["firestore-demo-001"]["active_run_id"] == "run-first"
    assert client.data["qaoa_keys"]["firestore-demo-002"]["active_run_id"] == "run-second"


def test_firestore_run_lock_allows_same_key_until_max_parallel_runs():
    client = _FakeFirestoreClient()
    client.data["qaoa_keys"]["firestore-demo-001"] = {
        "key_id": "firestore-demo-001",
        "usage_level": "qualified_demo",
        "status": "active",
        "max_runs": 10,
        "used_runs": 0,
        "remaining_runs": 10,
        "max_parallel_runs": 3,
        "expires_at": "2099-01-01T00:00:00Z",
    }
    key_record = {
        **client.data["qaoa_keys"]["firestore-demo-001"],
        "level": "qualified_demo",
        "_firestore_doc_id": "firestore-demo-001",
    }
    ledger = FirestoreRunLedger(client=client, transactional=_identity_transactional)

    for index in range(1, 4):
        lock_info = ledger.acquire_run_lock(key_record, f"run-{index}", policy_result=_policy_result())
        assert lock_info["acquired"] is True
        assert lock_info["active_run_count"] == index

    assert client.data["qaoa_key_run_state_v9"]["firestore-demo-001"]["active_count"] == 3
    assert client.data["qaoa_keys"]["firestore-demo-001"]["active_run_count"] == 3
    assert set(client.data["qaoa_keys"]["firestore-demo-001"]["active_run_ids"]) == {"run-1", "run-2", "run-3"}

    with pytest.raises(ApiError) as exc_info:
        ledger.acquire_run_lock(key_record, "run-4", policy_result=_policy_result())

    assert exc_info.value.status_code == 409
    assert exc_info.value.code == "active_run_exists"
    assert exc_info.value.details["active_run_count"] == 3
    assert exc_info.value.details["max_parallel_runs"] == 3
    assert set(exc_info.value.details["active_run_ids"]) == {"run-1", "run-2", "run-3"}


def test_firestore_run_lock_release_one_parallel_run_leaves_others_active():
    client = _FakeFirestoreClient()
    client.data["qaoa_keys"]["firestore-demo-001"] = {
        "key_id": "firestore-demo-001",
        "usage_level": "qualified_demo",
        "status": "active",
        "max_runs": 10,
        "used_runs": 0,
        "remaining_runs": 10,
        "max_parallel_runs": 3,
        "expires_at": "2099-01-01T00:00:00Z",
    }
    key_record = {
        **client.data["qaoa_keys"]["firestore-demo-001"],
        "level": "qualified_demo",
        "_firestore_doc_id": "firestore-demo-001",
    }
    ledger = FirestoreRunLedger(client=client, transactional=_identity_transactional)
    ledger.acquire_run_lock(key_record, "run-1", policy_result=_policy_result())
    ledger.acquire_run_lock(key_record, "run-2", policy_result=_policy_result())

    assert ledger.release_run_lock(key_record, "run-1") is True

    assert client.data["qaoa_key_run_locks_v9"]["firestore-demo-001__run-1"]["status"] == "released"
    assert client.data["qaoa_key_run_locks_v9"]["firestore-demo-001__run-2"]["status"] == "running"
    assert client.data["qaoa_key_run_state_v9"]["firestore-demo-001"]["active_count"] == 1
    assert client.data["qaoa_keys"]["firestore-demo-001"]["active_run_count"] == 1
    assert client.data["qaoa_keys"]["firestore-demo-001"]["active_run_ids"] == ["run-2"]


def test_firestore_run_lock_stale_lock_is_cleared_and_overwritten():
    client = _FakeFirestoreClient()
    old_started_at = dt.datetime.now(dt.timezone.utc) - dt.timedelta(hours=3)
    client.data["qaoa_keys"]["firestore-demo-001"] = {
        "key_id": "firestore-demo-001",
        "usage_level": "qualified_demo",
        "status": "active",
        "max_runs": 2,
        "used_runs": 0,
        "remaining_runs": 2,
        "expires_at": "2099-01-01T00:00:00Z",
        "active_run_id": "run-stale",
        "active_run_started_at": old_started_at,
        "active_run_status": "running",
    }
    key_record = {
        **client.data["qaoa_keys"]["firestore-demo-001"],
        "level": "qualified_demo",
        "_firestore_doc_id": "firestore-demo-001",
    }
    ledger = FirestoreRunLedger(client=client, transactional=_identity_transactional)

    lock_info = ledger.acquire_run_lock(key_record, "run-new", policy_result=_policy_result())

    assert lock_info["acquired"] is True
    assert lock_info["stale_lock_cleared"] is True
    assert lock_info["previous_active_run_id"] == "run-stale"
    assert client.data["qaoa_keys"]["firestore-demo-001"]["active_run_id"] == "run-new"


def test_firestore_run_lock_clears_finished_job_lock_before_capacity_check():
    client = _FakeFirestoreClient()
    client.data["qaoa_keys"]["firestore-demo-001"] = {
        "key_id": "firestore-demo-001",
        "usage_level": "qualified_demo",
        "status": "active",
        "max_runs": 10,
        "used_runs": 0,
        "remaining_runs": 10,
        "max_parallel_runs": 1,
        "expires_at": "2099-01-01T00:00:00Z",
    }
    started_at = dt.datetime.now(dt.timezone.utc)
    client.data["qaoa_key_run_state_v9"]["firestore-demo-001"] = {"active_count": 1, "max_parallel_runs": 1}
    client.data["qaoa_key_run_locks_v9"]["firestore-demo-001__run-finished"] = {
        "run_id": "run-finished",
        "key_id": "firestore-demo-001",
        "key_doc_id": "firestore-demo-001",
        "status": "running",
        "started_at": started_at,
    }
    client.data["qaoa_jobs_v9"]["run-finished"] = {
        "job_id": "run-finished",
        "status": "failed",
        "heartbeat_at": started_at,
    }
    key_record = {
        **client.data["qaoa_keys"]["firestore-demo-001"],
        "level": "qualified_demo",
        "_firestore_doc_id": "firestore-demo-001",
    }
    ledger = FirestoreRunLedger(client=client, transactional=_identity_transactional)

    lock_info = ledger.acquire_run_lock(key_record, "run-new", policy_result=_policy_result())

    assert lock_info["acquired"] is True
    assert client.data["qaoa_key_run_locks_v9"]["firestore-demo-001__run-finished"]["status"] == "stale_released"
    assert client.data["qaoa_key_run_locks_v9"]["firestore-demo-001__run-new"]["status"] == "running"
    assert client.data["qaoa_key_run_state_v9"]["firestore-demo-001"]["active_count"] == 1


def test_run_qaoa_releases_lock_in_finally_on_success(monkeypatch):
    ledger = _RouteLockLedger()
    _patch_fast_run(monkeypatch, ledger)

    response = _post_run(headers={"X-API-Key": "DEMO-123"}, mode="classical_only")
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["status"] == "completed"
    assert ledger.acquired is True
    assert ledger.public_acquired is False
    assert ledger.completed is True
    assert ledger.released is True
    assert ledger.consumed is True


def test_run_qaoa_releases_lock_in_finally_on_error(monkeypatch):
    ledger = _RouteLockLedger()
    _patch_fast_run(monkeypatch, ledger, fail_execution=True)

    response = _post_run(headers={"X-API-Key": "DEMO-123"}, mode="classical_only")
    payload = response.get_json()

    assert response.status_code == 500
    assert payload["error"]["code"] == "internal_server_error"
    assert ledger.acquired is True
    assert ledger.failed is True
    assert ledger.released is True
    assert ledger.consumed is False


def test_public_run_acquires_and_releases_public_slot(monkeypatch):
    ledger = _RouteLockLedger()
    _patch_fast_run(monkeypatch, ledger)

    response = _post_run(mode="classical_only")
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["status"] == "completed"
    assert ledger.public_acquired is True
    assert ledger.public_released is True
    assert ledger.acquired is False
    assert ledger.completed is True
    assert ledger.consumed is False


def test_public_capacity_rejection_returns_429_and_does_not_start(monkeypatch):
    error = ApiError(
        429,
        "public_demo_capacity_exceeded",
        "The public demo is currently busy. Please try again in a few minutes.",
        {
            "usage_level": "public_demo",
            "active_public_runs": 5,
            "max_parallel_public_runs": 5,
        },
    )
    ledger = _RouteLockLedger(fail_public_slot=error)
    _patch_fast_run(monkeypatch, ledger)

    response = _post_run(mode="classical_only")
    payload = response.get_json()

    assert response.status_code == 429
    assert payload["status"] == "error"
    assert payload["error"]["code"] == "public_demo_capacity_exceeded"
    assert payload["error"]["message"] == "The public demo is currently busy. Please try again in a few minutes."
    assert payload["error"]["details"]["active_public_runs"] == 5
    assert payload["error"]["details"]["max_parallel_public_runs"] == 5
    assert payload["license"]["usage_level"] == "public_demo"
    assert ledger.rejected is True
    assert ledger.completed is False
    assert ledger.consumed is False
    assert ledger.public_released is False


def test_inspect_workbook_does_not_acquire_public_slot(monkeypatch):
    ledger = _RouteLockLedger(
        fail_public_slot=ApiError(429, "public_demo_capacity_exceeded", "should not be used")
    )
    _patch_fast_run(monkeypatch, ledger)

    response = _post_inspect(mode="classical_only")
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["status"] == "completed"
    assert ledger.public_acquired is False
    assert ledger.public_released is False


def test_local_public_stale_slot_is_cleared_before_acquire():
    _LOCAL_PUBLIC_RUNS.clear()
    old_started_at = dt.datetime.now(dt.timezone.utc) - dt.timedelta(hours=3)
    _LOCAL_PUBLIC_RUNS["stale-public-run"] = {
        "run_id": "stale-public-run",
        "status": "running",
        "started_at": old_started_at,
    }
    context = SimpleNamespace(
        authenticated=False,
        usage_level_name="public_demo",
        usage_level={"max_parallel_runs": 1},
    )

    lock_info = RunLedger().acquire_public_run_slot(context, "fresh-public-run", policy_result=_policy_result())

    assert lock_info["acquired"] is True
    assert lock_info["active_public_runs"] == 1
    assert lock_info["max_parallel_public_runs"] == 1
    assert lock_info["stale_locks_cleared"] == 1
    assert "stale-public-run" not in _LOCAL_PUBLIC_RUNS
    assert "fresh-public-run" in _LOCAL_PUBLIC_RUNS
    assert RunLedger().release_public_run_slot("fresh-public-run") is True
    _LOCAL_PUBLIC_RUNS.clear()


def test_firestore_public_semaphore_acquire_reject_and_release():
    client = _FakeFirestoreClient()
    ledger = FirestoreRunLedger(client=client, transactional=_identity_transactional)
    context = SimpleNamespace(
        authenticated=False,
        usage_level_name="public_demo",
        usage_level={"max_parallel_runs": 1},
    )

    first = ledger.acquire_public_run_slot(
        context,
        "public-run-001",
        policy_result=_policy_result(),
        mode="classical_only",
        response_level="full",
        filename="input.xlsx",
        user_agent="pytest-agent",
        ip_address="127.0.0.1",
    )

    assert first["acquired"] is True
    state = client.data["qaoa_public_run_state_v9"]["global"]
    lock = client.data["qaoa_public_run_locks_v9"]["public-run-001"]
    assert state["active_count"] == 1
    assert state["max_parallel_runs"] == 1
    assert lock["status"] == "running"
    assert lock["usage_level"] == "public_demo"
    assert lock["mode"] == "classical_only"
    assert lock["filename"] == "input.xlsx"
    assert lock["client_ip_hash"] != "127.0.0.1"
    assert lock["user_agent_hash"] != "pytest-agent"

    with pytest.raises(ApiError) as exc_info:
        ledger.acquire_public_run_slot(context, "public-run-002", policy_result=_policy_result())

    assert exc_info.value.status_code == 429
    assert exc_info.value.code == "public_demo_capacity_exceeded"
    assert exc_info.value.details["active_public_runs"] == 1
    assert exc_info.value.details["max_parallel_public_runs"] == 1

    assert ledger.release_public_run_slot("public-run-001") is True
    assert client.data["qaoa_public_run_state_v9"]["global"]["active_count"] == 0
    assert client.data["qaoa_public_run_locks_v9"]["public-run-001"]["status"] == "released"


def test_firestore_public_stale_lock_is_cleared_before_capacity_check():
    client = _FakeFirestoreClient()
    old_started_at = dt.datetime.now(dt.timezone.utc) - dt.timedelta(hours=3)
    client.data["qaoa_public_run_state_v9"]["global"] = {
        "active_count": 1,
        "max_parallel_runs": 1,
    }
    client.data["qaoa_public_run_locks_v9"]["public-stale"] = {
        "run_id": "public-stale",
        "usage_level": "public_demo",
        "status": "running",
        "started_at": old_started_at,
    }
    ledger = FirestoreRunLedger(client=client, transactional=_identity_transactional)
    context = SimpleNamespace(
        authenticated=False,
        usage_level_name="public_demo",
        usage_level={"max_parallel_runs": 1},
    )

    lock_info = ledger.acquire_public_run_slot(context, "public-fresh", policy_result=_policy_result())

    assert lock_info["acquired"] is True
    assert client.data["qaoa_public_run_locks_v9"]["public-stale"]["status"] == "stale_released"
    assert client.data["qaoa_public_run_locks_v9"]["public-fresh"]["status"] == "running"
    assert client.data["qaoa_public_run_state_v9"]["global"]["active_count"] == 1


def test_key_admin_create_writes_lock_fields_and_single_raw_key(monkeypatch, capsys):
    client = _FakeFirestoreClient()
    monkeypatch.setattr(key_admin_firestore.secrets, "token_urlsafe", lambda _size: "generated-secret")
    args = SimpleNamespace(
        key_id="admin-key-001",
        name="Admin Test",
        email="admin@example.com",
        organization="Qubit Lab",
        usage_level="qualified_demo",
        max_runs=25,
        expires_at="2099-01-01T00:00:00Z",
        display_name="Admin Demo",
        notes="unit test",
        created_by="pytest",
        status="active",
        prefix="qlab",
        max_parallel_runs=3,
        force=False,
    )

    assert key_admin_firestore._create_key(client, args) == 0
    output = capsys.readouterr().out
    stored = client.data["qaoa_keys"]["admin-key-001"]

    assert output.count("raw_api_key: qlab_generated-secret") == 1
    assert "key_hash" in stored
    assert stored["key_hash"] != "qlab_generated-secret"
    assert "qlab_generated-secret" not in json.dumps(stored, default=str)
    assert stored["max_parallel_runs"] == 3
    assert stored["active_run_id"] is None
    assert stored["active_run_ids"] == []
    assert stored["active_run_count"] == 0
    assert stored["active_run_started_at"] is None
    assert stored["active_run_status"] is None
    assert stored["active_run_user_agent"] is None
    assert stored["active_run_ip"] is None


def test_key_admin_show_handles_missing_lock_fields(capsys):
    client = _FakeFirestoreClient()
    client.data["qaoa_keys"]["legacy-key"] = {
        "key_id": "legacy-key",
        "key_hash": "hidden-hash",
        "usage_level": "qualified_demo",
        "status": "active",
        "max_runs": 10,
        "used_runs": 1,
    }

    assert key_admin_firestore._show_key(client, "legacy-key", show_hash=False) == 0
    payload = json.loads(capsys.readouterr().out)

    assert "key_hash" not in payload
    assert payload["max_parallel_runs"] == 1
    assert payload["active_run_id"] == "none"
    assert payload["active_run_ids"] == []
    assert payload["active_run_count"] == 0
    assert payload["active_run_started_at"] == "none"
    assert payload["active_run_status"] == "idle"
    assert payload["active_run_user_agent"] == "none"
    assert payload["active_run_ip"] == "none"


def test_key_admin_list_handles_missing_lock_fields(capsys):
    client = _FakeFirestoreClient()
    client.data["qaoa_keys"]["legacy-key"] = {
        "key_id": "legacy-key",
        "usage_level": "qualified_demo",
        "status": "active",
        "max_runs": 10,
        "used_runs": 1,
    }

    assert key_admin_firestore._list_keys(client, limit=20) == 0
    output = capsys.readouterr().out

    assert "active" in output
    assert "legacy-key" in output
    assert "idle" in output


def test_key_admin_clear_lock_requires_confirm(capsys):
    client = _FakeFirestoreClient()
    client.data["qaoa_keys"]["locked-key"] = {
        "key_id": "locked-key",
        "active_run_id": "run-active",
        "active_run_status": "running",
    }

    assert key_admin_firestore._clear_lock(client, "locked-key", confirm=False) == 2
    output = capsys.readouterr().out

    assert "requires --confirm" in output
    assert client.data["qaoa_keys"]["locked-key"]["active_run_id"] == "run-active"
    assert client.data["qaoa_keys"]["locked-key"]["active_run_status"] == "running"


def test_key_admin_clear_lock_clears_fields(capsys):
    client = _FakeFirestoreClient()
    client.data["qaoa_keys"]["locked-key"] = {
        "key_id": "locked-key",
        "active_run_id": "run-active",
        "active_run_ids": ["run-active"],
        "active_run_count": 1,
        "active_run_started_at": "2026-01-01T00:00:00Z",
        "active_run_status": "running",
        "active_run_user_agent": "pytest",
        "active_run_ip": "127.0.0.1",
    }
    client.data["qaoa_key_run_state_v9"]["locked-key"] = {"active_count": 1}
    client.data["qaoa_key_run_locks_v9"]["locked-key__run-active"] = {
        "run_id": "run-active",
        "key_id": "locked-key",
        "key_doc_id": "locked-key",
        "status": "running",
    }

    assert key_admin_firestore._clear_lock(client, "locked-key", confirm=True) == 0
    output = capsys.readouterr().out
    stored = client.data["qaoa_keys"]["locked-key"]

    assert "active_run_lock: cleared" in output
    assert stored["active_run_id"] is None
    assert stored["active_run_ids"] == []
    assert stored["active_run_count"] == 0
    assert stored["active_run_started_at"] is None
    assert stored["active_run_status"] is None
    assert stored["active_run_user_agent"] is None
    assert stored["active_run_ip"] is None
    assert "cleared_lock_at" in stored
    assert client.data["qaoa_key_run_state_v9"]["locked-key"]["active_count"] == 0
    assert client.data["qaoa_key_run_locks_v9"]["locked-key__run-active"]["status"] == "admin_cleared"


def test_key_admin_usage_uses_field_filter_without_positional_where(monkeypatch, capsys):
    client = _FakeFirestoreClient()
    client.data["qaoa_usage_events"]["run-001"] = {
        "run_id": "run-001",
        "key_id": "admin-key-001",
        "timestamp": "2026-01-01T00:00:00Z",
        "status": "completed",
        "mode": "classical_only",
        "response_level": "compact",
        "binary_variables": 3,
    }

    class _FieldFilter:
        def __init__(self, field_path, op_string, value):
            self.field_path = field_path
            self.op_string = op_string
            self.value = value

    google_module = types.ModuleType("google")
    cloud_module = types.ModuleType("google.cloud")
    firestore_v1_module = types.ModuleType("google.cloud.firestore_v1")
    base_query_module = types.ModuleType("google.cloud.firestore_v1.base_query")
    base_query_module.FieldFilter = _FieldFilter
    monkeypatch.setitem(sys.modules, "google", google_module)
    monkeypatch.setitem(sys.modules, "google.cloud", cloud_module)
    monkeypatch.setitem(sys.modules, "google.cloud.firestore_v1", firestore_v1_module)
    monkeypatch.setitem(sys.modules, "google.cloud.firestore_v1.base_query", base_query_module)

    assert key_admin_firestore._usage(client, "admin-key-001", limit=20) == 0
    output = capsys.readouterr().out

    assert "run-001" in output
    assert client.keyword_where_calls == 1
    assert client.positional_where_calls == 0


def test_key_admin_runs_lists_public_usage_events(capsys):
    client = _FakeFirestoreClient()
    now = dt.datetime.now(dt.timezone.utc)
    client.data["qaoa_usage_events"]["run-public-001"] = {
        "run_id": "run-public-001",
        "key_id": "anonymous",
        "usage_level": "public_demo",
        "timestamp": now,
        "status": "completed",
        "mode": "classical_only",
        "response_level": "full",
        "binary_variables": 4,
        "actual_runtime_sec": 1.2,
    }
    client.data["qaoa_usage_events"]["run-auth-001"] = {
        "run_id": "run-auth-001",
        "key_id": "demo-qualified-001",
        "usage_level": "qualified_demo",
        "timestamp": now,
        "status": "completed",
        "mode": "classical_only",
    }

    assert key_admin_firestore._runs(
        client,
        days=30,
        usage_level="public_demo",
        limit=50,
        scan_limit=100,
    ) == 0
    output = capsys.readouterr().out

    assert "run-public-001" in output
    assert "public_demo" in output
    assert "anonymous" in output
    assert "run-auth-001" not in output


def test_key_admin_usage_summary_counts_public_and_capacity_rejections(capsys):
    client = _FakeFirestoreClient()
    now = dt.datetime.now(dt.timezone.utc)
    client.data["qaoa_usage_events"]["run-public-ok"] = {
        "run_id": "run-public-ok",
        "key_id": "anonymous",
        "usage_level": "public_demo",
        "timestamp": now,
        "status": "completed",
        "consumed_run": False,
    }
    client.data["qaoa_usage_events"]["run-public-capacity"] = {
        "run_id": "run-public-capacity",
        "key_id": "anonymous",
        "usage_level": "public_demo",
        "timestamp": now,
        "status": "rejected",
        "error_code": "public_demo_capacity_exceeded",
        "consumed_run": False,
    }
    client.data["qaoa_usage_events"]["run-auth-ok"] = {
        "run_id": "run-auth-ok",
        "key_id": "demo-qualified-001",
        "usage_level": "qualified_demo",
        "timestamp": now,
        "status": "completed",
        "consumed_run": True,
    }

    assert key_admin_firestore._usage_summary(client, days=30, scan_limit=100) == 0
    payload = json.loads(capsys.readouterr().out)

    assert payload["total_runs"] == 3
    assert payload["completed_runs"] == 2
    assert payload["rejected_runs"] == 1
    assert payload["consumed_runs"] == 1
    assert payload["public_demo"]["total"] == 2
    assert payload["public_demo"]["capacity_rejections"] == 1
    assert payload["authenticated"]["total"] == 1
    assert payload["by_usage_level"]["public_demo"] == 2
    assert payload["by_error_code"]["public_demo_capacity_exceeded"] == 1


def test_key_admin_public_status_shows_state_and_running_locks_without_hashes(capsys):
    client = _FakeFirestoreClient()
    now = dt.datetime.now(dt.timezone.utc)
    client.data["qaoa_public_run_state_v9"]["global"] = {
        "active_count": 2,
        "max_parallel_runs": 5,
        "updated_at": now,
    }
    client.data["qaoa_public_run_locks_v9"]["public-run-001"] = {
        "run_id": "public-run-001",
        "usage_level": "public_demo",
        "status": "running",
        "started_at": now,
        "stale_after_sec": 3600,
        "mode": "qaoa_limited",
        "response_level": "full",
        "filename": "input.xlsx",
        "estimated_runtime_sec": 12.5,
        "n_qubits": 7,
        "client_ip_hash": "hidden-client-hash",
        "user_agent_hash": "hidden-agent-hash",
    }

    assert key_admin_firestore._public_status(client, limit=20) == 0
    payload = json.loads(capsys.readouterr().out)

    assert payload["state_document"] == "qaoa_public_run_state_v9/global"
    assert payload["locks_collection"] == "qaoa_public_run_locks_v9"
    assert payload["active_public_runs"] == 2
    assert payload["max_parallel_runs"] == 5
    assert payload["running_locks"][0]["run_id"] == "public-run-001"
    assert payload["running_locks"][0]["n_qubits"] == 7
    assert "hidden-client-hash" not in json.dumps(payload)
    assert "hidden-agent-hash" not in json.dumps(payload)


def test_production_secret_configuration_requires_key_hash_secret(monkeypatch):
    monkeypatch.delenv("QAOA_RQP_LOCAL_DEV", raising=False)
    monkeypatch.delenv("KEY_HASH_SECRET", raising=False)

    with pytest.raises(RuntimeError, match="KEY_HASH_SECRET is required"):
        validate_secret_configuration()

    monkeypatch.setenv("KEY_HASH_SECRET", "unit-test-secret")
    monkeypatch.setenv("QAOA_JOB_BUCKET", "unit-test-job-bucket")
    validate_secret_configuration()


def test_production_configuration_rejects_local_stores(monkeypatch):
    monkeypatch.delenv("QAOA_RQP_LOCAL_DEV", raising=False)
    monkeypatch.setenv("KEY_HASH_SECRET", "unit-test-secret")
    monkeypatch.setenv("QAOA_RQP_KEY_STORE", "local")

    with pytest.raises(RuntimeError, match="QAOA_KEY_STORE=yaml/local"):
        validate_secret_configuration()

    monkeypatch.setenv("QAOA_RQP_KEY_STORE", "firestore")
    monkeypatch.setenv("QAOA_RQP_LEDGER_STORE", "local")
    with pytest.raises(RuntimeError, match="QAOA_RQP_LEDGER_STORE=local"):
        validate_secret_configuration()


def test_job_storage_mode_uses_gcs_when_bucket_is_configured(monkeypatch):
    monkeypatch.setenv("QAOA_RQP_LOCAL_DEV", "1")
    monkeypatch.delenv("QAOA_JOB_STORAGE", raising=False)
    monkeypatch.setenv("QAOA_JOB_BUCKET", "unit-test-job-bucket")

    assert main_module.Config.job_storage_mode() == "gcs"


def test_job_storage_local_is_local_dev_only(monkeypatch):
    monkeypatch.delenv("QAOA_RQP_LOCAL_DEV", raising=False)
    monkeypatch.setenv("KEY_HASH_SECRET", "unit-test-secret")
    monkeypatch.setenv("QAOA_JOB_STORAGE", "local")

    with pytest.raises(RuntimeError, match="QAOA_JOB_STORAGE=local"):
        validate_secret_configuration()


def test_gcs_job_storage_round_trip(monkeypatch, tmp_path):
    monkeypatch.setenv("QAOA_JOB_BUCKET", "unit-test-job-bucket")
    source = tmp_path / "input.xlsx"
    source.write_bytes(b"test workbook bytes")
    storage = GcsJobStorage(bucket_name="unit-test-job-bucket", client=_FakeGcsClient())

    input_info = storage.save_input_from_path("job_storage_test", source, "input.xlsx")
    downloaded = tmp_path / "downloaded.xlsx"
    storage.download_input_to(input_info["storage_path"], downloaded)
    result_path = storage.write_result_json("job_storage_test", {"ok": True, "value": 7})
    result = storage.read_result_json(result_path)

    assert input_info["storage_mode"] == "gcs"
    assert input_info["storage_path"] == "gs://unit-test-job-bucket/jobs/job_storage_test/input.xlsx"
    assert downloaded.read_bytes() == b"test workbook bytes"
    assert result_path == "gs://unit-test-job-bucket/jobs/job_storage_test/result.json"
    assert result == {"ok": True, "value": 7}


def test_dockerfile_cmd_uses_json_exec_form():
    dockerfile = (VERSION_DIR / "Dockerfile").read_text(encoding="utf-8")
    cmd_lines = [line.strip() for line in dockerfile.splitlines() if line.strip().startswith("CMD ")]

    assert len(cmd_lines) == 1
    command = json.loads(cmd_lines[0].removeprefix("CMD ").strip())
    assert command[:3] == ["gunicorn", "--config", "app/gunicorn_conf.py"]
    assert command[-1] == "app.main:app"


def test_demo_run_with_demo_key_still_works():
    response = app.test_client().post("/demo-run", headers={"X-API-Key": "DEMO-123"})
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["status"] == "completed"
    assert payload["mode"] == "demo"


def test_no_key_gets_public_demo_and_is_rejected_for_sample_workbook_size():
    response = _post_run(mode="classical_only")
    payload = response.get_json()

    assert response.status_code == 403
    assert payload["error"]["code"] == "qubit_limit_exceeded"
    assert payload["error"]["message"] == "Problem exceeds the license-level binary-variable limit."
    assert payload["error"]["details"]["mode"] == "classical_only"
    assert payload["error"]["details"]["usage_level"] == "public_demo"
    assert payload["error"]["details"]["binary_variables"] > 8
    assert payload["license"]["usage_level"] == "public_demo"


def test_public_demo_default_response_level_is_full(tmp_path):
    workbook_path = _limited_workbook(tmp_path, qubits=4)
    response = _post_run_file(workbook_path, mode="classical_only")
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["status"] == "completed"
    assert payload["license"]["usage_level"] == "public_demo"
    assert payload["binary_variables"] == 4
    assert "best_candidate" in payload
    assert "top_candidates" in payload
    assert payload["reporting"]["classical_candidates"]
    assert payload["diagnostics"]["effective_settings"]["shots_mode"] == "disabled"


def test_classical_run_reports_indicative_market_cost_rows(tmp_path):
    limited = _limited_workbook(tmp_path, qubits=4)
    workbook_path = _indicative_cost_workbook(tmp_path, source_workbook=limited)
    response = _post_run_file(workbook_path, mode="classical_only")
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["diagnostics"]["cost_column_used"] == "Indicative Market Cost USD"
    assert payload["selected_blocks"]
    assert "Indicative Market Cost USD" in payload["selected_blocks"][0]


def test_all_usage_levels_accept_all_response_levels():
    usage_levels = load_usage_config()["usage_levels"]

    for usage_level in usage_levels.values():
        assert usage_level["allowed_response_levels"] == ["compact", "standard", "full"]


def test_inspect_workbook_demo_key_returns_summary_without_consuming_run(monkeypatch, tmp_path):
    _enable_temp_ledger(monkeypatch, tmp_path)
    before = app.test_client().get("/license-status", headers={"X-API-Key": "DEMO-123"}).get_json()
    response = _post_inspect(
        headers={"X-API-Key": "DEMO-123"},
        mode="classical_only",
        layers="2",
        iterations="20",
        restarts="1",
        warm_start="true",
    )
    payload = response.get_json()
    after = app.test_client().get("/license-status", headers={"X-API-Key": "DEMO-123"}).get_json()
    summary = app.test_client().get("/ledger-summary").get_json()

    assert response.status_code == 200
    assert payload["status"] == "completed"
    assert payload["filename"] == SAMPLE_WORKBOOK.name
    assert payload["license"]["usage_level"] == "qualified_demo"
    assert payload["workbook_summary"]["n_qubits"] == 14
    assert payload["workbook_summary"]["decision_variables"] == 14
    assert payload["workbook_summary"]["decision_state_space"] == "2^14 = 16,384 bitstrings"
    assert payload["workbook_summary"]["currency_code"] == "USD"
    assert payload["workbook_summary"]["qubo_shape"] == [14, 14]
    assert payload["runtime_estimate"]["mode"] == "classical_only"
    assert payload["runtime_estimate"]["within_limit"] is True
    assert payload["runtime_estimate"]["basis"]["warm_start"] is True
    assert 0 < len(payload["diagnostics"]["logs"]) <= 50
    assert any("Inspection only" in line for line in payload["diagnostics"]["logs"])
    assert before["used_runs"] == 0
    assert after["used_runs"] == 0
    assert after["remaining_runs"] == 1000
    assert summary["total_runs"] == 0
    assert summary["consumed_runs"] == 0


def test_inspect_workbook_uses_workbook_settings_when_form_missing():
    response = _post_inspect(headers={"X-API-Key": "DEMO-123"}, mode="classical_only")
    payload = response.get_json()
    effective = payload["diagnostics"]["effective_settings"]

    assert response.status_code == 200
    assert effective["layers"] == 6
    assert effective["iterations"] == 100
    assert effective["restarts"] == 3
    assert effective["lambda_budget"] == 50.0
    assert effective["lambda_variance"] == 6.0
    assert effective["risk_free_rate"] == 0.04
    assert effective["sources"]["layers"] == "workbook"
    assert effective["sources"]["iterations"] == "workbook"
    assert effective["qaoa_shots_display"] == "not_applicable"


def test_inspect_workbook_uses_indicative_market_cost_column(tmp_path):
    workbook_path = _indicative_cost_workbook(tmp_path)

    response = _post_inspect(workbook_path, headers={"X-API-Key": "DEMO-123"}, mode="classical_only")
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["diagnostics"]["cost_column_used"] == "Indicative Market Cost USD"
    assert payload["diagnostics"]["cost_column_internal"] == "Indicative Market Cost USD"
    assert payload["diagnostics"]["cost_column_normalized"] is False
    assert payload["workbook_summary"]["cost_column_used"] == "Indicative Market Cost USD"


def test_indicative_market_cost_overrides_legacy_approx_cost_when_both_exist(tmp_path):
    workbook_path = _indicative_cost_workbook(tmp_path, keep_legacy=True, add_usd=1.0)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    assets = workbook["Assets"]
    headers = [cell.value for cell in assets[2]]
    indicative_col = headers.index("Indicative Market Cost USD") + 1
    expected_universe = sum(
        float(assets.cell(row_idx, indicative_col).value)
        for row_idx in range(3, assets.max_row + 1)
        if assets.cell(row_idx, indicative_col).value not in (None, "")
    )
    workbook.close()

    response = _post_inspect(workbook_path, headers={"X-API-Key": "DEMO-123"}, mode="classical_only")
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["workbook_summary"]["variable_candidate_universe"] == pytest.approx(expected_universe)
    assert payload["diagnostics"]["cost_column_conflicting_row_count"] > 0
    assert any("Indicative Market Cost USD was used" in warning for warning in payload["diagnostics"]["workbook_warnings"])


def test_existing_workbook_missing_additional_type_constraints_defaults_to_zero():
    response = _post_inspect(headers={"X-API-Key": "DEMO-123"}, mode="classical_only")
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["workbook_summary"]["additional_type_constraints_count"] == 0
    assert payload["workbook_summary"]["additional_type_constraints"] == []
    assert payload["diagnostics"]["additional_type_constraints_count"] == 0
    assert payload["diagnostics"]["additional_type_constraints"] == []


def test_additional_type_constraints_zero_adds_no_qubo_terms(tmp_path):
    base_path = _limited_workbook(tmp_path, qubits=3)
    zero_path = _type_constraint_workbook(tmp_path, qubits=3, constraint_count=0)

    base = build_qubo_from_workbook(base_path, lambda _message: None)
    zero = build_qubo_from_workbook(zero_path, lambda _message: None)

    assert zero.additional_type_constraints_count == 0
    np.testing.assert_allclose(zero.Q, base.Q)
    assert zero.constant == pytest.approx(base.constant)


def test_one_additional_type_constraint_parses_and_inspects(tmp_path):
    workbook_path = _type_constraint_workbook(tmp_path, qubits=3, constraint_count=1)

    response = _post_inspect(workbook_path, headers={"X-API-Key": "DEMO-123"}, mode="classical_only")
    payload = response.get_json()

    assert response.status_code == 200
    constraints = payload["diagnostics"]["additional_type_constraints"]
    assert payload["workbook_summary"]["additional_type_constraints_count"] == 1
    assert constraints[0]["id"] == "type_a"
    assert constraints[0]["name"] == "Bond"
    assert constraints[0]["budget"] == pytest.approx(100.0)
    assert constraints[0]["penalty"] == pytest.approx(30.0)
    assert constraints[0]["size_column"] == "Type A Size"


def test_multiple_additional_type_constraints_parse(tmp_path):
    workbook_path = _type_constraint_workbook(
        tmp_path,
        qubits=3,
        constraint_count=2,
        include_type_b_size=True,
    )
    optimizer = build_qubo_from_workbook(workbook_path, lambda _message: None)

    assert optimizer.additional_type_constraints_count == 2
    assert [item["id"] for item in optimizer.additional_type_constraints] == ["type_a", "type_b"]
    assert [item["label"] for item in optimizer.additional_type_constraints] == ["Bond", "Equity"]


def test_additional_type_constraint_qubo_expansion_includes_fixed_offset(tmp_path):
    roles = ["fixed", "variable", "variable", "variable"]
    type_a_values = [40.0, 20.0, 40.0, 80.0]
    base_path = _type_constraint_workbook(
        tmp_path,
        qubits=4,
        constraint_count=0,
        type_a_values=type_a_values,
        decision_roles=roles,
    )
    typed_path = _type_constraint_workbook(
        tmp_path,
        qubits=4,
        constraint_count=1,
        type_a_values=type_a_values,
        type_a_budget=100.0,
        type_a_penalty=30.0,
        decision_roles=roles,
    )
    base = build_qubo_from_workbook(base_path, lambda _message: None)
    typed = build_qubo_from_workbook(typed_path, lambda _message: None)

    b = typed.variable_options_df["Type A Size"].astype(float).to_numpy() / 100.0
    fixed_exposure = float(typed.fixed_options_df["Type A Size"].astype(float).sum())
    c = fixed_exposure / 100.0 - 1.0
    expected_delta = np.zeros_like(base.Q)
    for i in range(len(b)):
        expected_delta[i, i] += 30.0 * (2.0 * c * b[i] + b[i] ** 2)
        for j in range(i + 1, len(b)):
            expected_delta[i, j] += 2.0 * 30.0 * b[i] * b[j]

    np.testing.assert_allclose(typed.Q - base.Q, expected_delta)
    assert fixed_exposure == pytest.approx(40.0)
    assert typed.constant - base.constant == pytest.approx(30.0 * c**2)

    bitvec = np.array([1, 1, 0], dtype=int)
    normalized = float(fixed_exposure / 100.0 + np.dot(b, bitvec))
    expected_penalty = 30.0 * (normalized - 1.0) ** 2
    breakdown = typed.qubo_term_breakdown(bitvec)
    stats = typed.portfolio_stats(bitvec)

    assert breakdown["type_budget_term"] == pytest.approx(expected_penalty)
    assert breakdown["type_a_penalty"] == pytest.approx(expected_penalty)
    assert breakdown["qubo_reconstructed"] == pytest.approx(typed.qubo_value(bitvec))
    assert stats["type_a_fixed"] == pytest.approx(40.0)
    assert stats["type_a_variable_selected"] == pytest.approx(60.0)
    assert stats["type_a_achieved"] == pytest.approx(100.0)
    assert stats["type_a_relative_deviation"] == pytest.approx(0.0)


def test_type_constraint_can_be_reached_by_fixed_plus_variable_exposure(tmp_path):
    workbook_path = _type_constraint_workbook(
        tmp_path,
        qubits=4,
        constraint_count=1,
        type_a_values=[40.0, 20.0, 40.0, 80.0],
        type_a_budget=100.0,
        type_a_penalty=30.0,
        decision_roles=["fixed", "variable", "variable", "variable"],
    )
    optimizer = build_qubo_from_workbook(workbook_path, lambda _message: None)

    stats = optimizer.portfolio_stats(np.array([1, 1, 0], dtype=int))

    assert stats["type_a_fixed"] == pytest.approx(40.0)
    assert stats["type_a_variable_selected"] == pytest.approx(60.0)
    assert stats["type_a_achieved"] == pytest.approx(100.0)
    assert stats["type_a_variable_selected"] < stats["type_a_budget"]
    assert stats["type_a_penalty"] == pytest.approx(0.0)


def test_multiple_additional_type_constraints_are_additive(tmp_path):
    base_path = _type_constraint_workbook(tmp_path, qubits=3, constraint_count=0)
    typed_path = _type_constraint_workbook(
        tmp_path,
        qubits=3,
        constraint_count=2,
        include_type_b_size=True,
        type_a_values=[20.0, 50.0, 80.0],
        type_b_values=[5.0, 15.0, 30.0],
        type_a_budget=100.0,
        type_a_penalty=30.0,
        type_b_budget=50.0,
        type_b_penalty=12.0,
    )
    base = build_qubo_from_workbook(base_path, lambda _message: None)
    typed = build_qubo_from_workbook(typed_path, lambda _message: None)

    a = typed.variable_options_df["Type A Size"].astype(float).to_numpy() / 100.0
    b = typed.variable_options_df["Type B Size"].astype(float).to_numpy() / 50.0
    expected_delta = np.zeros_like(base.Q)
    for sizes, penalty in ((a, 30.0), (b, 12.0)):
        for i in range(len(sizes)):
            expected_delta[i, i] += penalty * (sizes[i] ** 2 - 2.0 * sizes[i])
            for j in range(i + 1, len(sizes)):
                expected_delta[i, j] += 2.0 * penalty * sizes[i] * sizes[j]

    np.testing.assert_allclose(typed.Q - base.Q, expected_delta)
    assert typed.constant - base.constant == pytest.approx(42.0)


def test_missing_active_type_size_column_fails_clearly(tmp_path):
    workbook_path = _type_constraint_workbook(
        tmp_path,
        qubits=3,
        constraint_count=1,
        include_type_a_size=False,
    )

    response = _post_inspect(workbook_path, headers={"X-API-Key": "DEMO-123"}, mode="classical_only")
    payload = response.get_json()

    assert response.status_code == 422
    assert payload["error"]["code"] == "optimization_error"
    assert "Type A Size column is required" in payload["error"]["message"]


def test_additional_type_constraints_count_must_be_integer_range(tmp_path):
    workbook_path = _limited_workbook(tmp_path, qubits=3)
    workbook = load_workbook(workbook_path)
    _set_setting(workbook["Settings"], "Additional Type Constraints", 1.5)
    workbook.save(workbook_path)

    response = _post_inspect(workbook_path, headers={"X-API-Key": "DEMO-123"}, mode="classical_only")
    payload = response.get_json()

    assert response.status_code == 422
    assert "Additional Type Constraints must be an integer from 0 to 5" in payload["error"]["message"]


def test_type_size_values_must_be_numeric_when_present(tmp_path):
    workbook_path = _type_constraint_workbook(tmp_path, qubits=3, constraint_count=1)
    workbook = load_workbook(workbook_path)
    assets = workbook["Assets"]
    headers = [cell.value for cell in assets[2]]
    type_a_col = headers.index("Type A Size") + 1
    assets.cell(3, type_a_col).value = "bad-size"
    workbook.save(workbook_path)

    response = _post_inspect(workbook_path, headers={"X-API-Key": "DEMO-123"}, mode="classical_only")
    payload = response.get_json()

    assert response.status_code == 422
    assert "Type A Size values must be numeric" in payload["error"]["message"]


def test_invalid_type_budget_setting_fails_clearly(tmp_path):
    workbook_path = _type_constraint_workbook(
        tmp_path,
        qubits=3,
        constraint_count=1,
        type_a_budget="not-a-number",
    )

    response = _post_inspect(workbook_path, headers={"X-API-Key": "DEMO-123"}, mode="classical_only")
    payload = response.get_json()

    assert response.status_code == 422
    assert "Type A Budget must be numeric" in payload["error"]["message"]


def test_type_budget_must_be_positive_for_normalization(tmp_path):
    workbook_path = _type_constraint_workbook(
        tmp_path,
        qubits=3,
        constraint_count=1,
        type_a_budget=0.0,
    )

    response = _post_inspect(workbook_path, headers={"X-API-Key": "DEMO-123"}, mode="classical_only")
    payload = response.get_json()

    assert response.status_code == 422
    assert "Type A Budget must be > 0" in payload["error"]["message"]


def test_run_qaoa_output_includes_additional_type_budget_diagnostics(tmp_path):
    workbook_path = _type_constraint_workbook(
        tmp_path,
        qubits=4,
        constraint_count=1,
        type_a_values=[40.0, 20.0, 40.0, 80.0],
        decision_roles=["fixed", "variable", "variable", "variable"],
    )

    response = _post_run_file(workbook_path, mode="classical_only")
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["diagnostics"]["additional_type_constraints_count"] == 1
    assert payload["diagnostics"]["additional_type_constraints"][0]["id"] == "type_a"
    achievement = payload["diagnostics"]["additional_type_budget_achievements"][0]
    assert achievement["id"] == "type_a"
    assert achievement["total_achieved_exposure"] == pytest.approx(
        achievement["fixed_exposure"] + achievement["variable_selected_exposure"]
    )
    assert "type_budget_term" in payload["components"]
    assert "type_a_achieved" in payload["best_candidate"]
    assert payload["best_candidate"]["type_a_fixed"] == pytest.approx(40.0)
    assert payload["best_candidate"]["type_a_achieved"] == pytest.approx(
        payload["best_candidate"]["type_a_fixed"] + payload["best_candidate"]["type_a_variable_selected"]
    )
    assert "type_a_penalty" in payload["top_candidates"][0]
    assert "type_a_fixed" in payload["top_candidates"][0]
    assert "type_a_variable_selected" in payload["top_candidates"][0]
    assert "Type A Size" in payload["selected_blocks"][0]


def test_inspect_workbook_form_values_override_workbook_settings():
    response = _post_inspect(
        headers={"X-API-Key": "DEMO-123"},
        mode="classical_only",
        layers="2",
        iterations="20",
        restarts="1",
        warm_start="true",
        qaoa_shots="123",
        lambda_budget="11",
        lambda_variance="2",
        risk_free_rate="0.03",
        restart_perturbation="0.2",
    )
    payload = response.get_json()
    effective = payload["diagnostics"]["effective_settings"]

    assert response.status_code == 200
    assert effective["layers"] == 2
    assert effective["iterations"] == 20
    assert effective["restarts"] == 1
    assert effective["warm_start"] is True
    assert effective["lambda_budget"] == 11.0
    assert effective["lambda_variance"] == 2.0
    assert effective["risk_free_rate"] == 0.03
    assert effective["restart_perturbation"] == 0.2
    assert effective["sources"]["layers"] == "form"
    assert effective["sources"]["lambda_budget"] == "form"
    assert payload["runtime_estimate"]["basis"]["layers"] == 2


def test_inspect_workbook_random_seed_form_field_is_effective():
    response = _post_inspect(headers={"X-API-Key": "DEMO-123"}, mode="classical_only", random_seed="12345")
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["diagnostics"]["effective_settings"]["random_seed"] == 12345
    assert payload["diagnostics"]["effective_settings"]["rng_seed"] == 12345
    assert payload["diagnostics"]["effective_settings"]["sources"]["random_seed"] == "form"
    assert payload["diagnostics"]["runtime_inputs"]["random_seed"] == 12345
    assert payload["diagnostics"]["random_seed"] == 12345
    assert any("Random seed: 12345" in line for line in payload["diagnostics"]["logs"])


def test_inspect_workbook_without_random_seed_still_uses_workbook_seed_if_present():
    response = _post_inspect(headers={"X-API-Key": "DEMO-123"}, mode="classical_only")
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["diagnostics"]["effective_settings"]["random_seed"] == 42
    assert payload["diagnostics"]["effective_settings"]["sources"]["random_seed"] == "workbook"


def test_inspect_workbook_random_seed_zero_is_valid():
    response = _post_inspect(headers={"X-API-Key": "DEMO-123"}, mode="classical_only", random_seed="0")
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["diagnostics"]["effective_settings"]["random_seed"] == 0
    assert payload["diagnostics"]["random_seed"] == 0


def test_inspect_workbook_rejects_non_integer_random_seed():
    response = _post_inspect(headers={"X-API-Key": "DEMO-123"}, mode="classical_only", random_seed="not-an-int")
    payload = response.get_json()

    assert response.status_code == 400
    assert payload["error"]["code"] == "invalid_random_seed"
    assert payload["error"]["details"]["field"] == "random_seed"


def test_inspect_workbook_rejects_out_of_range_random_seed():
    response = _post_inspect(headers={"X-API-Key": "DEMO-123"}, mode="classical_only", random_seed=str(2**32))
    payload = response.get_json()

    assert response.status_code == 400
    assert payload["error"]["code"] == "invalid_random_seed"
    assert payload["error"]["details"]["max"] == 4294967295


def test_workbook_budget_sanity_warnings_are_non_blocking():
    optimizer = _fake_budget_warning_optimizer()

    warnings = workbook_warnings(optimizer)

    assert "Fixed holdings exceed the configured budget before any variable block is selected." in warnings
    assert "Configured budget is smaller than fixed holdings plus the cheapest variable block." in warnings
    assert "Configured budget may leave little or no room for typical variable selections." in warnings


def test_inspect_workbook_includes_budget_sanity_warnings(monkeypatch):
    optimizer = _fake_budget_warning_optimizer()
    monkeypatch.setattr(main_module, "validate_required_input_sheets", lambda _path: None)
    monkeypatch.setattr(main_module, "workbook_structure", lambda _path: {})
    monkeypatch.setattr(main_module, "build_qubo_from_workbook", lambda _path, _log, _form=None: optimizer)

    response = _post_inspect(headers={"X-API-Key": "DEMO-123"}, mode="classical_only")
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["diagnostics"]["workbook_warning_count"] >= 3
    assert any("Fixed holdings exceed" in warning for warning in payload["diagnostics"]["workbook_warnings"])
    assert any("Workbook warning: Fixed holdings exceed" in line for line in payload["diagnostics"]["logs"])


def test_async_submission_logs_budget_sanity_warnings(monkeypatch, tmp_path):
    optimizer = _fake_budget_warning_optimizer()
    monkeypatch.setattr(main_module.Config, "LOCAL_JOB_DIR", tmp_path / "jobs")
    ledger = _RouteLockLedger()
    monkeypatch.setattr(main_module, "get_run_ledger", lambda: ledger)
    monkeypatch.setattr(main_module, "validate_required_input_sheets", lambda _path: None)
    monkeypatch.setattr(main_module, "workbook_structure", lambda _path: {})
    monkeypatch.setattr(main_module, "build_qubo_from_workbook", lambda _path, _log, _form=None: optimizer)
    monkeypatch.setattr(main_module, "trigger_cloud_run_job", lambda job_id: {"triggered": False, "mode": "test"})

    response = _post_async(headers={"X-API-Key": "DEMO-123"}, mode="classical_only")
    payload = response.get_json()
    job = get_job_store().get_job(payload["job_id"])

    assert response.status_code == 202
    assert any("Workbook warning: Fixed holdings exceed" in line for line in job["logs_tail"])
    assert ledger.acquired is True


def test_inspect_workbook_public_demo_enforces_qubit_limit():
    response = _post_inspect(mode="classical_only")
    payload = response.get_json()

    assert response.status_code == 403
    assert payload["error"]["code"] == "qubit_limit_exceeded"
    assert payload["license"]["usage_level"] == "public_demo"


def test_public_demo_qaoa_limited_rejects_above_small_limits():
    response = _post_run(
        mode="qaoa_limited",
        response_level="compact",
        qaoa_p="1",
        qaoa_maxiter="2",
        qaoa_multistart_restarts="1",
        warm_start="false",
    )
    payload = response.get_json()

    assert response.status_code == 403
    assert payload["error"]["code"] == "qubit_limit_exceeded"
    assert payload["error"]["message"]
    assert payload["error"]["details"]["usage_level"] == "public_demo"
    assert payload["error"]["details"]["mode"] == "qaoa_lightning_sim"
    assert payload["error"]["details"]["binary_variables"] == 14
    assert payload["error"]["details"]["max_qubits"] == 8


def test_inspect_workbook_qaoa_limited_returns_estimate_without_execution(tmp_path):
    workbook_path = _limited_workbook(tmp_path, qubits=7)
    response = _post_inspect(
        workbook_path,
        headers={"X-API-Key": "TESTER-123"},
        mode="qaoa_limited",
        layers="2",
        iterations="10",
        restarts="1",
        warm_start="false",
        qaoa_shots="64",
        restart_perturbation="0.1",
    )
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["workbook_summary"]["n_qubits"] == 7
    assert payload["runtime_estimate"]["mode"] == "qaoa_lightning_sim"
    assert payload["runtime_estimate"]["basis"]["n_qubits"] == 7
    assert payload["runtime_estimate"]["basis"]["layers"] == 2
    assert payload["runtime_estimate"]["basis"]["iterations"] == 10
    assert payload["runtime_estimate"]["basis"]["restarts"] == 1
    assert payload["runtime_estimate"]["basis"]["warm_start"] is False
    assert payload["runtime_estimate"]["basis"]["qaoa_shots_display"] == "exact"
    assert payload["runtime_estimate"]["within_limit"] is True
    assert payload["runtime_estimate"]["max_estimated_runtime_sec"] == 2700
    assert payload["diagnostics"]["effective_settings"]["qaoa_shots"] is None
    assert payload["diagnostics"]["effective_settings"]["qaoa_shots_display"] == "exact"
    assert payload["diagnostics"]["effective_settings"]["shots_mode"] == "exact"
    assert payload["diagnostics"]["requested_run_mode"] == "qaoa_limited"
    assert payload["diagnostics"]["run_mode"] == "qaoa_lightning_sim"
    assert payload["diagnostics"]["simulation_backend"] == "lightning.qubit"
    assert payload["diagnostics"]["legacy_run_mode_alias"] is True
    assert payload["diagnostics"]["hardware_replay"] is False
    assert any("optimization execution skipped" in line for line in payload["diagnostics"]["logs"])
    assert any("not executed during workbook inspection" in line for line in payload["diagnostics"]["logs"])


def test_inspect_workbook_qaoa_lightning_sim_is_accepted_with_diagnostics(tmp_path):
    workbook_path = _limited_workbook(tmp_path, qubits=4)
    response = _post_inspect(
        workbook_path,
        headers={"X-API-Key": "TESTER-123"},
        mode="qaoa_lightning_sim",
    )
    payload = response.get_json()

    assert response.status_code == 200
    diagnostics = payload["diagnostics"]
    assert diagnostics["requested_run_mode"] == "qaoa_lightning_sim"
    assert diagnostics["run_mode"] == "qaoa_lightning_sim"
    assert diagnostics["simulation_backend"] == "lightning.qubit"
    assert diagnostics["legacy_run_mode_alias"] is False
    assert diagnostics["hardware_replay"] is False
    assert diagnostics["effective_settings"]["response_level"] == "full"


def test_inspect_workbook_qaoa_tensor_sim_selects_tensor_backend(tmp_path):
    workbook_path = _limited_workbook(tmp_path, qubits=4)
    response = _post_inspect(
        workbook_path,
        headers={"X-API-Key": "TESTER-123"},
        mode="qaoa_tensor_sim",
    )
    payload = response.get_json()

    assert response.status_code == 200
    diagnostics = payload["diagnostics"]
    assert diagnostics["requested_run_mode"] == "qaoa_tensor_sim"
    assert diagnostics["run_mode"] == "qaoa_tensor_sim"
    assert diagnostics["simulation_backend"] == "default.tensor"
    assert diagnostics["legacy_run_mode_alias"] is False
    assert diagnostics["hardware_replay"] is False
    assert diagnostics["effective_settings"]["response_level"] == "full"


def test_inspect_workbook_missing_mode_defaults_to_qaoa_lightning_sim(tmp_path):
    workbook_path = _limited_workbook(tmp_path, qubits=4)
    response = _post_inspect(workbook_path, headers={"X-API-Key": "TESTER-123"})
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["runtime_estimate"]["mode"] == "qaoa_lightning_sim"
    assert payload["diagnostics"]["requested_run_mode"] == "qaoa_lightning_sim"
    assert payload["diagnostics"]["run_mode"] == "qaoa_lightning_sim"
    assert payload["diagnostics"]["simulation_backend"] == "lightning.qubit"
    assert payload["diagnostics"]["legacy_run_mode_alias"] is False
    assert payload["diagnostics"]["effective_settings"]["response_level"] == "full"


def test_policy_rejects_qaoa_full_estimate():
    usage_levels = load_usage_config()["usage_levels"]
    context = UsageContext(
        usage_level_name="tester",
        usage_level=usage_levels["tester"],
        key_record={"key_id": "tester"},
        authenticated=True,
    )
    optimizer = SimpleNamespace(
        n=3,
        qaoa_p=2,
        qaoa_maxiter=5,
        qaoa_multistart_restarts=1,
        qaoa_layerwise_warm_start=False,
        qaoa_shots=256,
        qaoa_restart_perturbation=0.05,
        lambda_budget=50.0,
        lambda_variance=6.0,
        risk_free=0.04,
        settings={},
        classical_results=None,
    )

    with pytest.raises(ApiError) as exc_info:
        validate_problem_policy(context, optimizer, "qaoa_full", {})

    assert exc_info.value.status_code == 400
    assert exc_info.value.code == "unsupported_mode"


def test_qaoa_limited_runtime_estimate_is_calibrated_for_24_qubit_exact_mode():
    usage_levels = load_usage_config()["usage_levels"]
    context = UsageContext(
        usage_level_name="internal_power",
        usage_level=usage_levels["internal_power"],
        key_record={"key_id": INTERNAL_POWER_KEY_ID},
        authenticated=True,
    )
    optimizer = SimpleNamespace(
        n=24,
        qaoa_p=1,
        qaoa_maxiter=30,
        qaoa_multistart_restarts=1,
        qaoa_layerwise_warm_start=False,
        qaoa_shots=None,
        qaoa_restart_perturbation=None,
        lambda_budget=50.0,
        lambda_variance=6.0,
        risk_free=0.04,
        settings={},
        classical_results=None,
    )
    runtime_inputs = RuntimeInputs(layers=1, iterations=30, restarts=1, warm_start=False)

    policy_result = estimate_policy_result(
        context,
        optimizer,
        "qaoa_limited",
        runtime_inputs=runtime_inputs,
        candidate_count=1 << 24,
    )

    assert 500 <= policy_result.raw_estimated_runtime_sec <= 510
    assert policy_result.estimated_runtime_sec >= 1200
    assert policy_result.estimated_runtime_sec > policy_result.raw_estimated_runtime_sec
    assert policy_result.within_limit is True


def test_qaoa_limited_keeps_5000_export_rows_when_within_safety_cap(monkeypatch):
    optimizer = SimpleNamespace(n=24, qaoa_max_export_rows=5000)
    runtime_inputs = RuntimeInputs(layers=1, iterations=30, restarts=1, warm_start=False)

    qaoa_engine_module._configure_limited_qaoa(optimizer, runtime_inputs, max_qubits=24)

    assert optimizer.qaoa_export_requested_rows == 5000
    assert optimizer.qaoa_max_export_rows == 5000
    assert optimizer.qaoa_export_cap_applied is False
    assert optimizer.qaoa_export_cap_reason == "requested_rows_within_safety_cap"


def test_qaoa_limited_export_safety_cap_is_explicit(monkeypatch):
    monkeypatch.setenv("QAOA_LIMITED_MAX_EXPORT_ROWS_CAP", "5000")
    optimizer = SimpleNamespace(
        n=24,
        qaoa_max_export_rows=8000,
        qaoa_export_cap_applied=True,
        qaoa_export_cap_reason="qaoa_limited_exact_export_safety_cap",
        samples_df=pd.DataFrame({"bitstring": ["0" * 24]}),
        enable_qaoa=True,
        qaoa_total_states_considered=1 << 24,
        qaoa_min_probability_to_export=0.0,
        qaoa_export_mode="top_k",
        qaoa_export_sort_by="probability",
        qaoa_export_feasible_only=False,
        qaoa_exact_probability_max_qubits=24,
    )
    runtime_inputs = RuntimeInputs(layers=1, iterations=30, restarts=1, warm_start=False)

    qaoa_engine_module._configure_limited_qaoa(optimizer, runtime_inputs, max_qubits=24)
    diagnostics = candidate_export_diagnostics(optimizer)

    assert optimizer.qaoa_export_requested_rows == 8000
    assert optimizer.qaoa_max_export_rows == 5000
    assert diagnostics["qaoa_export_cap_applied"] is True
    assert diagnostics["qaoa_export_cap_reason"] == "qaoa_limited_exact_export_safety_cap"
    assert diagnostics["qaoa_exact_state_space"] == 1 << 24


def test_qaoa_sim_switches_to_sampling_above_exact_probability_cap():
    optimizer = SimpleNamespace(n=26, qaoa_max_export_rows=5000, qaoa_shots=4096)
    runtime_inputs = RuntimeInputs(layers=1, iterations=10, restarts=1, warm_start=False, qaoa_shots=2048)

    qaoa_engine_module._configure_limited_qaoa(optimizer, runtime_inputs, max_qubits=30)

    assert optimizer.qaoa_exact_probability_max_qubits == 24
    assert optimizer.qaoa_limited_exact_probabilities is False
    assert optimizer.qaoa_sim_exact_probabilities is False
    assert optimizer.qaoa_shots == 2048


def test_qaoa_sampling_normalization_preserves_sampled_scope():
    optimizer = SimpleNamespace(
        n=26,
        enable_qaoa=True,
        qaoa_max_export_rows=5000,
        qaoa_total_states_considered=2048,
        qaoa_limited_exact_probabilities=False,
        qaoa_sim_exact_probabilities=False,
        samples_df=pd.DataFrame(
            {
                "bitstring": ["0" * 26],
                "qubo_value": [1.23],
                "probability": [0.5],
                "selection_scope": ["unique sampled states from 2,048 shots"],
            }
        ),
    )
    optimizer.sort_candidates = lambda frame: frame.sort_values("qubo_value")

    qaoa_engine_module._normalize_limited_qaoa_outputs(optimizer, run_mode="qaoa_lightning_sim")
    diagnostics = candidate_export_diagnostics(optimizer)

    assert optimizer.samples_df.iloc[0]["source"] == "qaoa_lightning_sim"
    assert optimizer.samples_df.iloc[0]["selection_scope"] == "unique sampled states from 2,048 shots"
    assert optimizer.qaoa_exact_states_exported == 0
    assert optimizer.qaoa_sampled_states_exported == 1
    assert diagnostics["qaoa_exact_states_evaluated"] is None
    assert diagnostics["qaoa_exact_states_exported"] == 0
    assert diagnostics["qaoa_sampled_states_evaluated"] == 2048
    assert diagnostics["qaoa_sampled_states_exported"] == 1


def test_policy_reports_sampling_shots_display_above_exact_probability_cap():
    usage_levels = load_usage_config()["usage_levels"]
    context = UsageContext(
        usage_level_name="internal_qaoa_30",
        usage_level=usage_levels["internal_qaoa_30"],
        key_record={"key_id": "internal-qaoa-30"},
        authenticated=True,
    )
    optimizer = SimpleNamespace(
        n=26,
        qaoa_p=1,
        qaoa_maxiter=10,
        qaoa_multistart_restarts=1,
        qaoa_layerwise_warm_start=False,
        qaoa_shots=2048,
        qaoa_restart_perturbation=None,
        lambda_budget=50.0,
        lambda_variance=6.0,
        risk_free=0.04,
        settings={},
        classical_results=None,
    )
    runtime_inputs = RuntimeInputs(layers=1, iterations=10, restarts=1, warm_start=False, qaoa_shots=2048)

    policy_result = estimate_policy_result(
        context,
        optimizer,
        "qaoa_lightning_sim",
        runtime_inputs=runtime_inputs,
        candidate_count=2048,
    )

    assert policy_result.effective_settings["shots_mode"] == "sampling"
    assert policy_result.effective_settings["qaoa_shots"] == 2048
    assert policy_result.effective_settings["qaoa_shots_display"] == "2048"


def test_policy_reports_sampling_shots_display_for_tensor_sim_under_exact_cap():
    usage_levels = load_usage_config()["usage_levels"]
    context = UsageContext(
        usage_level_name="tester",
        usage_level=usage_levels["tester"],
        key_record={"key_id": "tester"},
        authenticated=True,
    )
    optimizer = SimpleNamespace(
        n=7,
        qaoa_p=1,
        qaoa_maxiter=10,
        qaoa_multistart_restarts=1,
        qaoa_layerwise_warm_start=False,
        qaoa_shots=1024,
        qaoa_restart_perturbation=None,
        lambda_budget=50.0,
        lambda_variance=6.0,
        risk_free=0.04,
        settings={},
        classical_results=None,
    )
    runtime_inputs = RuntimeInputs(layers=1, iterations=10, restarts=1, warm_start=False, qaoa_shots=1024)

    policy_result = estimate_policy_result(
        context,
        optimizer,
        "qaoa_tensor_sim",
        runtime_inputs=runtime_inputs,
        candidate_count=128,
    )

    assert policy_result.effective_settings["shots_mode"] == "sampling"
    assert policy_result.effective_settings["qaoa_shots"] == 1024
    assert policy_result.effective_settings["qaoa_shots_display"] == "1024"


def test_classical_export_diagnostics_explain_fewer_unique_candidates():
    optimizer = SimpleNamespace(
        n=24,
        random_search_samples=8000,
        local_search_starts=40,
        classical_results=pd.DataFrame({"bitstring": [f"{idx:024b}" for idx in range(7119)]}),
        samples_df=pd.DataFrame(),
        enable_qaoa=False,
        top_n_export=20,
        overview_classical_pool=300,
        overview_qaoa_pool=500,
        result_candidate_limit_per_solver=500,
    )

    diagnostics = candidate_export_diagnostics(optimizer)

    assert diagnostics["classical_export_requested_rows"] == 8024
    assert diagnostics["classical_export_actual_rows"] == 7119
    assert diagnostics["classical_export_cap_applied"] is True
    assert diagnostics["classical_export_cap_reason"] == "unique_candidate_count_after_duplicate_removal_or_search_convergence"


def test_demo_key_gets_qualified_demo_full_response_by_default():
    response = _post_run(headers={"X-API-Key": "DEMO-123"}, mode="classical_only")
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["status"] == "completed"
    assert payload["model_version"] == "9.1.0"
    assert payload["mode"] == "classical_only"
    assert payload["license"]["usage_level"] == "qualified_demo"
    assert payload["license"]["max_runs"] == 1000
    assert payload["license"]["used_runs"] == 0
    assert payload["license"]["remaining_runs"] == 1000
    assert payload["binary_variables"] == 14
    assert payload["best_bitstring"]
    assert payload["qubo_value"] is not None
    assert payload["selected_usd"] is not None
    assert payload["diagnostics"]["estimated_runtime_sec"] is not None
    assert payload["diagnostics"]["estimated_runtime_sec"] >= 1.0
    assert payload["diagnostics"]["actual_runtime_sec"] is not None
    assert payload["diagnostics"]["runtime_ratio"] is not None
    assert payload["diagnostics"]["n_qubits"] == 14
    assert 0 < payload["diagnostics"]["candidate_count"] <= 2**14
    assert "top_candidates" in payload
    assert "best_candidate" in payload
    assert "effective_settings" in payload["diagnostics"]
    assert payload["diagnostics"]["effective_settings"]["qaoa_shots_display"] == "not_applicable"
    assert payload["diagnostics"]["effective_settings"]["shots_mode"] == "disabled"
    assert payload["diagnostics"]["runtime_inputs"]["layers"] == payload["diagnostics"]["effective_settings"]["layers"]
    assert len(payload["diagnostics"]["logs"]) > 0
    assert any("Selected mode: classical_only" in line for line in payload["diagnostics"]["logs"])
    assert any("Refresh of Data:" in line for line in payload["diagnostics"]["logs"])
    assert not any("Refresh with yfinance" in line for line in payload["diagnostics"]["logs"])
    assert "reporting" in payload
    summary = payload["reporting"]["summary"]
    assert summary["classical_candidate_count"] > 0
    assert summary["qaoa_enabled"] is False
    assert summary["classical_result_summary"]["status"] == "Available"
    assert summary["classical_result_summary"]["best_bitstring"] == payload["best_bitstring"]
    assert summary["classical_result_summary"]["qubo_value"] == payload["qubo_value"]
    assert summary["quantum_result_summary"]["status"] == "Disabled / Not available"
    assert summary["quantum_result_summary"]["available"] is False
    assert summary["quantum_result_summary"]["best_bitstring"] is None
    assert summary["quantum_result_summary"]["qubo_value"] is None
    assert summary["quantum_result_summary"]["selected_usd"] is None
    assert payload["reporting"]["classical_candidates"]
    assert "charts" in payload["reporting"]


def test_demo_key_can_request_standard_response_with_limited_candidates():
    response = _post_run(
        headers={"X-API-Key": "DEMO-123"},
        mode="classical_only",
        response_level="standard",
    )
    payload = response.get_json()

    assert response.status_code == 200
    assert "components" in payload
    assert "top_candidates" in payload
    assert len(payload["top_candidates"]) <= 5
    assert 0 < len(payload["diagnostics"]["logs"]) <= 50
    assert any("QUBO matrix size" in line for line in payload["diagnostics"]["logs"])
    reporting = payload["reporting"]
    assert reporting["summary"]["classical_candidate_count"] > 0
    assert reporting["summary"]["classical_result_summary"]["source"] == "Classical_Candidates"
    assert reporting["summary"]["quantum_result_summary"]["future_source"] == "QAOA_Samples and QAOA_Best_QUBO"
    assert reporting["classical_candidates"]
    assert len(reporting["classical_candidates"]) <= 20
    assert reporting["solver_comparison"]
    assert reporting["portfolio_contents"]
    assert "quantum_samples" not in reporting
    assert "charts" not in reporting


def test_tester_key_can_request_full_response():
    response = _post_run(
        headers={"X-API-Key": "TESTER-123"},
        mode="classical_only",
        response_level="full",
    )
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["license"]["usage_level"] == "tester"
    assert "best_candidate" in payload
    assert "top_candidates" in payload
    assert "logs" in payload["diagnostics"]
    assert "input_sheet_names" in payload["diagnostics"]
    assert "ignored_output_sheets" in payload["diagnostics"]
    reporting = payload["reporting"]
    assert reporting["summary"]["qaoa_status"] == "disabled_or_not_available_in_version_7_classical_only"
    assert reporting["summary"]["quantum_result_summary"]["status"] == "Disabled / Not available"
    assert reporting["summary"]["quantum_result_summary"]["qubo_value"] is None
    assert reporting["classical_candidates"]
    assert reporting["portfolio_contents"]
    assert reporting["solver_comparison"]
    assert reporting["quantum_samples"] == []
    assert reporting["qaoa_best_qubo"] == []
    assert isinstance(reporting["optimization_history"], list)
    assert reporting["circuit"]["available"] is False
    assert payload["diagnostics"]["circuit"]["available"] is False
    assert reporting["charts"]["risk_return_sharpe"].startswith("data:image/png;base64,")
    assert reporting["charts"]["risk_return_qubo"].startswith("data:image/png;base64,")
    assert reporting["charts"]["qubo_breakdown"].startswith("data:image/png;base64,")
    assert reporting["charts"]["qubo_breakdown_classical"].startswith("data:image/png;base64,")
    assert reporting["charts"]["qubo_breakdown_quantum"] is None
    assert reporting["charts"]["optimization_history"] is None
    assert reporting["charts"]["circuit_overview"] is None
    assert reporting["charts"]["solver_comparison"].startswith("data:image/png;base64,")


def test_invalid_key_rejected():
    response = _post_run(headers={"X-API-Key": "NOT-A-REAL-KEY"}, mode="classical_only")
    payload = response.get_json()

    assert response.status_code == 401
    assert payload["error"]["code"] == "invalid_api_key"


def test_demo_key_full_response_is_allowed_and_consumes_success(monkeypatch, tmp_path):
    _enable_temp_ledger(monkeypatch, tmp_path)
    response = _post_run(
        headers={"X-API-Key": "DEMO-123"},
        mode="classical_only",
        response_level="full",
    )
    payload = response.get_json()
    status = app.test_client().get("/license-status", headers={"X-API-Key": "DEMO-123"}).get_json()
    summary = app.test_client().get("/ledger-summary").get_json()

    assert response.status_code == 200
    assert payload["status"] == "completed"
    assert "best_candidate" in payload
    assert status["used_runs"] == 1
    assert status["remaining_runs"] == 999
    assert summary["completed_runs"] == 1
    assert summary["consumed_runs"] == 1


def test_public_oversized_rejection_does_not_consume_run(monkeypatch, tmp_path):
    _enable_temp_ledger(monkeypatch, tmp_path)
    response = _post_run(mode="classical_only")
    summary = app.test_client().get("/ledger-summary").get_json()

    assert response.status_code == 403
    assert response.get_json()["error"]["code"] == "qubit_limit_exceeded"
    assert summary["rejected_runs"] == 1
    assert summary["consumed_runs"] == 0


def test_qaoa_mode_returns_unsupported_and_does_not_consume(monkeypatch, tmp_path):
    _enable_temp_ledger(monkeypatch, tmp_path)
    response = _post_run(headers={"X-API-Key": "TESTER-123"}, mode="qaoa")
    payload = response.get_json()
    status = app.test_client().get("/license-status", headers={"X-API-Key": "TESTER-123"}).get_json()
    summary = app.test_client().get("/ledger-summary").get_json()

    assert response.status_code == 400
    assert payload["error"]["code"] == "unsupported_mode"
    assert payload["error"]["details"]["received_mode"] == "qaoa"
    assert status["used_runs"] == 0
    assert status["remaining_runs"] == 1000
    assert summary["rejected_runs"] == 1
    assert summary["consumed_runs"] == 0


def test_qaoa_full_mode_returns_unsupported_error():
    response = _post_run(headers={"X-API-Key": "TESTER-123"}, mode="qaoa_full")
    payload = response.get_json()

    assert response.status_code == 400
    assert payload["error"]["code"] == "unsupported_mode"
    assert payload["error"]["message"] == "Unsupported mode."
    assert payload["error"]["details"]["received_mode"] == "qaoa_full"


def test_public_demo_qaoa_limited_runs_within_small_limits(tmp_path):
    workbook_path = _limited_workbook(tmp_path, qubits=3)
    response = _post_run_file(
        workbook_path,
        mode="qaoa_limited",
        response_level="compact",
        qaoa_p="1",
        qaoa_maxiter="2",
        qaoa_multistart_restarts="1",
        warm_start="false",
    )
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["status"] == "completed"
    assert payload["license"]["usage_level"] == "public_demo"
    assert payload["mode"] == "qaoa_lightning_sim"
    assert payload["binary_variables"] == 3
    assert payload["diagnostics"]["qaoa_enabled"] is True
    assert payload["diagnostics"]["effective_settings"]["qaoa_shots"] is None
    assert payload["diagnostics"]["effective_settings"]["qaoa_shots_display"] == "exact"
    assert payload["diagnostics"]["circuit"]["shots_mode"] == "exact"


def test_qaoa_limited_rejects_over_16_qubit_safety_cap():
    response = _post_run_file(
        EXTENDED_WORKBOOK,
        headers={"X-API-Key": "TESTER-123"},
        mode="qaoa_limited",
        qaoa_p="1",
        qaoa_maxiter="2",
        qaoa_multistart_restarts="1",
    )
    payload = response.get_json()

    assert response.status_code == 403
    assert payload["error"]["code"] == "qaoa_runtime_limit_exceeded"
    assert payload["error"]["details"]["field"] == "max_qubits"
    assert payload["error"]["details"]["requested"] > 16
    assert payload["error"]["details"]["allowed"] == 16
    assert payload["error"]["details"]["binary_variables"] > 16
    assert payload["error"]["details"]["license_max_qubits"] == 24
    assert payload["error"]["details"]["qaoa_max_qubits"] == 16
    assert payload["license"]["usage_level"] == "tester"


def test_tester_qaoa_limited_runs_exact_statevector_full_response(tmp_path):
    workbook_path = _limited_workbook(tmp_path, qubits=7)

    response = _post_run_file(
        workbook_path,
        headers={"X-API-Key": "TESTER-123"},
        mode="qaoa_limited",
        response_level="full",
        qaoa_p="1",
        qaoa_maxiter="4",
        qaoa_multistart_restarts="1",
        qaoa_shots="8",
    )
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["status"] == "completed"
    assert payload["mode"] == "qaoa_lightning_sim"
    assert payload["solver"] == "classical_heuristic+qaoa_lightning_sim"
    assert payload["binary_variables"] == 7
    assert payload["diagnostics"]["qaoa_enabled"] is True
    assert payload["diagnostics"]["qaoa_available"] is True
    assert payload["diagnostics"]["qaoa_status"] == "available"
    assert payload["diagnostics"]["qaoa_mode"] == "qaoa_lightning_sim"
    assert payload["diagnostics"]["requested_run_mode"] == "qaoa_limited"
    assert payload["diagnostics"]["run_mode"] == "qaoa_lightning_sim"
    assert payload["diagnostics"]["simulation_backend"] == "lightning.qubit"
    assert payload["diagnostics"]["legacy_run_mode_alias"] is True
    assert payload["diagnostics"]["hardware_replay"] is False
    assert payload["diagnostics"]["qaoa_candidate_count"] > 0
    assert payload["diagnostics"]["qaoa_p"] == 1
    assert payload["diagnostics"]["qaoa_iterations"] == 4
    assert payload["diagnostics"]["qaoa_restarts"] == 1
    assert payload["diagnostics"]["qaoa_exact_probabilities"] is True
    assert payload["diagnostics"]["qaoa_runtime_sec"] is not None
    assert payload["diagnostics"]["circuit"]["available"] is True
    assert payload["diagnostics"]["circuit"]["counts_are_estimated"] is True
    assert payload["diagnostics"]["circuit"]["n_qubits"] == 7
    assert payload["diagnostics"]["circuit"]["layers"] == 1
    assert payload["diagnostics"]["circuit"]["shots_mode"] == "exact"
    assert payload["diagnostics"]["circuit"]["qaoa_shots"] is None
    assert payload["diagnostics"]["circuit"]["qaoa_shots_display"] == "exact"
    ibm_circuit = payload["diagnostics"]["circuit"]["ibm"]
    assert ibm_circuit["available"] is True
    assert ibm_circuit["dry_run"] is True
    assert ibm_circuit["qiskit_available"] is True
    assert ibm_circuit["hardware_submission"] == "not_configured"
    assert ibm_circuit["measurement_required_for_sampler"] is True
    assert ibm_circuit["counts_decoder"] == "reverse_qiskit_count_key"
    assert ibm_circuit["classical_bits"] == 7
    assert ibm_circuit["qiskit_depth_with_measurements"] >= ibm_circuit["qiskit_depth_without_measurements"]
    assert ibm_circuit["qiskit_size_with_measurements"] > ibm_circuit["qiskit_size_without_measurements"]
    assert ibm_circuit["qiskit_gate_counts_with_measurements"]["measure"] == 7
    assert ibm_circuit["openqasm3"]["available"] is True
    assert ibm_circuit["openqasm3"]["preview"].startswith("OPENQASM 3.0;")
    assert payload["diagnostics"]["effective_settings"]["qaoa_shots"] is None
    assert payload["diagnostics"]["effective_settings"]["qaoa_shots_display"] == "exact"

    reporting = payload["reporting"]
    assert reporting["classical_candidates"]
    assert reporting["summary"]["classical_result_summary"]["status"] == "Available"
    assert reporting["summary"]["quantum_result_summary"]["status"] == "Available"
    assert reporting["summary"]["quantum_result_summary"]["source"] == "QAOA_Best_QUBO"
    assert reporting["summary"]["quantum_result_summary"]["probability"] is not None
    assert reporting["summary"]["qaoa_enabled"] is True
    assert reporting["summary"]["qaoa_available"] is True
    assert reporting["summary"]["qaoa_mode"] == "qaoa_lightning_sim"
    assert reporting["summary"]["qaoa_candidate_count"] > 0
    assert reporting["quantum_samples"]
    assert reporting["quantum_samples"][0]["source"] == "qaoa_lightning_sim"
    assert reporting["quantum_samples"][0]["selection_scope"] == "qaoa exact probability sample"
    assert reporting["qaoa_best_qubo"]
    assert reporting["qaoa_best_qubo"][0]["source"] == "qaoa_lightning_sim"
    assert reporting["qaoa_best_qubo"][0]["selection_scope"] == "qaoa exact probability sample"
    assert reporting["qaoa_best_qubo"][0]["qubo_value"] == min(row["qubo_value"] for row in reporting["qaoa_best_qubo"])
    assert reporting["solver_comparison"]
    assert {row["solver"] for row in reporting["solver_comparison"]} >= {"Classical Heuristic"}
    assert any("QAOA" in row["solver"] for row in reporting["solver_comparison"])
    assert any(row.get("source") == "qaoa_best_qubo" for row in reporting["portfolio_contents"])
    assert reporting["circuit"]["available"] is True
    assert reporting["circuit"]["counts_are_estimated"] is True
    assert reporting["circuit"]["n_qubits"] == 7
    assert reporting["circuit"]["ibm"]["qiskit_available"] is True
    assert reporting["charts"]["qubo_breakdown_quantum"].startswith("data:image/png;base64,")
    assert reporting["charts"]["optimization_history"].startswith("data:image/png;base64,")
    assert reporting["charts"]["circuit_overview"].startswith("data:image/png;base64,")
    assert reporting["charts"]["qubo_breakdown_classical"].startswith("data:image/png;base64,")


def test_tester_qaoa_tensor_sim_uses_sampling_shots(tmp_path):
    workbook_path = _limited_workbook(tmp_path, qubits=3)

    response = _post_run_file(
        workbook_path,
        headers={"X-API-Key": "TESTER-123"},
        mode="qaoa_tensor_sim",
        response_level="full",
        qaoa_p="1",
        qaoa_maxiter="2",
        qaoa_multistart_restarts="1",
        qaoa_shots="8",
    )
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["status"] == "completed"
    assert payload["mode"] == "qaoa_tensor_sim"
    assert payload["solver"] == "classical_heuristic+qaoa_tensor_sim"
    assert payload["diagnostics"]["requested_run_mode"] == "qaoa_tensor_sim"
    assert payload["diagnostics"]["run_mode"] == "qaoa_tensor_sim"
    assert payload["diagnostics"]["simulation_backend"] == "default.tensor"
    assert payload["diagnostics"]["qaoa_enabled"] is True
    assert payload["diagnostics"]["qaoa_available"] is True
    assert payload["diagnostics"]["qaoa_mode"] == "qaoa_tensor_sim"
    assert payload["diagnostics"]["qaoa_exact_probabilities"] is False
    assert payload["diagnostics"]["qaoa_candidate_count"] > 0
    assert payload["diagnostics"]["circuit"]["shots_mode"] == "sampling"
    assert payload["diagnostics"]["circuit"]["qaoa_shots"] == 8
    assert payload["diagnostics"]["effective_settings"]["qaoa_shots"] == 8
    assert payload["diagnostics"]["effective_settings"]["qaoa_shots_display"] == "8"
    assert payload["reporting"]["summary"]["quantum_result_summary"]["available"] is True
    assert payload["reporting"]["quantum_samples"]
    assert payload["reporting"]["quantum_samples"][0]["source"] == "qaoa_tensor_sim"
    assert "shots" in payload["reporting"]["quantum_samples"][0]["selection_scope"]


def test_qaoa_limited_same_random_seed_repeats_best_bitstring(tmp_path):
    workbook_path = _limited_workbook(tmp_path, qubits=3)
    common_fields = {
        "mode": "qaoa_limited",
        "response_level": "compact",
        "layers": "1",
        "iterations": "2",
        "restarts": "1",
        "warm_start": "false",
        "random_seed": "2468",
    }

    first = _post_run_file(workbook_path, headers={"X-API-Key": "TESTER-123"}, **common_fields)
    second = _post_run_file(workbook_path, headers={"X-API-Key": "TESTER-123"}, **common_fields)
    first_payload = first.get_json()
    second_payload = second.get_json()

    assert first.status_code == 200
    assert second.status_code == 200
    assert first_payload["diagnostics"]["random_seed"] == 2468
    assert second_payload["diagnostics"]["random_seed"] == 2468
    assert first_payload["best_bitstring"] == second_payload["best_bitstring"]


def test_tester_qaoa_limited_runs_14_qubit_small_workbook():
    response = _post_run(
        headers={"X-API-Key": "TESTER-123"},
        mode="qaoa_limited",
        response_level="compact",
        qaoa_p="1",
        qaoa_maxiter="10",
        qaoa_multistart_restarts="1",
    )
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["status"] == "completed"
    assert payload["mode"] == "qaoa_lightning_sim"
    assert payload["binary_variables"] == 14
    assert payload["diagnostics"]["qaoa_enabled"] is True
    assert payload["diagnostics"]["qaoa_available"] is True
    assert payload["diagnostics"]["qaoa_mode"] == "qaoa_lightning_sim"
    assert payload["diagnostics"]["qaoa_exact_probabilities"] is True
    assert payload["diagnostics"]["qaoa_p"] == 1
    assert payload["diagnostics"]["qaoa_iterations"] == 10
    assert payload["diagnostics"]["qaoa_restarts"] == 1
    assert payload["reporting"]["summary"]["quantum_result_summary"]["available"] is True


def test_qaoa_limited_rejects_runtime_limits(tmp_path):
    workbook_path = _limited_workbook(tmp_path, qubits=4)

    response = _post_run_file(
        workbook_path,
        headers={"X-API-Key": "TESTER-123"},
        mode="qaoa_limited",
        qaoa_p="7",
        qaoa_maxiter="201",
        qaoa_multistart_restarts="4",
    )
    payload = response.get_json()

    assert response.status_code == 403
    assert payload["error"]["code"] == "qaoa_runtime_limit_exceeded"
    assert payload["error"]["details"]["usage_level"] == "tester"
    assert payload["error"]["details"]["field"] == "layers"
    assert payload["error"]["details"]["requested"] == 7
    assert payload["error"]["details"]["allowed"] == 6
    exceeded = {entry["field"]: entry for entry in payload["error"]["details"]["exceeded"]}
    assert exceeded["layers"]["allowed"] == 6
    assert exceeded["iterations"]["allowed"] == 200
    assert exceeded["restarts"]["allowed"] == 3
    assert payload["license"]["usage_level"] == "tester"


def test_internal_power_qaoa_limited_accepts_14_qubits_above_tester_layer_cap(monkeypatch):
    import app.main as main_module

    def _mock_qaoa_lightning(optimizer, runtime_inputs, logs=None, max_qubits=None, **_kwargs):
        logs = logs if logs is not None else []
        optimizer.qaoa_limited_runtime_inputs = {
            "layers": int(runtime_inputs.layers),
            "iterations": int(runtime_inputs.iterations),
            "restarts": int(runtime_inputs.restarts),
        }
        optimizer.qaoa_limited_exact_probabilities = True
        optimizer.qaoa_max_qubits_allowed = int(max_qubits)
        logs.append("Mocked qaoa_limited execution for policy boundary test.")
        return optimizer, logs

    monkeypatch.setattr(main_module, "run_qaoa_sim", _mock_qaoa_lightning)
    response = _post_run(
        headers={"X-API-Key": INTERNAL_POWER_KEY},
        mode="qaoa_limited",
        response_level="compact",
        qaoa_p="7",
        qaoa_maxiter="10",
        qaoa_multistart_restarts="1",
    )
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["status"] == "completed"
    assert payload["license"]["usage_level"] == "internal_power"
    assert payload["mode"] == "qaoa_lightning_sim"
    assert payload["binary_variables"] == 14


def test_internal_power_qaoa_limited_rejects_above_24_qubits():
    usage_levels = load_usage_config()["usage_levels"]
    context = UsageContext(
        usage_level_name="internal_power",
        usage_level=usage_levels["internal_power"],
        key_record={"key_id": INTERNAL_POWER_KEY_ID},
        authenticated=True,
    )
    optimizer = SimpleNamespace(
        n=25,
        qaoa_p=1,
        qaoa_maxiter=2,
        qaoa_multistart_restarts=1,
        classical_results=None,
    )

    with pytest.raises(ApiError) as exc_info:
        validate_problem_policy(context, optimizer, "qaoa_limited", {})

    error = exc_info.value
    assert error.status_code == 403
    assert error.code == "qubit_limit_exceeded"
    assert error.details["license_max_qubits"] == 24


def test_qaoa_limited_execution_error_is_controlled_and_not_consumed(monkeypatch, tmp_path):
    import app.main as main_module

    _enable_temp_ledger(monkeypatch, tmp_path)
    workbook_path = _limited_workbook(tmp_path, qubits=3)

    def _raise_qaoa_error(*_args, **_kwargs):
        raise main_module.QAOAExecutionError("unit-test qaoa failure")

    monkeypatch.setattr(main_module, "run_qaoa_sim", _raise_qaoa_error)
    response = _post_run_file(
        workbook_path,
        headers={"X-API-Key": "TESTER-123"},
        mode="qaoa_limited",
        qaoa_p="1",
        qaoa_maxiter="2",
        qaoa_multistart_restarts="1",
    )
    payload = response.get_json()
    status = app.test_client().get("/license-status", headers={"X-API-Key": "TESTER-123"}).get_json()
    summary = app.test_client().get("/ledger-summary").get_json()

    assert response.status_code == 500
    assert payload["error"]["code"] == "qaoa_execution_error"
    assert payload["error"]["details"]["run_status"] == "failed"
    assert status["used_runs"] == 0
    assert status["remaining_runs"] == 1000
    assert summary["failed_runs"] == 1
    assert summary["consumed_runs"] == 0


def test_local_ledger_success_increments_used_runs(monkeypatch, tmp_path):
    ledger_path = _enable_temp_ledger(monkeypatch, tmp_path)
    before = app.test_client().get("/license-status", headers={"X-API-Key": "DEMO-123"}).get_json()
    response = _post_run(headers={"X-API-Key": "DEMO-123"}, mode="classical_only")
    payload = response.get_json()
    after = app.test_client().get("/license-status", headers={"X-API-Key": "DEMO-123"}).get_json()
    summary = app.test_client().get("/ledger-summary").get_json()

    assert response.status_code == 200
    assert before["used_runs"] == 0
    assert payload["license"]["used_runs"] == 1
    assert payload["license"]["remaining_runs"] == 999
    assert after["used_runs"] == 1
    assert after["remaining_runs"] == 999
    assert summary["completed_runs"] == 1
    assert summary["consumed_runs"] == 1
    assert summary["runs_by_key_id"][DEMO_KEY_ID] == 1

    ledger_record = json.loads(ledger_path.read_text(encoding="utf-8"))["runs"][0]
    assert ledger_record["status"] == "completed"
    assert ledger_record["actual_runtime_sec"] is not None
    assert ledger_record["estimated_runtime_sec"] >= 1.0
    assert ledger_record["runtime_ratio"] is not None
    assert ledger_record["n_qubits"] == 14
    assert ledger_record["mode"] == "classical_only"
    assert ledger_record["layers"] >= 1
    assert ledger_record["iterations"] >= 1
    assert ledger_record["restarts"] >= 1
    assert 0 < ledger_record["candidate_count"] <= 2**14


def test_run_limit_exceeded_returns_403(monkeypatch, tmp_path):
    ledger_path = _enable_temp_ledger(monkeypatch, tmp_path)
    ledger_path.write_text(
        json.dumps(
            {
                "runs": [
                    {
                        "run_id": f"seed-{idx}",
                        "status": "completed",
                        "key_id": DEMO_KEY_ID,
                        "usage_level": "qualified_demo",
                        "consumed_run": True,
                    }
                    for idx in range(1000)
                ]
            }
        ),
        encoding="utf-8",
    )

    response = _post_run(headers={"X-API-Key": "DEMO-123"}, mode="classical_only")
    payload = response.get_json()

    assert response.status_code == 403
    assert payload["error"]["code"] == "run_limit_exceeded"
    assert payload["license"]["remaining_runs"] == 0


def test_ledger_summary_only_local_dev(monkeypatch, tmp_path):
    _enable_temp_ledger(monkeypatch, tmp_path)
    response = app.test_client().get("/ledger-summary")
    payload = response.get_json()

    assert response.status_code == 200
    assert "key_hash" not in str(payload)
    assert "DEMO-123" not in str(payload)

    monkeypatch.delenv("QAOA_RQP_LOCAL_DEV", raising=False)
    disabled_response = app.test_client().get("/ledger-summary")
    assert disabled_response.status_code == 403


def test_ledger_summary_not_registered_in_production_mode(monkeypatch):
    monkeypatch.delenv("QAOA_RQP_LOCAL_DEV", raising=False)
    monkeypatch.setenv("KEY_HASH_SECRET", "unit-test-secret")
    monkeypatch.setenv("QAOA_JOB_BUCKET", "unit-test-job-bucket")
    production_app = create_app()

    response = production_app.test_client().get("/ledger-summary")
    payload = response.get_json()

    assert response.status_code == 404
    assert payload["error"]["code"] == "not_found"


def test_upload_excel_demo_still_uses_key_gate():
    response = app.test_client().post(
        "/upload-excel-demo",
        headers={"X-API-Key": "DEMO-123"},
        data=_xlsx_upload_data(),
        content_type="multipart/form-data",
    )
    payload = response.get_json()

    assert response.status_code == 200
    assert payload["status"] == "ok"
    assert "Assets" in payload["input_sheet_names"]
    assert "Settings" in payload["input_sheet_names"]
