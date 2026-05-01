#!/usr/bin/env python3
"""
QAOA Optimizer Core - Extracted from Jupyter notebook
Handles all quantum and classical optimization logic
"""

import random
import time
import numpy as np
import pandas as pd
from pathlib import Path
from typing import Callable, Optional, Dict, List, Any
import heapq

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from scipy.optimize import minimize

try:
    import yfinance as yf
except ImportError:
    yf = None

try:
    import pennylane as qml
except ImportError:
    qml = None


class OptimizationError(Exception):
    """Custom exception for optimization errors."""
    pass


class QAOAOptimizer:
    """Main optimizer class integrating classical and quantum approaches."""
    
    def __init__(self, 
                 xlsx_path: str,
                 refresh_data: bool = False,
                 enable_qaoa: bool = True,
                 enable_classical: bool = True,
                 qaoa_p: int = 1,
                 qaoa_maxiter: int = 60,
                 qaoa_shots: int = 4096,
                 progress_callback: Optional[Callable] = None,
                 log_callback: Optional[Callable] = None,
                 stop_check: Optional[Callable] = None):
        
        self.xlsx_path = Path(xlsx_path)
        self.refresh_data = refresh_data
        self.enable_qaoa = enable_qaoa
        self.enable_classical = enable_classical
        self.qaoa_p = qaoa_p
        self.qaoa_maxiter = qaoa_maxiter
        self.qaoa_shots = qaoa_shots
        
        self.progress_callback = progress_callback or (lambda msg, p: None)
        self.log_callback = log_callback or (lambda msg: None)
        self.stop_check = stop_check or (lambda: False)
        
        # State
        self.assets_df = None
        self.settings = {}
        self.options_df = None
        self.opt_tickers = None
        self.opt_cost = None
        self.opt_ret = None
        self.Sigma_opt = None
        self.Q = None
        self.constant = 0.0
        self.n = 0
        
        self.classical_results = pd.DataFrame()
        self.samples_df = pd.DataFrame()
        self.overview_df = pd.DataFrame()
        self.portfolios_df = pd.DataFrame()
        self.history_df = pd.DataFrame()
        
        self.best_gammas = np.array([])
        self.best_betas = np.array([])
        self.qaoa_total_states = 0
        self.qaoa_nonzero_states = 0
        self.assets_header_row = None
        self.settings_header_row = None
        
    def _log(self, message: str):
        """Log message to callback."""
        self.log_callback(message)
        
    def _progress(self, message: str, progress: float = None):
        """Update progress to callback."""
        self.progress_callback(message, progress)
        
    def _should_stop(self) -> bool:
        """Check if optimization should stop."""
        return self.stop_check()
        
    def load_input(self):
        """Load Excel input file."""
        if not self.xlsx_path.exists():
            raise OptimizationError(f"Excel file not found: {self.xlsx_path}")
            
        self._log(f"Loading {self.xlsx_path.name}...")
        
        xls = pd.ExcelFile(self.xlsx_path)
        self.assets_header_row = self._detect_header_row(xls, "Assets", ("ticker",))
        self.settings_header_row = self._detect_header_row(xls, "Settings", ("key", "value"))

        self.assets_df = pd.read_excel(xls, sheet_name="Assets", header=self.assets_header_row)
        settings_df = pd.read_excel(xls, sheet_name="Settings", header=self.settings_header_row)

        self.assets_df = self._standardize_assets_columns(self.assets_df)
        settings_df = self._standardize_settings_columns(settings_df)
        self.settings = dict(zip(settings_df["Key"], settings_df["Value"]))
        
        self._log(f"Loaded {len(self.assets_df)} assets")
        self._log(f"Loaded {len(self.settings)} settings")
        
    def refresh_market_data(self):
        """Refresh market data from Yahoo Finance."""
        if not yf:
            raise OptimizationError("yfinance not installed. Cannot refresh market data.")
            
        self._log("Refreshing market data from Yahoo Finance...")
        
        option_rows = self.assets_df.loc[self.assets_df["Ticker"].notna()].copy()
        if "Allowed" in option_rows.columns:
            option_rows = option_rows.loc[option_rows["Allowed"].fillna(1).astype(int) == 1].copy()
            
        tickers = list(dict.fromkeys(option_rows["Ticker"].astype(str).tolist()))
        if len(tickers) == 0:
            raise OptimizationError("No tickers available to refresh.")
            
        self._log(f"Downloading data for {len(tickers)} tickers...")
        prices = yf.download(
            tickers=tickers,
            period="12mo",
            interval="1d",
            auto_adjust=True,
            progress=False,
        )["Close"]
        
        if isinstance(prices, pd.Series):
            prices = prices.to_frame()
            
        prices = prices.dropna(how="all").ffill().dropna()
        rets = prices.pct_change().dropna()
        
        total_return_12m = prices.iloc[-1] / prices.iloc[0] - 1
        ann_vol = rets.std() * np.sqrt(252)
        latest_price = prices.iloc[-1]
        
        self._log("Updating Excel file with new data...")
        
        wb = load_workbook(self.xlsx_path)
        sh_assets = wb["Assets"]
        
        header_row = (self.assets_header_row or 0) + 1
        data_start_row = header_row + 1
        asset_headers = [sh_assets.cell(header_row, c).value for c in range(1, sh_assets.max_column + 1)]
        col_idx = self._build_column_index(asset_headers, self._asset_column_aliases())
        
        for r_idx in range(data_start_row, sh_assets.max_row + 1):
            ticker = sh_assets.cell(r_idx, col_idx["Ticker"]).value
            if not ticker or ticker not in latest_price.index:
                continue

            latest = float(latest_price[ticker])

            if "Current Price (USD)" in col_idx:
                sh_assets.cell(r_idx, col_idx["Current Price (USD)"], latest)

            if "Approx Cost USD" in col_idx:
                approx_cost = latest
                if "Shares" in col_idx:
                    shares = sh_assets.cell(r_idx, col_idx["Shares"]).value
                    try:
                        approx_cost = float(shares) * latest
                    except (TypeError, ValueError):
                        approx_cost = latest
                sh_assets.cell(r_idx, col_idx["Approx Cost USD"], approx_cost)

            if "Expected Return Proxy" in col_idx:
                sh_assets.cell(r_idx, col_idx["Expected Return Proxy"], float(total_return_12m[ticker]))

            if "Annual Volatility" in col_idx:
                sh_assets.cell(r_idx, col_idx["Annual Volatility"], float(ann_vol[ticker]))
            
        wb.save(self.xlsx_path)
        self._log("Market data updated successfully")
        
        # Reload
        self.load_input()
        
    def build_qubo(self):
        """Build QUBO optimization problem."""
        self._log("Building QUBO problem...")
        
        # Extract settings
        budget_usd = self._setting_float("budget_usd", 1_000_000.0)
        risk_free = self._setting_float("risk_free_rate_annual", 0.04)
        lambda_budget = self._setting_float("lambda_budget", 50.0)
        lambda_variance = self._setting_float("lambda_variance", 6.0)
        rng_seed = self._setting_int("rng_seed", 42)
        
        random.seed(rng_seed)
        np.random.seed(rng_seed)
        
        # Parse options
        options_df = self.assets_df.loc[self.assets_df["Ticker"].notna()].copy()
        options_df["Ticker"] = options_df["Ticker"].astype(str)
        
        if "Allowed" in options_df.columns:
            options_df = options_df.loc[options_df["Allowed"].fillna(1).astype(int) == 1].copy()
            
        required_cols = ["Ticker", "Approx Cost USD", "Expected Return Proxy", "Annual Volatility"]
        missing = [c for c in required_cols if c not in options_df.columns]
        if missing:
            raise OptimizationError(
                "Assets sheet is missing required columns: "
                f"{missing}. Provide 'Approx Cost USD' directly, or include "
                "'Current Price (USD)' so the app can derive it."
            )
            
        for col in ["Approx Cost USD", "Expected Return Proxy", "Annual Volatility"]:
            options_df[col] = pd.to_numeric(options_df[col], errors="coerce")
            
        if options_df["Approx Cost USD"].isna().any():
            raise OptimizationError("Approx Cost USD has NaN values")
            
        if (options_df["Approx Cost USD"] <= 0).any():
            raise OptimizationError("Approx Cost USD must be > 0")
            
        if "decision_id" not in options_df.columns:
            options_df["decision_id"] = [f"{t}_opt{i+1}" for i, t in enumerate(options_df["Ticker"])]
            
        self.options_df = options_df.reset_index(drop=True)
        self.n = len(self.options_df)
        
        self._log(f"Problem size: {self.n} decision variables")
        
        # Extract vectors
        self.opt_tickers = self.options_df["Ticker"].values
        self.opt_cost = self.options_df["Approx Cost USD"].values
        self.opt_ret = self.options_df["Expected Return Proxy"].values
        
        # Build covariance matrix (simplified - use diagonal)
        vols = self.options_df["Annual Volatility"].values
        self.Sigma_opt = np.diag(vols ** 2)
        
        # Build QUBO
        Q = np.zeros((self.n, self.n))
        
        # Return term (negative, to maximize)
        for i in range(self.n):
            Q[i, i] -= lambda_variance * self.opt_ret[i]
            
        # Risk term (variance penalty)
        for i in range(self.n):
            for j in range(self.n):
                Q[i, j] += lambda_variance * self.Sigma_opt[i, j]
                
        # Budget term
        cost_sum = np.sum(self.opt_cost)
        for i in range(self.n):
            Q[i, i] += lambda_budget * (2 * self.opt_cost[i] - budget_usd)
        for i in range(self.n):
            for j in range(i+1, self.n):
                Q[i, j] += 2 * lambda_budget * self.opt_cost[i]
                
        self.Q = Q
        self.constant = lambda_budget * budget_usd ** 2
        
        self._log(f"QUBO matrix size: {self.Q.shape}")
        
    def run_classical_search(self):
        """Run classical search algorithms."""
        self._log("Running classical search...")
        
        max_qubits = self._setting_int("classical_max_qubits_allowed", 30)
        if self.n > max_qubits:
            self._log(f"Problem size ({self.n}) exceeds classical limit ({max_qubits}), skipping")
            return
            
        # Simple random search
        num_samples = self._setting_int("classical_random_search_samples", 8000)
        candidates = []
        
        for _ in range(num_samples):
            if self._should_stop():
                break
            x = np.random.randint(0, 2, self.n)
            energy = self._qubo_value(x)
            candidates.append({
                "bitstring": "".join(map(str, x)),
                "source": "classical_random",
                "qubo_value": energy
            })
            
        self.classical_results = pd.DataFrame(candidates)
        self._log(f"Classical search generated {len(self.classical_results)} candidates")
        
    def run_qaoa(self):
        """Run QAOA optimization."""
        if not qml:
            raise OptimizationError("PennyLane not installed. Cannot run QAOA.")
            
        self._log(f"Running QAOA (p={self.qaoa_p})...")
        
        if self.n > self._setting_int("qaoa_max_qubits_allowed", 24):
            self._log(f"Problem size exceeds QAOA limit")
            return
            
        # Convert QUBO to Ising
        h, J, ising_offset = self._qubo_to_ising(self.Q, self.constant)
        
        # Build Hamiltonian
        coeffs = []
        ops = []
        for i, coeff in enumerate(h):
            if abs(coeff) > 1e-12:
                coeffs.append(float(coeff))
                ops.append(qml.PauliZ(i))
        for (i, j), coeff in J.items():
            if abs(coeff) > 1e-12:
                coeffs.append(float(coeff))
                ops.append(qml.PauliZ(i) @ qml.PauliZ(j))
                
        if not coeffs:
            coeffs = [0.0]
            ops = [qml.Identity(0)]
            
        H = qml.Hamiltonian(coeffs, ops)
        
        exact_mode = self.n <= self._setting_int("qaoa_exact_probability_max_qubits", 20)
        dev = qml.device("lightning.qubit", wires=self.n, shots=None if exact_mode else self.qaoa_shots)
        
        def apply_qaoa_layer(gamma, beta):
            for i, coeff in enumerate(h):
                if abs(coeff) > 1e-12:
                    qml.RZ(2 * gamma * coeff, wires=i)
            for (i, j), coeff in J.items():
                if abs(coeff) > 1e-12:
                    qml.CNOT(wires=[i, j])
                    qml.RZ(2 * gamma * coeff, wires=j)
                    qml.CNOT(wires=[i, j])
            for w in range(self.n):
                qml.RX(2 * beta, wires=w)
                
        @qml.qnode(dev, interface="autograd")
        def qaoa_energy(params):
            gammas = params[:self.qaoa_p]
            betas = params[self.qaoa_p:]
            for w in range(self.n):
                qml.Hadamard(wires=w)
            for layer in range(self.qaoa_p):
                apply_qaoa_layer(gammas[layer], betas[layer])
            return qml.expval(H)
            
        @qml.qnode(dev, interface="autograd")
        def qaoa_probs(params):
            gammas = params[:self.qaoa_p]
            betas = params[self.qaoa_p:]
            for w in range(self.n):
                qml.Hadamard(wires=w)
            for layer in range(self.qaoa_p):
                apply_qaoa_layer(gammas[layer], betas[layer])
            return qml.probs(wires=range(self.n))
            
        @qml.qnode(dev, interface="autograd")
        def qaoa_samples(params):
            gammas = params[:self.qaoa_p]
            betas = params[self.qaoa_p:]
            for w in range(self.n):
                qml.Hadamard(wires=w)
            for layer in range(self.qaoa_p):
                apply_qaoa_layer(gammas[layer], betas[layer])
            return qml.sample(wires=range(self.n))
            
        # Optimize
        progress_rows = []
        start_time = time.time()
        
        def objective_logged(x):
            energy = float(qaoa_energy(x))
            elapsed = time.time() - start_time
            current_best = min([energy] + [r["energy"] for r in progress_rows]) if progress_rows else energy
            
            row = {"iteration": len(progress_rows) + 1, "energy": energy, "best_energy": current_best, "elapsed_sec": elapsed}
            progress_rows.append(row)
            
            if len(progress_rows) % 5 == 0:
                self._log(f"QAOA iter {len(progress_rows)}: energy={energy:.6f}, best={current_best:.6f}")
                
            return energy
            
        gamma_init = np.linspace(0.30, 0.60, self.qaoa_p)
        beta_init = np.linspace(0.40, 0.20, self.qaoa_p)
        x0 = np.concatenate([gamma_init, beta_init]).astype(float)
        
        self._log("Optimizing QAOA parameters...")
        res = minimize(objective_logged, x0=x0, method="COBYLA", options={"maxiter": self.qaoa_maxiter})
        
        self.best_gammas = np.array(res.x[:self.qaoa_p])
        self.best_betas = np.array(res.x[self.qaoa_p:])
        self.history_df = pd.DataFrame(progress_rows)
        
        self._log("Extracting QAOA samples...")
        
        q_rows = []
        params_opt = np.concatenate([self.best_gammas, self.best_betas])
        
        if exact_mode:
            probs = np.array(qaoa_probs(params_opt))
            self.qaoa_total_states = int(len(probs))
            self.qaoa_nonzero_states = int(np.count_nonzero(probs > 0))
            
            for idx, prob in enumerate(probs):
                if prob < 1e-12:
                    continue
                bitstring = format(idx, f"0{self.n}b")[::-1]
                bits = np.array(list(map(int, bitstring)), dtype=int)
                q_rows.append({
                    "bitstring": "".join(map(str, bits.astype(int))),
                    "source": f"qaoa_p{self.qaoa_p}",
                    "probability": float(prob),
                    "qubo_value": self._qubo_value(bits)
                })
        else:
            raw_samples = np.asarray(qaoa_samples(params_opt))
            if raw_samples.ndim == 1:
                raw_samples = raw_samples.reshape(1, -1)
                
            sample_counts = {}
            for sample in raw_samples:
                bits = np.asarray(sample, dtype=int).reshape(-1)
                bitstring = "".join(map(str, bits.astype(int)))
                sample_counts[bitstring] = sample_counts.get(bitstring, 0) + 1
                
            total_shots = sum(sample_counts.values())
            self.qaoa_total_states = int(total_shots)
            self.qaoa_nonzero_states = int(len(sample_counts))
            
            for bitstring, count in sample_counts.items():
                bits = np.array(list(map(int, bitstring)), dtype=int)
                q_rows.append({
                    "bitstring": bitstring,
                    "source": f"qaoa_p{self.qaoa_p}",
                    "probability": float(count / total_shots),
                    "qubo_value": self._qubo_value(bits)
                })
                
        self.samples_df = pd.DataFrame(q_rows)
        self._log(f"QAOA generated {len(self.samples_df)} unique solutions")
        
    def generate_results(self):
        """Generate result summaries."""
        self._log("Generating results summary...")
        
        # Combine and sort results
        frames = []
        if len(self.classical_results) > 0:
            frames.append(self.classical_results)
        if len(self.samples_df) > 0:
            frames.append(self.samples_df)
            
        if frames:
            combined = pd.concat(frames, ignore_index=True)
            combined = combined.sort_values("qubo_value").reset_index(drop=True)
            self.overview_df = combined.head(self._setting_int("top_n_export", 20)).copy()
        else:
            self.overview_df = pd.DataFrame()
            
        self._log(f"Generated overview with {len(self.overview_df)} results")
        
    def write_results(self):
        """Write results back to Excel."""
        self._log("Writing results to Excel...")
        
        wb = load_workbook(self.xlsx_path)
        
        # Remove old result sheets
        for name in ["Results_Summary", "Results_Overview", "Classical_Candidates", "QAOA_Samples"]:
            if name in wb.sheetnames:
                wb.remove(wb[name])
                
        # Create new sheets
        dark = PatternFill("solid", fgColor="1F4E78")
        white_bold = Font(color="FFFFFF", bold=True)
        
        # Summary sheet
        summary_ws = wb.create_sheet("Results_Summary")
        self._write_summary_sheet(summary_ws, dark, white_bold)
        
        # Overview sheet
        if len(self.overview_df) > 0:
            overview_ws = wb.create_sheet("Results_Overview")
            self._write_dataframe_to_sheet(overview_ws, self.overview_df, dark, white_bold)
            
        # Classical results
        if len(self.classical_results) > 0:
            classical_ws = wb.create_sheet("Classical_Candidates")
            self._write_dataframe_to_sheet(classical_ws, self.classical_results, dark, white_bold)
            
        # QAOA results
        if len(self.samples_df) > 0:
            qaoa_ws = wb.create_sheet("QAOA_Samples")
            self._write_dataframe_to_sheet(qaoa_ws, self.samples_df, dark, white_bold)
            
        wb.save(self.xlsx_path)
        self._log("Results saved successfully")
        
    # Helper methods
    def _setting_value(self, key: str, default=None):
        """Get setting value."""
        return self.settings.get(key, default) if key in self.settings and pd.notna(self.settings.get(key)) else default
        
    def _setting_int(self, key: str, default: int = 0) -> int:
        val = self._setting_value(key, default)
        try:
            return int(float(val))
        except:
            return int(default)
            
    def _setting_float(self, key: str, default: float = 0.0) -> float:
        val = self._setting_value(key, default)
        try:
            return float(val)
        except:
            return float(default)
            
    def _qubo_value(self, x: np.ndarray) -> float:
        """Compute QUBO value for solution."""
        return float(x @ self.Q @ x + self.constant)
        
    def _qubo_to_ising(self, Q, constant=0.0):
        """Convert QUBO to Ising model."""
        n = Q.shape[0]
        h = np.zeros(n, dtype=float)
        J = {}
        offset = float(constant)
        
        for i in range(n):
            qii = Q[i, i]
            offset += qii / 2
            h[i] += -qii / 2
            
        for i in range(n):
            for j in range(i + 1, n):
                qij = Q[i, j]
                if abs(qij) > 1e-12:
                    offset += qij / 4
                    h[i] += -qij / 4
                    h[j] += -qij / 4
                    J[(i, j)] = qij / 4
                    
        return h, J, offset
        
    def _write_dataframe_to_sheet(self, ws, df: pd.DataFrame, dark_fill, white_bold):
        """Write dataframe to worksheet."""
        for j, col in enumerate(df.columns, start=1):
            ws.cell(1, j, col)
            ws.cell(1, j).fill = dark_fill
            ws.cell(1, j).font = white_bold
            
        for i, row in enumerate(df.itertuples(index=False), start=2):
            for j, val in enumerate(row, start=1):
                ws.cell(i, j, val)
                
    def _write_summary_sheet(self, ws, dark_fill, white_bold):
        """Write summary information."""
        ws["A1"] = "Optimization Results Summary"
        ws["A1"].fill = dark_fill
        ws["A1"].font = white_bold
        
        items = [
            ("Classical results count", len(self.classical_results)),
            ("QAOA results count", len(self.samples_df)),
            ("Top-N exported", len(self.overview_df)),
            ("Decision variables", self.n),
            ("QAOA enabled", int(self.enable_qaoa)),
            ("Classical enabled", int(self.enable_classical)),
            ("QAOA p", self.qaoa_p),
            ("QAOA total states", self.qaoa_total_states),
            ("QAOA nonzero states", self.qaoa_nonzero_states),
        ]
        
        for i, (label, value) in enumerate(items, start=2):
            ws.cell(i, 1, label)
            ws.cell(i, 2, value)

    @staticmethod
    def _normalize_header(value) -> str:
        """Convert worksheet header labels into a stable comparison format."""
        if value is None:
            return ""
        return " ".join(str(value).replace("\n", " ").split()).strip().lower()

    def _detect_header_row(self, xls, sheet_name: str, required_headers, max_rows: int = 8) -> int:
        """Detect whether a sheet uses row 1 or row 2 for headers."""
        preview = pd.read_excel(xls, sheet_name=sheet_name, header=None, nrows=max_rows)
        required = {self._normalize_header(value) for value in required_headers}

        for idx in range(len(preview)):
            labels = {self._normalize_header(value) for value in preview.iloc[idx].tolist()}
            labels.discard("")
            if required.issubset(labels):
                return int(idx)

        raise OptimizationError(
            f"Could not find the header row in sheet '{sheet_name}'. "
            "Use either headers on the first row, or a title row followed by headers."
        )

    @staticmethod
    def _build_column_index(headers, aliases):
        """Map workbook headers to canonical names."""
        col_idx = {}
        for idx, header in enumerate(headers, start=1):
            key = QAOAOptimizer._normalize_header(header)
            canonical = aliases.get(key)
            if canonical:
                col_idx[canonical] = idx
        return col_idx

    @staticmethod
    def _rename_columns(df: pd.DataFrame, aliases) -> pd.DataFrame:
        """Rename dataframe columns using normalized aliases."""
        rename_map = {}
        for col in df.columns:
            key = QAOAOptimizer._normalize_header(col)
            canonical = aliases.get(key)
            if canonical and col != canonical:
                rename_map[col] = canonical
        return df.rename(columns=rename_map)

    @staticmethod
    def _asset_column_aliases():
        """Aliases accepted for workbook asset columns."""
        return {
            "decision_id": "decision_id",
            "ticker": "Ticker",
            "company": "Company",
            "shares": "Shares",
            "current price (usd)": "Current Price (USD)",
            "current price usd": "Current Price (USD)",
            "approx cost usd": "Approx Cost USD",
            "expected return proxy": "Expected Return Proxy",
            "annual volatility": "Annual Volatility",
            "allowed": "Allowed",
        }

    def _standardize_assets_columns(self, assets_df: pd.DataFrame) -> pd.DataFrame:
        """Accept both the sample workbook format and the simplified quickstart format."""
        assets_df = self._rename_columns(assets_df, self._asset_column_aliases())

        if "Approx Cost USD" not in assets_df.columns and "Current Price (USD)" in assets_df.columns:
            if "Shares" in assets_df.columns:
                shares = pd.to_numeric(assets_df["Shares"], errors="coerce")
                prices = pd.to_numeric(assets_df["Current Price (USD)"], errors="coerce")
                assets_df["Approx Cost USD"] = shares * prices
            else:
                assets_df["Approx Cost USD"] = assets_df["Current Price (USD)"]
        elif "Approx Cost USD" in assets_df.columns and "Current Price (USD)" in assets_df.columns and "Shares" in assets_df.columns:
            approx_cost = pd.to_numeric(assets_df["Approx Cost USD"], errors="coerce")
            shares = pd.to_numeric(assets_df["Shares"], errors="coerce")
            prices = pd.to_numeric(assets_df["Current Price (USD)"], errors="coerce")
            assets_df["Approx Cost USD"] = approx_cost.fillna(shares * prices)

        return assets_df

    def _standardize_settings_columns(self, settings_df: pd.DataFrame) -> pd.DataFrame:
        """Normalize settings columns so both supported header layouts behave the same."""
        aliases = {
            "key": "Key",
            "value": "Value",
        }
        return self._rename_columns(settings_df, aliases)
