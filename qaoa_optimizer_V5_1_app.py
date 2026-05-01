#!/usr/bin/env python3
"""Tkinter GUI for the Version 5.1 QAOA optimizer.

V5.1 keeps the V5 GUI layout and points it at the V5.1 core. The app-level
edits are intentionally narrow:

* Imports now point at ``qaoa_optimizer_V5_1_core``.
* The driver class is renamed ``OptimizerGUIV51``.
* The banner / summary strings advertise Version 5.1.
* Internal type hints reference ``QAOAOptimizerV51``.

All V5.1-specific algorithmic fixes live in ``qaoa_optimizer_V5_1_core.py``
and are driven from the Excel workbook's ``Settings`` sheet, so no new UI
controls were added.
"""

import io
import threading
import tkinter as tk
import traceback
from pathlib import Path
from tkinter import filedialog, messagebox, scrolledtext, ttk

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from PIL import Image, ImageTk
from types import SimpleNamespace

from qaoa_optimizer_V5_1_core import OptimizationCancelled, OptimizationError, QAOAOptimizerV51


class OptimizerGUIV51:
    def __init__(self, root):
        self.root = root
        self.root.title("QAOA RQP - Rapid Quantum Prototyping - by qubit-lab.ch")
        self.root.geometry("1180x920")
        self.root.minsize(1040, 820)

        self.optimizer = None
        self.optimizer_thread = None
        self.is_running = False
        self.logo_image = None
        self.result_images = []
        self.results_window = None
        self.qaoa_shots_numeric_value = 4096
        self.qaoa_shots_exact_mode = False

        self.colors = {
            "bg": "#050816",
            "panel": "#0A1022",
            "panel_alt": "#101933",
            "text": "#F5F9FF",
            "muted": "#B7C6E0",
            "accent": "#1EC8FF",
            "accent_alt": "#1B8DFF",
            "border": "#1E335F",
            "button_text": "#03111D",
            "entry_bg": "#091224",
        }

        self._setup_styles()
        self._setup_ui()

    def _setup_styles(self):
        self.root.configure(bg=self.colors["bg"])
        style = ttk.Style()
        style.theme_use("clam")

        style.configure(
            ".",
            background=self.colors["bg"],
            foreground=self.colors["text"],
            fieldbackground=self.colors["entry_bg"],
        )
        style.configure("App.TFrame", background=self.colors["bg"])
        style.configure("Panel.TFrame", background=self.colors["panel"])
        style.configure(
            "Card.TLabelframe",
            background=self.colors["panel"],
            bordercolor=self.colors["border"],
            relief="solid",
            borderwidth=1,
            padding=10,
        )
        style.configure(
            "Card.TLabelframe.Label",
            background=self.colors["panel"],
            foreground=self.colors["accent"],
            font=("Helvetica", 11, "bold"),
        )
        style.configure(
            "Title.TLabel",
            background=self.colors["bg"],
            foreground=self.colors["accent"],
            font=("Helvetica", 17, "bold"),
        )
        style.configure(
            "Subtitle.TLabel",
            background=self.colors["bg"],
            foreground=self.colors["text"],
            font=("Helvetica", 11),
        )
        style.configure(
            "Body.TLabel",
            background=self.colors["panel"],
            foreground=self.colors["muted"],
            font=("Helvetica", 10),
        )
        style.configure(
            "TLabel",
            background=self.colors["panel"],
            foreground=self.colors["text"],
        )
        style.configure(
            "TCheckbutton",
            background=self.colors["panel"],
            foreground=self.colors["text"],
        )
        style.map(
            "TCheckbutton",
            background=[("active", self.colors["panel"])],
            foreground=[("disabled", "#6D7F9E")],
        )
        style.configure(
            "TEntry",
            fieldbackground=self.colors["entry_bg"],
            foreground=self.colors["text"],
            insertcolor=self.colors["text"],
            bordercolor=self.colors["border"],
            lightcolor=self.colors["border"],
            darkcolor=self.colors["border"],
        )
        style.configure(
            "TSpinbox",
            fieldbackground=self.colors["entry_bg"],
            foreground=self.colors["text"],
            arrowsize=14,
            arrowcolor=self.colors["accent"],
            bordercolor=self.colors["border"],
            lightcolor=self.colors["border"],
            darkcolor=self.colors["border"],
        )
        style.configure(
            "Primary.TButton",
            background=self.colors["accent"],
            foreground=self.colors["button_text"],
            borderwidth=0,
            focusthickness=0,
            padding=(16, 10),
            font=("Helvetica", 10, "bold"),
        )
        style.map(
            "Primary.TButton",
            background=[("active", "#56D8FF"), ("disabled", "#6D7F9E")],
            foreground=[("disabled", "#E3ECF7")],
        )
        style.configure(
            "Secondary.TButton",
            background=self.colors["panel_alt"],
            foreground=self.colors["text"],
            bordercolor=self.colors["accent"],
            lightcolor=self.colors["accent"],
            darkcolor=self.colors["accent"],
            borderwidth=1,
            focusthickness=0,
            padding=(16, 10),
            font=("Helvetica", 10, "bold"),
        )
        style.map(
            "Secondary.TButton",
            background=[("active", "#142347"), ("disabled", "#182544")],
            foreground=[("disabled", "#6D7F9E")],
        )
        style.configure(
            "TProgressbar",
            background=self.colors["accent"],
            troughcolor=self.colors["panel_alt"],
            bordercolor=self.colors["border"],
            lightcolor=self.colors["accent"],
            darkcolor=self.colors["accent_alt"],
            thickness=12,
        )

    def _setup_ui(self):
        main_frame = ttk.Frame(self.root, padding="8", style="App.TFrame")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(5, weight=1)

        header_frame = ttk.Frame(main_frame, style="App.TFrame")
        header_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 8))
        header_frame.columnconfigure(1, weight=1)

        logo_label = ttk.Label(header_frame)
        logo_label.grid(row=0, column=0, rowspan=2, sticky=tk.W, padx=(0, 10))
        self._apply_logo(logo_label)

        ttk.Label(
            header_frame,
            text="QAOA RQP - Rapid Quantum Prototyping - by qubit-lab.ch",
            style="Title.TLabel",
        ).grid(row=0, column=1, sticky=tk.W)
        ttk.Label(
            header_frame,
            text="Version 5.1 with scale-aware COBYLA, adaptive p=1 presearch, and canonicalized p=1 warm starts for stronger layerwise clustering",
            style="Subtitle.TLabel",
        ).grid(row=1, column=1, sticky=tk.W)

        file_frame = ttk.LabelFrame(main_frame, text="Input File", style="Card.TLabelframe")
        file_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 8))
        file_frame.columnconfigure(1, weight=1)

        self.file_path_var = tk.StringVar()
        self.file_path_var.trace_add("write", lambda *_: self._on_file_path_changed())
        ttk.Label(file_frame, text="Workbook:").grid(row=0, column=0, sticky=tk.W)
        ttk.Entry(file_frame, textvariable=self.file_path_var).grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(8, 8))
        ttk.Button(file_frame, text="Browse", command=self._browse_file, style="Secondary.TButton").grid(
            row=0, column=2, sticky=tk.E
        )

        controls_frame = ttk.Frame(main_frame, style="App.TFrame")
        controls_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 8))
        controls_frame.columnconfigure(0, weight=5)
        controls_frame.columnconfigure(1, weight=7)

        note_frame = ttk.LabelFrame(controls_frame, text="Run Mode", style="Card.TLabelframe")
        note_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 6))
        note_frame.columnconfigure(0, weight=1)

        self.use_workbook_settings_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            note_frame,
            text="Use workbook settings exactly as defined in the Excel file",
            variable=self.use_workbook_settings_var,
            command=self._toggle_override_state,
        ).grid(row=0, column=0, sticky=tk.W, pady=(0, 8))
        ttk.Label(
            note_frame,
            text="Leave this checked for notebook-equivalent behavior. Uncheck it only to force app overrides.",
            wraplength=760,
            style="Body.TLabel",
        ).grid(row=1, column=0, sticky=tk.W)

        override_frame = ttk.LabelFrame(controls_frame, text="Optional Overrides", style="Card.TLabelframe")
        override_frame.grid(row=0, column=1, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(6, 0))
        override_frame.columnconfigure(1, weight=1)

        self.refresh_data_var = tk.BooleanVar(value=False)
        self.enable_qaoa_var = tk.BooleanVar(value=True)
        self.enable_classical_var = tk.BooleanVar(value=True)
        self.qaoa_p_var = tk.StringVar(value="1")
        self.qaoa_maxiter_var = tk.StringVar(value="60")
        self.qaoa_shots_var = tk.StringVar(value="4096")
        self.lambda_budget_var = tk.StringVar(value="50")
        self.lambda_variance_var = tk.StringVar(value="6")
        self.risk_free_rate_var = tk.StringVar(value="0.04")
        self.rng_seed_var = tk.StringVar(value="42")
        self.qaoa_multistart_restarts_var = tk.StringVar(value="2")
        self.qaoa_layerwise_warm_start_var = tk.BooleanVar(value=True)
        self.qaoa_restart_perturbation_var = tk.StringVar(value="0.05")

        self.override_widgets = []

        for col in range(6):
            override_frame.columnconfigure(col, weight=1 if col % 2 == 1 else 0)

        widget = ttk.Checkbutton(
            override_frame,
            text="Refresh market data from Yahoo Finance",
            variable=self.refresh_data_var,
        )
        widget.grid(row=0, column=0, columnspan=6, sticky=tk.W, pady=2)
        self.override_widgets.append(widget)

        widget = ttk.Checkbutton(
            override_frame,
            text="Enable QAOA",
            variable=self.enable_qaoa_var,
        )
        widget.grid(row=1, column=0, columnspan=3, sticky=tk.W, pady=2)
        self.override_widgets.append(widget)

        widget = ttk.Checkbutton(
            override_frame,
            text="Enable classical search",
            variable=self.enable_classical_var,
        )
        widget.grid(row=1, column=3, columnspan=3, sticky=tk.W, pady=2)
        self.override_widgets.append(widget)

        ttk.Label(override_frame, text="QAOA layers (p):").grid(row=2, column=0, sticky=tk.W, pady=2)
        widget = ttk.Spinbox(override_frame, from_=1, to=5, textvariable=self.qaoa_p_var, width=10)
        widget.grid(row=2, column=1, sticky=tk.W, pady=2, padx=(6, 16))
        self.override_widgets.append(widget)

        ttk.Label(override_frame, text="QAOA max iterations:").grid(row=2, column=2, sticky=tk.W, pady=2)
        widget = ttk.Spinbox(override_frame, from_=1, to=5000, textvariable=self.qaoa_maxiter_var, width=10)
        widget.grid(row=2, column=3, sticky=tk.W, pady=2, padx=(6, 16))
        self.override_widgets.append(widget)

        ttk.Label(override_frame, text="QAOA shots:").grid(row=2, column=4, sticky=tk.W, pady=2)
        widget = ttk.Spinbox(override_frame, from_=1, to=1000000, textvariable=self.qaoa_shots_var, width=12)
        widget.grid(row=2, column=5, sticky=tk.W, pady=2)
        self.qaoa_shots_widget = widget
        self.override_widgets.append(widget)

        ttk.Label(override_frame, text="Budget lambda:").grid(row=3, column=0, sticky=tk.W, pady=2)
        widget = ttk.Entry(override_frame, textvariable=self.lambda_budget_var, width=12)
        widget.grid(row=3, column=1, sticky=tk.W, pady=2, padx=(6, 16))
        self.override_widgets.append(widget)

        ttk.Label(override_frame, text="Risk lambda:").grid(row=3, column=2, sticky=tk.W, pady=2)
        widget = ttk.Entry(override_frame, textvariable=self.lambda_variance_var, width=12)
        widget.grid(row=3, column=3, sticky=tk.W, pady=2, padx=(6, 16))
        self.override_widgets.append(widget)

        ttk.Label(override_frame, text="Risk-free rate:").grid(row=3, column=4, sticky=tk.W, pady=2)
        widget = ttk.Entry(override_frame, textvariable=self.risk_free_rate_var, width=12)
        widget.grid(row=3, column=5, sticky=tk.W, pady=2)
        self.override_widgets.append(widget)

        ttk.Label(override_frame, text="Random seed:").grid(row=4, column=0, sticky=tk.W, pady=2)
        widget = ttk.Entry(override_frame, textvariable=self.rng_seed_var, width=12)
        widget.grid(row=4, column=1, sticky=tk.W, pady=2, padx=(6, 16))
        self.override_widgets.append(widget)

        ttk.Label(override_frame, text="QAOA restarts:").grid(row=4, column=2, sticky=tk.W, pady=2)
        widget = ttk.Spinbox(override_frame, from_=1, to=50, textvariable=self.qaoa_multistart_restarts_var, width=10)
        widget.grid(row=4, column=3, sticky=tk.W, pady=2, padx=(6, 16))
        self.override_widgets.append(widget)

        ttk.Label(override_frame, text="Restart perturbation:").grid(row=4, column=4, sticky=tk.W, pady=2)
        widget = ttk.Entry(override_frame, textvariable=self.qaoa_restart_perturbation_var, width=12)
        widget.grid(row=4, column=5, sticky=tk.W, pady=2)
        self.override_widgets.append(widget)

        widget = ttk.Checkbutton(
            override_frame,
            text="Layerwise warm start",
            variable=self.qaoa_layerwise_warm_start_var,
        )
        widget.grid(row=5, column=0, columnspan=3, sticky=tk.W, pady=2)
        self.override_widgets.append(widget)

        control_frame = ttk.Frame(main_frame, style="App.TFrame")
        control_frame.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=(0, 8))
        self.run_btn = ttk.Button(
            control_frame, text="Run Optimization", command=self._run_optimization, style="Primary.TButton"
        )
        self.run_btn.pack(side=tk.LEFT, padx=(0, 8))
        self.stop_btn = ttk.Button(
            control_frame, text="Stop", command=self._stop_optimization, state=tk.DISABLED, style="Secondary.TButton"
        )
        self.stop_btn.pack(side=tk.LEFT)
        self.show_results_btn = ttk.Button(
            control_frame,
            text="Show Existing Results",
            command=self._show_existing_results,
            state=tk.DISABLED,
            style="Secondary.TButton",
        )
        self.show_results_btn.pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(control_frame, text="Exit", command=self.root.quit, style="Secondary.TButton").pack(side=tk.RIGHT)

        status_frame = ttk.LabelFrame(main_frame, text="Status", style="Card.TLabelframe")
        status_frame.grid(row=4, column=0, sticky=(tk.W, tk.E), pady=(0, 8))
        status_frame.columnconfigure(0, weight=1)
        self.status_label = ttk.Label(status_frame, text="Ready", style="Subtitle.TLabel")
        self.status_label.grid(row=0, column=0, sticky=tk.W, pady=(0, 4))
        self.progress_var = tk.DoubleVar(value=0)
        ttk.Progressbar(status_frame, variable=self.progress_var, maximum=100).grid(
            row=1, column=0, sticky=(tk.W, tk.E)
        )

        log_frame = ttk.LabelFrame(main_frame, text="Execution Log", style="Card.TLabelframe")
        log_frame.grid(row=5, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        self.log_text = scrolledtext.ScrolledText(
            log_frame,
            height=22,
            state=tk.DISABLED,
            bg=self.colors["entry_bg"],
            fg=self.colors["text"],
            insertbackground=self.colors["text"],
            selectbackground=self.colors["accent_alt"],
            font=("Menlo", 9),
            relief=tk.FLAT,
            highlightthickness=1,
            highlightbackground=self.colors["border"],
            highlightcolor=self.colors["accent"],
        )
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        self._toggle_override_state()
        self._refresh_results_button_state()

    def _toggle_override_state(self):
        state = tk.DISABLED if self.use_workbook_settings_var.get() else tk.NORMAL
        for widget in self.override_widgets:
            widget.configure(state=state)
        self._update_qaoa_shots_field_mode()

    def _browse_file(self):
        filename = filedialog.askopenfilename(
            title="Select Excel file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if filename:
            self.file_path_var.set(filename)

    def _on_file_path_changed(self):
        self._refresh_results_button_state()
        self._load_workbook_settings_into_fields()

    def _workbook_has_results(self, xlsx_path: str) -> bool:
        path = Path(xlsx_path)
        if not path.exists():
            return False
        try:
            xls = pd.ExcelFile(path)
            sheets = set(xls.sheet_names)
        except Exception:
            return False
        required = {"Results_Overview", "Solver_Comparison"}
        return required.issubset(sheets)

    def _refresh_results_button_state(self):
        state = tk.NORMAL if self._workbook_has_results(self.file_path_var.get().strip()) else tk.DISABLED
        if self.is_running:
            state = tk.DISABLED
        self.show_results_btn.configure(state=state)

    def _load_workbook_settings_into_fields(self):
        xlsx_path = self.file_path_var.get().strip()
        path = Path(xlsx_path)
        if not path.exists():
            self._update_qaoa_shots_field_mode(force_exact_mode=False)
            return
        try:
            settings_df = pd.read_excel(path, sheet_name="Settings", header=1)
            settings = dict(zip(settings_df["Key"], settings_df["Value"]))
        except Exception:
            self._update_qaoa_shots_field_mode(force_exact_mode=False)
            return

        def set_str(var, key, default):
            value = settings.get(key, default)
            if pd.isna(value):
                value = default
            var.set(str(value))

        def set_bool(var, key, default):
            value = settings.get(key, default)
            if isinstance(value, str):
                value = value.strip().lower() in {"1", "true", "yes", "y", "on"}
            else:
                try:
                    value = bool(int(value))
                except Exception:
                    value = bool(value)
            var.set(value)

        set_bool(self.refresh_data_var, "refresh_market_data", False)
        set_bool(self.enable_qaoa_var, "enable_qaoa", True)
        set_bool(self.enable_classical_var, "enable_classical_search", True)
        set_str(self.qaoa_p_var, "qaoa_p", 1)
        set_str(self.qaoa_maxiter_var, "qaoa_maxiter", 60)
        set_str(self.qaoa_shots_var, "qaoa_shots", 4096)
        self.qaoa_shots_numeric_value = self._parse_int_value(self.qaoa_shots_var.get(), 4096)
        set_str(self.lambda_budget_var, "lambda_budget", 50.0)
        set_str(self.lambda_variance_var, "lambda_variance", 6.0)
        set_str(self.risk_free_rate_var, "risk_free_rate_annual", 0.04)
        set_str(self.rng_seed_var, "rng_seed", 42)
        set_str(self.qaoa_multistart_restarts_var, "qaoa_multistart_restarts", 2)
        set_bool(self.qaoa_layerwise_warm_start_var, "qaoa_layerwise_warm_start", True)
        set_str(self.qaoa_restart_perturbation_var, "qaoa_restart_perturbation", 0.05)
        exact_mode, _, _ = self._detect_exact_mode_from_workbook(path)
        self._update_qaoa_shots_field_mode(force_exact_mode=exact_mode)

    def _parse_int_value(self, raw_value, default):
        try:
            return int(float(str(raw_value).strip()))
        except Exception:
            return int(default)

    def _current_qaoa_shots_value(self):
        raw_value = self.qaoa_shots_var.get().strip()
        try:
            parsed = int(float(raw_value))
            self.qaoa_shots_numeric_value = parsed
            return parsed
        except Exception:
            return int(self.qaoa_shots_numeric_value)

    def _detect_exact_mode_from_workbook(self, xlsx_path: Path):
        try:
            settings_df = pd.read_excel(xlsx_path, sheet_name="Settings", header=1)
            settings = dict(zip(settings_df["Key"], settings_df["Value"]))
            assets_df = pd.read_excel(xlsx_path, sheet_name="Assets", header=1)
        except Exception:
            return False, 0, 20

        option_rows = assets_df.loc[assets_df["Ticker"].notna()].copy() if "Ticker" in assets_df.columns else pd.DataFrame()
        if "Allowed" in option_rows.columns:
            option_rows = option_rows.loc[option_rows["Allowed"].fillna(1).astype(int) == 1].copy()
        n = int(len(option_rows))
        exact_threshold = self._parse_int_value(settings.get("qaoa_exact_probability_max_qubits", 20), 20)
        return n <= exact_threshold and n > 0, n, exact_threshold

    def _update_qaoa_shots_field_mode(self, force_exact_mode=None):
        if force_exact_mode is None:
            xlsx_path = self.file_path_var.get().strip()
            exact_mode = False
            if xlsx_path and Path(xlsx_path).exists():
                exact_mode, _, _ = self._detect_exact_mode_from_workbook(Path(xlsx_path))
        else:
            exact_mode = bool(force_exact_mode)

        current_text = self.qaoa_shots_var.get().strip()
        if exact_mode:
            try:
                self.qaoa_shots_numeric_value = int(float(current_text))
            except Exception:
                pass
            self.qaoa_shots_exact_mode = True
            self.qaoa_shots_var.set("(exact mode)")
        else:
            if self.qaoa_shots_exact_mode or current_text == "(exact mode)":
                self.qaoa_shots_var.set(str(self.qaoa_shots_numeric_value))
            self.qaoa_shots_exact_mode = False

        if hasattr(self, "qaoa_shots_widget"):
            base_state = tk.DISABLED if self.use_workbook_settings_var.get() else tk.NORMAL
            state = tk.DISABLED if self.qaoa_shots_exact_mode else base_state
            self.qaoa_shots_widget.configure(state=state)

    def _write_gui_settings_to_workbook(self, xlsx_path: str):
        wb = load_workbook(xlsx_path)
        if "Settings" not in wb.sheetnames:
            raise ValueError("Workbook does not contain a Settings sheet.")
        ws = wb["Settings"]

        header_row = 2
        key_col = None
        value_col = None
        for col in range(1, ws.max_column + 1):
            header = ws.cell(header_row, col).value
            if header == "Key":
                key_col = col
            elif header == "Value":
                value_col = col
        if key_col is None or value_col is None:
            raise ValueError("Settings sheet must contain Key and Value columns.")

        updates = {
            "refresh_market_data": int(self.refresh_data_var.get()),
            "enable_qaoa": int(self.enable_qaoa_var.get()),
            "enable_classical_search": int(self.enable_classical_var.get()),
            "qaoa_p": int(float(self.qaoa_p_var.get())),
            "qaoa_maxiter": int(float(self.qaoa_maxiter_var.get())),
            "qaoa_shots": self._current_qaoa_shots_value(),
            "lambda_budget": float(self.lambda_budget_var.get()),
            "lambda_variance": float(self.lambda_variance_var.get()),
            "risk_free_rate_annual": float(self.risk_free_rate_var.get()),
            "rng_seed": int(float(self.rng_seed_var.get())),
            "qaoa_multistart_restarts": int(float(self.qaoa_multistart_restarts_var.get())),
            "qaoa_layerwise_warm_start": int(self.qaoa_layerwise_warm_start_var.get()),
            "qaoa_restart_perturbation": float(self.qaoa_restart_perturbation_var.get()),
        }

        found = set()
        for row in range(3, ws.max_row + 1):
            key = ws.cell(row, key_col).value
            if key in updates:
                ws.cell(row, value_col, updates[key])
                found.add(key)

        missing = [key for key in updates if key not in found]
        if missing:
            next_row = ws.max_row + 1
            for key in missing:
                ws.cell(next_row, key_col, key)
                ws.cell(next_row, value_col, updates[key])
                next_row += 1

        wb.save(xlsx_path)

    def _find_logo_path(self):
        candidates = [
            Path("assets/qubit-lab-light-blue-transparent.png"),
            Path("assets/logo_ceramic_transparent.png"),
            Path("assets/qubit-lab-logo-white.png"),
            Path("assets/qubit-lab-logo.png"),
            Path("assets/qubit_lab_logo.png"),
            Path("assets/qubit-lab.ch-logo.png"),
            Path("qubit-lab-logo.png"),
            Path("qubit_lab_logo.png"),
            Path("logo.png"),
        ]
        for candidate in candidates:
            if candidate.exists():
                return candidate
        return None

    def _apply_logo(self, label):
        logo_path = self._find_logo_path()
        if logo_path is None:
            label.configure(text="[qubit-lab.ch logo]", width=18, anchor=tk.CENTER)
            return
        try:
            image = Image.open(logo_path).convert("RGBA")
            target_size = 96
            image.thumbnail((target_size, target_size), Image.Resampling.LANCZOS)
            self.logo_image = ImageTk.PhotoImage(image)
            label.configure(image=self.logo_image, text="")
        except Exception:
            label.configure(text=logo_path.name, width=18, anchor=tk.CENTER)

    @staticmethod
    def _is_main_thread():
        return threading.current_thread() is threading.main_thread()

    def _call_on_ui_thread(self, fn):
        if self._is_main_thread():
            fn()
        else:
            self.root.after(0, fn)

    def _log_message(self, message):
        def update():
            self.log_text.configure(state=tk.NORMAL)
            self.log_text.insert(tk.END, message + "\n")
            self.log_text.see(tk.END)
            self.log_text.configure(state=tk.DISABLED)
            self.root.update_idletasks()

        self._call_on_ui_thread(update)

    def _update_status(self, message, progress=None):
        def update():
            self.status_label.configure(text=message)
            if progress is not None:
                self.progress_var.set(progress)
            self.root.update_idletasks()

        self._call_on_ui_thread(update)

    def _results_output_dir(self, xlsx_path: str) -> Path:
        workbook = Path(xlsx_path)
        out_dir = workbook.parent / f"{workbook.stem}_charts"
        out_dir.mkdir(exist_ok=True)
        return out_dir

    @staticmethod
    def _top_mask(df: pd.DataFrame, column: str, ascending: bool) -> np.ndarray:
        if df is None or len(df) == 0 or column not in df.columns:
            return np.array([], dtype=bool)
        n_top = max(1, int(np.ceil(len(df) * 0.10)))
        ranked = df[column].replace([np.inf, -np.inf], np.nan)
        order = ranked.sort_values(ascending=ascending).index[:n_top]
        return df.index.isin(order)

    def _save_risk_return_metric_chart(
        self,
        optimizer: QAOAOptimizerV51,
        out_dir: Path,
        *,
        metric: str,
        metric_title: str,
        classical_cmap,
        quantum_cmap,
        classical_label: str,
        quantum_label: str,
        classical_ascending: bool,
        quantum_ascending: bool,
        quantum_selection: str,
        filename: str,
        x_col: str = "portfolio_vol",
        y_col: str = "portfolio_return",
        x_label: str = "Portfolio Volatility",
        y_label: str = "Portfolio Return",
    ) -> Path:
        classical = optimizer.classical_results.head(100).copy()
        if len(optimizer.samples_df) and quantum_selection == "qubo":
            quantum = optimizer.samples_df.nsmallest(100, "qubo_value").copy()
        else:
            quantum = (
                optimizer.samples_df.nlargest(100, "probability").copy()
                if len(optimizer.samples_df) and "probability" in optimizer.samples_df.columns
                else optimizer.samples_df.head(100).copy()
            )

        fig, ax = plt.subplots(figsize=(11, 7))
        fig.patch.set_facecolor(self.colors["bg"])
        ax.set_facecolor(self.colors["panel"])

        metric_frames = []
        if len(classical):
            metric_frames.append(classical[metric])
        if len(quantum):
            metric_frames.append(quantum[metric])
        if metric_frames:
            combined_metric = pd.concat(metric_frames).replace([np.inf, -np.inf], np.nan).dropna()
            if len(combined_metric):
                norm = mcolors.Normalize(vmin=float(combined_metric.min()), vmax=float(combined_metric.max()))
            else:
                norm = mcolors.Normalize(vmin=0.0, vmax=1.0)
        else:
            norm = mcolors.Normalize(vmin=0.0, vmax=1.0)

        def plot_solver(df, marker, label, cmap, top_mask):
            if df is None or len(df) == 0 or x_col not in df.columns or y_col not in df.columns:
                return None
            values = df[metric].replace([np.inf, -np.inf], np.nan).fillna(0.0)
            sizes = np.full(len(df), 52.0)
            if "probability" in df.columns and df["probability"].notna().any():
                sizes = 40 + 260 * df["probability"].fillna(0.0).to_numpy()
            edgecolors = np.where(top_mask, "#FFFFFF", "#B9C8E8")
            linewidths = np.where(top_mask, 1.8, 0.5)
            scatter = ax.scatter(
                df[x_col],
                df[y_col],
                c=values,
                s=sizes,
                cmap=cmap,
                norm=norm,
                alpha=0.78,
                edgecolors=edgecolors,
                linewidths=linewidths,
                marker=marker,
                label=label,
            )
            return scatter

        classical_mask = self._top_mask(classical, metric, classical_ascending)
        quantum_mask = self._top_mask(quantum, metric, quantum_ascending)

        classical_scatter = plot_solver(classical, "o", classical_label, classical_cmap, classical_mask)
        quantum_scatter = plot_solver(quantum, "D", quantum_label, quantum_cmap, quantum_mask)

        for df, label in [(classical.head(1), "Best Classical"), (quantum.head(1), "Best Quantum")]:
            if len(df) and x_col in df.columns and y_col in df.columns:
                row = df.iloc[0]
                ax.annotate(
                    label,
                    (row[x_col], row[y_col]),
                    xytext=(8, 8),
                    textcoords="offset points",
                    color=self.colors["text"],
                    fontsize=9,
                )

        ax.set_title(metric_title, color=self.colors["accent"], fontsize=15, fontweight="bold")
        ax.set_xlabel(x_label, color=self.colors["text"])
        ax.set_ylabel(y_label, color=self.colors["text"])
        ax.tick_params(colors=self.colors["muted"])
        for spine in ax.spines.values():
            spine.set_color(self.colors["border"])
        ax.grid(True, color="#233A68", alpha=0.35)
        ax.legend(facecolor=self.colors["panel_alt"], edgecolor=self.colors["border"], labelcolor=self.colors["text"])
        fig.subplots_adjust(right=0.82)
        if classical_scatter is not None:
            classical_cax = fig.add_axes([0.845, 0.15, 0.022, 0.72])
            cbar = fig.colorbar(classical_scatter, cax=classical_cax)
            cbar.set_label(f"Classical {metric_title.split('-')[-1].strip()}", color=self.colors["text"])
            cbar.ax.yaxis.set_tick_params(color=self.colors["muted"])
            plt.setp(cbar.ax.get_yticklabels(), color=self.colors["muted"])
            cbar.outline.set_edgecolor(self.colors["border"])
        if quantum_scatter is not None:
            quantum_cax = fig.add_axes([0.905, 0.15, 0.022, 0.72])
            cbar = fig.colorbar(quantum_scatter, cax=quantum_cax)
            cbar.set_label(f"Quantum {metric_title.split('-')[-1].strip()}", color=self.colors["text"])
            cbar.ax.yaxis.set_tick_params(color=self.colors["muted"])
            plt.setp(cbar.ax.get_yticklabels(), color=self.colors["muted"])
            cbar.outline.set_edgecolor(self.colors["border"])

        path = out_dir / filename
        fig.savefig(path, dpi=180, facecolor=fig.get_facecolor())
        plt.close(fig)
        return path

    def _save_risk_return_sharpe_chart(self, optimizer: QAOAOptimizerV51, out_dir: Path) -> Path:
        classical_cmap = mcolors.LinearSegmentedColormap.from_list(
            "classical_sharpe",
            ["#4A2C1A", "#A86A2A", "#FFE36E"],
        )
        quantum_cmap = mcolors.LinearSegmentedColormap.from_list(
            "quantum_sharpe",
            ["#0A2A7A", "#3A8DFF", "#FFFFFF"],
        )
        return self._save_risk_return_metric_chart(
            optimizer,
            out_dir,
            metric="sharpe_like",
            metric_title="Risk / Return - Sharpe Ratio",
            classical_cmap=classical_cmap,
            quantum_cmap=quantum_cmap,
            classical_label="Classical top 100 by QUBO",
            quantum_label="Quantum top 100 by probability",
            classical_ascending=False,
            quantum_ascending=False,
            quantum_selection="probability",
            filename="risk_return_sharpe_ratio.png",
        )

    def _save_risk_return_qubo_chart(self, optimizer: QAOAOptimizerV51, out_dir: Path) -> Path:
        classical_cmap = mcolors.LinearSegmentedColormap.from_list(
            "classical_qubo",
            ["#FFD166", "#B22222", "#2F0A0A"],
        )
        quantum_cmap = mcolors.LinearSegmentedColormap.from_list(
            "quantum_qubo",
            ["#DFF6FF", "#1D4ED8", "#061A40"],
        )
        return self._save_risk_return_metric_chart(
            optimizer,
            out_dir,
            metric="qubo_value",
            metric_title="Risk / Return - QUBO Value",
            classical_cmap=classical_cmap,
            quantum_cmap=quantum_cmap,
            classical_label="Classical top 100 by QUBO",
            quantum_label="Quantum top 100 by probability",
            classical_ascending=True,
            quantum_ascending=True,
            quantum_selection="probability",
            filename="risk_return_qubo_value.png",
        )

    def _save_risk_return_qubo_quantum_by_qubo_chart(self, optimizer: QAOAOptimizerV51, out_dir: Path) -> Path:
        classical_cmap = mcolors.LinearSegmentedColormap.from_list(
            "classical_qubo_quantum_by_qubo",
            ["#FFD166", "#B22222", "#2F0A0A"],
        )
        quantum_cmap = mcolors.LinearSegmentedColormap.from_list(
            "quantum_qubo_quantum_by_qubo",
            ["#DFF6FF", "#1D4ED8", "#061A40"],
        )
        return self._save_risk_return_metric_chart(
            optimizer,
            out_dir,
            metric="qubo_value",
            metric_title="Risk / Return - QUBO Value (Quantum by QUBO)",
            classical_cmap=classical_cmap,
            quantum_cmap=quantum_cmap,
            classical_label="Classical top 100 by QUBO",
            quantum_label="Quantum top 100 by QUBO",
            classical_ascending=True,
            quantum_ascending=True,
            quantum_selection="qubo",
            filename="risk_return_qubo_value_quantum_by_qubo.png",
        )

    def _save_risk_return_qubo_budget_normalized_chart(self, optimizer: QAOAOptimizerV51, out_dir: Path) -> Path:
        classical_cmap = mcolors.LinearSegmentedColormap.from_list(
            "classical_qubo_budget_norm",
            ["#FFD166", "#B22222", "#2F0A0A"],
        )
        quantum_cmap = mcolors.LinearSegmentedColormap.from_list(
            "quantum_qubo_budget_norm",
            ["#DFF6FF", "#1D4ED8", "#061A40"],
        )
        return self._save_risk_return_metric_chart(
            optimizer,
            out_dir,
            metric="qubo_value",
            metric_title="Risk / Return - QUBO Value (Budget-Normalized)",
            classical_cmap=classical_cmap,
            quantum_cmap=quantum_cmap,
            classical_label="Classical top 100 by QUBO",
            quantum_label="Quantum top 100 by probability",
            classical_ascending=True,
            quantum_ascending=True,
            quantum_selection="probability",
            filename="risk_return_qubo_value_budget_normalized.png",
            x_col="portfolio_vol_budget_normalized",
            y_col="portfolio_return_budget_normalized",
            x_label="Portfolio Volatility (Budget-Normalized)",
            y_label="Portfolio Return (Budget-Normalized)",
        )

    def _save_solver_comparison_chart(self, optimizer: QAOAOptimizerV51, out_dir: Path) -> Path:
        comparison = optimizer.solver_comparison_df.copy()
        fig, axes = plt.subplots(2, 2, figsize=(11, 8))
        fig.patch.set_facecolor(self.colors["bg"])
        metrics = [
            ("qubo_value", "QUBO Value"),
            ("portfolio_return", "Portfolio Return"),
            ("portfolio_vol", "Portfolio Volatility"),
            ("sharpe_like", "Sharpe Ratio"),
        ]

        if len(comparison) == 0:
            for ax in axes.flat:
                ax.set_facecolor(self.colors["panel"])
                ax.text(
                    0.5,
                    0.5,
                    "No solver comparison available",
                    ha="center",
                    va="center",
                    color=self.colors["text"],
                    fontsize=12,
                )
                ax.set_xticks([])
                ax.set_yticks([])
            path = out_dir / "solver_comparison.png"
            fig.tight_layout()
            fig.savefig(path, dpi=180, facecolor=fig.get_facecolor())
            plt.close(fig)
            return path

        colors = [self.colors["accent"], self.colors["accent_alt"]]
        labels = comparison["solver"].tolist()
        for ax, (metric, title) in zip(axes.flat, metrics):
            ax.set_facecolor(self.colors["panel"])
            values = comparison[metric].replace([np.inf, -np.inf], np.nan).fillna(0.0).tolist()
            ax.bar(labels, values, color=colors[: len(labels)], alpha=0.9)
            ax.set_title(title, color=self.colors["accent"], fontsize=12, fontweight="bold")
            ax.tick_params(axis="x", colors=self.colors["muted"], rotation=10)
            ax.tick_params(axis="y", colors=self.colors["muted"])
            for spine in ax.spines.values():
                spine.set_color(self.colors["border"])
            ax.grid(True, axis="y", color="#233A68", alpha=0.35)

        fig.suptitle("Best Classical vs Quantum", color=self.colors["text"], fontsize=16, fontweight="bold")
        path = out_dir / "solver_comparison.png"
        fig.tight_layout(rect=(0, 0, 1, 0.96))
        fig.savefig(path, dpi=180, facecolor=fig.get_facecolor())
        plt.close(fig)
        return path

    def _save_qubo_breakdown_chart(self, optimizer, out_dir: Path) -> Path:
        classical = optimizer.classical_results.head(10).copy() if len(optimizer.classical_results) else pd.DataFrame()
        quantum_prob = (
            optimizer.samples_df.sort_values("probability", ascending=False).head(10).copy()
            if len(optimizer.samples_df) and "probability" in optimizer.samples_df.columns
            else pd.DataFrame()
        )
        quantum_qubo = (
            self._positive_probability_quantum_df(optimizer).sort_values("qubo_value").head(10).copy()
            if len(self._positive_probability_quantum_df(optimizer))
            else pd.DataFrame()
        )

        fig, axes = plt.subplots(3, 1, figsize=(12, 12), sharex=False)
        fig.patch.set_facecolor(self.colors["bg"])

        sections = [
            (axes[0], classical, "Classical Top 10 by QUBO", self.colors["accent_alt"]),
            (axes[1], quantum_prob, "Quantum Top 10 by Probability", self.colors["accent"]),
            (axes[2], quantum_qubo, "Quantum Top 10 by QUBO", "#9BE7FF"),
        ]
        term_colors = {
            "return_term": "#2ED8A3",
            "risk_term": "#FFB04D",
            "budget_term": "#FF5E7A",
        }

        for ax, df, title, accent in sections:
            ax.set_facecolor(self.colors["panel"])
            if df is None or len(df) == 0:
                ax.text(0.5, 0.5, "No data available", ha="center", va="center", color=self.colors["text"])
                ax.set_xticks([])
                ax.set_yticks([])
                continue

            labels = []
            x = np.arange(len(df))
            bottoms_pos = np.zeros(len(df))
            bottoms_neg = np.zeros(len(df))
            for idx, row in enumerate(df.itertuples(index=False), start=1):
                source = getattr(row, "source", "solver")
                labels.append(f"{idx}:{'Q' if 'qaoa' in str(source).lower() else 'C'}")

            for term in ["return_term", "risk_term", "budget_term"]:
                values = df[term].fillna(0.0).to_numpy(dtype=float)
                positive = np.where(values > 0, values, 0.0)
                negative = np.where(values < 0, values, 0.0)
                ax.bar(x, positive, bottom=bottoms_pos, color=term_colors[term], alpha=0.9, label=term if term not in ax.get_legend_handles_labels()[1] else None)
                ax.bar(x, negative, bottom=bottoms_neg, color=term_colors[term], alpha=0.9)
                bottoms_pos += positive
                bottoms_neg += negative

            ax.plot(x, df["qubo_value"].fillna(0.0).to_numpy(dtype=float), color="#FFFFFF", linewidth=2.0, marker="o", label="QUBO")
            ax.set_title(title, color=accent, fontsize=13, fontweight="bold")
            ax.set_ylabel("Contribution", color=self.colors["text"])
            ax.set_xticks(x, labels)
            ax.tick_params(axis="x", colors=self.colors["muted"])
            ax.tick_params(axis="y", colors=self.colors["muted"])
            for spine in ax.spines.values():
                spine.set_color(self.colors["border"])
            ax.grid(True, axis="y", color="#233A68", alpha=0.35)
            ax.legend(facecolor=self.colors["panel_alt"], edgecolor=self.colors["border"], labelcolor=self.colors["text"], ncol=4, fontsize=9)

        fig.suptitle("QUBO Breakdown by Portfolio", color=self.colors["text"], fontsize=16, fontweight="bold")
        path = out_dir / "qubo_breakdown.png"
        fig.tight_layout(rect=(0, 0, 1, 0.96))
        fig.savefig(path, dpi=180, facecolor=fig.get_facecolor())
        plt.close(fig)
        return path

    def _save_qaoa_history_chart(self, optimizer: QAOAOptimizerV51, out_dir: Path) -> Path | None:
        if optimizer.history_df is None or len(optimizer.history_df) == 0:
            return None
        fig, ax = plt.subplots(figsize=(11, 6))
        fig.patch.set_facecolor(self.colors["bg"])
        ax.set_facecolor(self.colors["panel"])
        ax.plot(optimizer.history_df["iteration"], optimizer.history_df["energy"], color=self.colors["accent"], label="Energy")
        ax.plot(
            optimizer.history_df["iteration"],
            optimizer.history_df["best_energy"],
            color="#7BE0FF",
            linestyle="--",
            label="Best so far",
        )
        ax.set_title("QAOA Optimization History", color=self.colors["accent"], fontsize=15, fontweight="bold")
        ax.set_xlabel("Iteration", color=self.colors["text"])
        ax.set_ylabel("Energy", color=self.colors["text"])
        ax.tick_params(colors=self.colors["muted"])
        for spine in ax.spines.values():
            spine.set_color(self.colors["border"])
        ax.grid(True, color="#233A68", alpha=0.35)
        ax.legend(facecolor=self.colors["panel_alt"], edgecolor=self.colors["border"], labelcolor=self.colors["text"])
        path = out_dir / "qaoa_history.png"
        fig.tight_layout()
        fig.savefig(path, dpi=180, facecolor=fig.get_facecolor())
        plt.close(fig)
        return path

    def _best_row_text(self, row: pd.Series, label: str) -> str:
        prob = row.get("probability", np.nan)
        prob_str = "n/a" if pd.isna(prob) else f"{prob:.4f}"
        text = (
            f"{label}\n"
            f"source: {row.get('source', 'n/a')}\n"
            f"QUBO: {row.get('qubo_value', np.nan):.6f}\n"
            f"Invested: {row.get('selected_usd', np.nan):,.2f} USD\n"
            f"Budget gap: {row.get('budget_gap', np.nan):,.2f} USD\n"
            f"Return (invested-only): {row.get('portfolio_return', np.nan):.6f}\n"
            f"Volatility (invested-only): {row.get('portfolio_vol', np.nan):.6f}\n"
            f"Sharpe ratio (invested-only): {row.get('sharpe_like', np.nan):.6f}\n"
            f"Probability: {prob_str}"
        )
        cash_weight = row.get("cash_weight", np.nan)
        ret_budget = row.get("portfolio_return_budget_normalized", np.nan)
        vol_budget = row.get("portfolio_vol_budget_normalized", np.nan)
        sharpe_budget = row.get("sharpe_like_budget_normalized", np.nan)
        if not all(pd.isna(val) for val in [cash_weight, ret_budget, vol_budget, sharpe_budget]):
            text += (
                f"\nCash weight: {100.0 * float(cash_weight):.2f}%\n"
                f"Return (budget-normalized): {float(ret_budget):.6f}\n"
                f"Volatility (budget-normalized): {float(vol_budget):.6f}\n"
                f"Sharpe ratio (budget-normalized): {float(sharpe_budget):.6f}"
            )
        return text

    def _quantum_feasibility_tolerance(self, optimizer) -> float:
        val = getattr(optimizer, "qaoa_feasibility_budget_tolerance_usd", np.nan)
        return float(val) if pd.notna(val) else 2500.0

    def _quantum_analysis(self, optimizer) -> dict:
        df = self._positive_probability_quantum_df(optimizer)
        analysis = {
            "positive_df": df,
            "most_likely": None,
            "best_by_qubo": None,
            "best_topk_probability": None,
            "topk_k": 0,
            "mass_qubo_lt_1": np.nan,
            "mass_qubo_lt_2": np.nan,
            "mass_qubo_lt_3": np.nan,
            "mass_qubo_lt_5": np.nan,
            "mass_budget_gap_lt_2500": np.nan,
            "mass_budget_gap_lt_5000": np.nan,
            "mass_budget_gap_lt_25000": np.nan,
            "mass_budget_gap_lt_50000": np.nan,
            "expected_selected_usd": np.nan,
            "expected_budget_gap": np.nan,
            "expected_qubo": np.nan,
            "mass_label_suffix": "",
        }
        if df is None or len(df) == 0:
            return analysis

        probs = pd.to_numeric(df["probability"], errors="coerce").fillna(0.0)
        qubo = pd.to_numeric(df["qubo_value"], errors="coerce")
        abs_budget_gap = pd.to_numeric(df.get("abs_budget_gap", np.nan), errors="coerce").fillna(np.inf)
        selected_usd = pd.to_numeric(df.get("selected_usd", np.nan), errors="coerce").fillna(0.0)
        budget_gap = pd.to_numeric(df.get("budget_gap", np.nan), errors="coerce").fillna(0.0)

        analysis["most_likely"] = df.loc[probs.idxmax()]
        analysis["best_by_qubo"] = df.loc[qubo.idxmin()]

        topk_k = min(50, len(df))
        if topk_k > 0:
            topk_df = df.nlargest(topk_k, "probability").copy()
            analysis["best_topk_probability"] = topk_df.loc[pd.to_numeric(topk_df["qubo_value"], errors="coerce").idxmin()]
            analysis["topk_k"] = topk_k

        analysis["mass_qubo_lt_1"] = float(probs[qubo < 1.0].sum())
        analysis["mass_qubo_lt_2"] = float(probs[qubo < 2.0].sum())
        analysis["mass_qubo_lt_3"] = float(probs[qubo < 3.0].sum())
        analysis["mass_qubo_lt_5"] = float(probs[qubo < 5.0].sum())
        analysis["mass_budget_gap_lt_2500"] = float(probs[abs_budget_gap <= 2500.0].sum())
        analysis["mass_budget_gap_lt_5000"] = float(probs[abs_budget_gap <= 5000.0].sum())
        analysis["mass_budget_gap_lt_25000"] = float(probs[abs_budget_gap <= 25000.0].sum())
        analysis["mass_budget_gap_lt_50000"] = float(probs[abs_budget_gap <= 50000.0].sum())

        prob_mass = float(probs.sum())
        if prob_mass > 0:
            normalized = probs / prob_mass
            analysis["expected_selected_usd"] = float((normalized * selected_usd).sum())
            analysis["expected_budget_gap"] = float((normalized * budget_gap).sum())
            analysis["expected_qubo"] = float((normalized * qubo.fillna(0.0)).sum())

        total_states = getattr(optimizer, "qaoa_total_states_considered", np.nan)
        exported_states = len(df)
        if pd.notna(total_states) and int(total_states) > exported_states:
            analysis["mass_label_suffix"] = " (within exported quantum states)"
        return analysis

    def _positive_probability_quantum_df(self, optimizer) -> pd.DataFrame:
        if optimizer.samples_df is None or len(optimizer.samples_df) == 0:
            return pd.DataFrame()
        if "probability" not in optimizer.samples_df.columns:
            return optimizer.samples_df.copy()
        df = optimizer.samples_df.copy()
        probs = pd.to_numeric(df["probability"], errors="coerce").fillna(0.0)
        return df.loc[probs > 0].copy()

    def _build_summary_text(self, optimizer: QAOAOptimizerV51, chart_dir: Path) -> str:
        lines = [
            "Optimization complete.",
            "",
            f"Workbook: {optimizer.xlsx_path.name}",
            f"Charts saved to: {chart_dir}",
            f"Classical candidates: {len(optimizer.classical_results)}",
            f"Quantum candidates: {len(optimizer.samples_df)}",
            f"Overview portfolios: {len(optimizer.overview_df)}",
            f"Budget lambda: {getattr(optimizer, 'lambda_budget', np.nan)}",
            f"Risk lambda: {getattr(optimizer, 'lambda_variance', np.nan)}",
            f"Risk-free rate: {getattr(optimizer, 'risk_free', np.nan)}",
        ]
        if len(optimizer.classical_results):
            lines.append("")
            lines.append(self._best_row_text(optimizer.classical_results.sort_values("qubo_value").iloc[0], "Best Classical by QUBO"))
        if len(optimizer.samples_df):
            quantum_analysis = self._quantum_analysis(optimizer)
            positive_quantum = quantum_analysis["positive_df"]
            lines.append("")
            if quantum_analysis["most_likely"] is not None:
                lines.append(self._best_row_text(quantum_analysis["most_likely"], "Most Likely Quantum State"))
            if quantum_analysis["best_topk_probability"] is not None:
                lines.append("")
                lines.append(
                    self._best_row_text(
                        quantum_analysis["best_topk_probability"],
                        f"Best Quantum among Top {quantum_analysis['topk_k']} by Probability",
                    )
                )
            if len(positive_quantum):
                lines.append("")
                lines.append(self._best_row_text(quantum_analysis["best_by_qubo"], "Best Quantum by QUBO"))
                lines.append("")
                lines.append("Quantum Probability Mass Diagnostics")
                lines.append(
                    f"Mass with QUBO < 1{quantum_analysis['mass_label_suffix']}: {quantum_analysis['mass_qubo_lt_1']:.4f}"
                )
                lines.append(
                    f"Mass with QUBO < 2{quantum_analysis['mass_label_suffix']}: {quantum_analysis['mass_qubo_lt_2']:.4f}"
                )
                lines.append(
                    f"Mass with QUBO < 3{quantum_analysis['mass_label_suffix']}: {quantum_analysis['mass_qubo_lt_3']:.4f}"
                )
                lines.append(
                    f"Mass with QUBO < 5{quantum_analysis['mass_label_suffix']}: {quantum_analysis['mass_qubo_lt_5']:.4f}"
                )
                lines.append(
                    f"Mass with |budget gap| <= 2,500 USD{quantum_analysis['mass_label_suffix']}: {quantum_analysis['mass_budget_gap_lt_2500']:.4f}"
                )
                lines.append(
                    f"Mass with |budget gap| <= 5,000 USD{quantum_analysis['mass_label_suffix']}: {quantum_analysis['mass_budget_gap_lt_5000']:.4f}"
                )
                lines.append(
                    f"Mass with |budget gap| <= 25,000 USD{quantum_analysis['mass_label_suffix']}: {quantum_analysis['mass_budget_gap_lt_25000']:.4f}"
                )
                lines.append(
                    f"Mass with |budget gap| <= 50,000 USD{quantum_analysis['mass_label_suffix']}: {quantum_analysis['mass_budget_gap_lt_50000']:.4f}"
                )
                lines.append(
                    f"Expected invested USD{quantum_analysis['mass_label_suffix']}: {quantum_analysis['expected_selected_usd']:,.2f}"
                )
                lines.append(
                    f"Expected budget gap USD{quantum_analysis['mass_label_suffix']}: {quantum_analysis['expected_budget_gap']:,.2f}"
                )
                lines.append(
                    f"Expected QUBO{quantum_analysis['mass_label_suffix']}: {quantum_analysis['expected_qubo']:.6f}"
                )
                max_state_prob = getattr(optimizer, "qaoa_max_state_probability", np.nan)
                top10_prob_mass = getattr(optimizer, "qaoa_top10_probability_mass", np.nan)
                if pd.notna(max_state_prob):
                    lines.append(f"Maximum single-state probability: {float(max_state_prob):.4f}")
                if pd.notna(top10_prob_mass):
                    lines.append(f"Top-10 probability mass: {float(top10_prob_mass):.4f}")
        return "\n".join(lines)

    def _display_columns(self, df: pd.DataFrame, preferred: list[str]) -> list[str]:
        return [col for col in preferred if col in df.columns]

    def _format_value(self, col: str, value, formatters: dict | None = None):
        if pd.isna(value):
            return ""
        if formatters and col in formatters:
            return formatters[col](value)
        return str(value)

    def _render_dataframe_text(self, df: pd.DataFrame, columns: list[str], formatters: dict | None = None) -> str:
        if df is None or len(df) == 0 or not columns:
            return "No data available."

        rendered = df[columns].copy()
        for col in columns:
            rendered[col] = rendered[col].apply(lambda v, c=col: self._format_value(c, v, formatters))
        return rendered.to_string(index=False)

    def _add_text_tab(self, notebook, tab_name: str, content: str):
        frame = ttk.Frame(notebook, style="App.TFrame")
        notebook.add(frame, text=tab_name)
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(0, weight=1)

        text = scrolledtext.ScrolledText(
            frame,
            bg=self.colors["entry_bg"],
            fg=self.colors["text"],
            insertbackground=self.colors["text"],
            selectbackground=self.colors["accent_alt"],
            font=("Menlo", 11),
            relief=tk.FLAT,
            highlightthickness=1,
            highlightbackground=self.colors["border"],
            highlightcolor=self.colors["accent"],
            wrap=tk.NONE,
        )
        text.grid(row=0, column=0, sticky="nsew", padx=8, pady=8)
        text.insert("1.0", content)
        text.configure(state=tk.DISABLED)

    def _build_portfolio_contents_df(self, optimizer) -> pd.DataFrame:
        frames = []
        existing_holdings = (
            optimizer.portfolios_df.copy()
            if getattr(optimizer, "portfolios_df", None) is not None and len(optimizer.portfolios_df)
            else pd.DataFrame()
        )
        options_df = self._portfolio_options_df(optimizer)

        candidate_groups = []
        if len(optimizer.classical_results):
            candidate_groups.append(("Classical", optimizer.classical_results.head(5).copy()))
        if len(optimizer.samples_df):
            candidate_groups.append(
                ("Quantum by Probability", optimizer.samples_df.sort_values("probability", ascending=False).head(5).copy())
            )
            quantum_qubo_df = self._positive_probability_quantum_df(optimizer).sort_values("qubo_value").head(5).copy()
            if len(quantum_qubo_df):
                candidate_groups.append(("Quantum by QUBO", quantum_qubo_df))

        for portfolio_group, candidate_df in candidate_groups:
            if candidate_df is None or len(candidate_df) == 0:
                continue
            for idx, row in candidate_df.reset_index(drop=True).iterrows():
                bitstring = str(row.get("bitstring", ""))
                group_rows = pd.DataFrame()
                if len(existing_holdings):
                    group_rows = existing_holdings.loc[existing_holdings["bitstring"].astype(str) == bitstring].copy()
                if len(group_rows) == 0:
                    group_rows = self._reconstruct_portfolio_rows(optimizer, row, options_df)
                if len(group_rows) == 0:
                    continue
                group_rows.insert(0, "portfolio_group", portfolio_group)
                if "rank" not in group_rows.columns:
                    group_rows.insert(1, "rank", idx + 1)
                else:
                    group_rows["rank"] = idx + 1
                if "source" not in group_rows.columns:
                    group_rows["source"] = row.get("source", "")
                if "probability" not in group_rows.columns:
                    group_rows["probability"] = row.get("probability", np.nan)
                frames.append(group_rows)

        if not frames:
            return pd.DataFrame()

        combined = pd.concat(frames, ignore_index=True)
        if "probability" not in combined.columns:
            combined["probability"] = np.nan
        if "selected_usd" in combined.columns and "Approx Cost USD" in combined.columns:
            combined["portfolio_weight"] = np.where(
                combined["selected_usd"].fillna(0) > 0,
                combined["Approx Cost USD"].fillna(0) / combined["selected_usd"].replace(0, np.nan),
                np.nan,
            )
        return combined

    def _portfolio_options_df(self, optimizer) -> pd.DataFrame:
        options_df = getattr(optimizer, "options_df", None)
        if options_df is not None and len(options_df):
            return options_df.copy()

        xlsx_path = getattr(optimizer, "xlsx_path", None)
        if not xlsx_path:
            return pd.DataFrame()

        try:
            assets_df = pd.read_excel(xlsx_path, sheet_name="Assets", header=1)
        except Exception:
            return pd.DataFrame()

        options_df = assets_df.loc[assets_df["Ticker"].notna()].copy() if "Ticker" in assets_df.columns else pd.DataFrame()
        if len(options_df) == 0:
            return options_df
        if "Allowed" in options_df.columns:
            options_df = options_df.loc[options_df["Allowed"].fillna(1).astype(int) == 1].copy()
        if "Company" not in options_df.columns:
            options_df["Company"] = options_df["Ticker"]
        if "Option Label" not in options_df.columns:
            options_df["Option Label"] = ""
        if "Shares" not in options_df.columns:
            options_df["Shares"] = np.nan
        if "decision_id" not in options_df.columns:
            options_df["decision_id"] = [
                f"{ticker}_opt{i + 1}" for i, ticker in enumerate(options_df["Ticker"].astype(str).tolist())
            ]
        for col in ["Approx Cost USD", "Expected Return Proxy", "Annual Volatility", "Shares"]:
            if col in options_df.columns:
                options_df[col] = pd.to_numeric(options_df[col], errors="coerce")
        return options_df.reset_index(drop=True)

    def _reconstruct_portfolio_rows(self, optimizer, candidate_row: pd.Series, options_df: pd.DataFrame) -> pd.DataFrame:
        if options_df is None or len(options_df) == 0:
            return pd.DataFrame()

        bitstring = str(candidate_row.get("bitstring", ""))
        if not bitstring:
            return pd.DataFrame()

        bits = np.array([int(ch) for ch in bitstring if ch in {"0", "1"}], dtype=int)
        if len(bits) != len(options_df):
            return pd.DataFrame()

        selected_idx = np.flatnonzero(bits)
        if selected_idx.size == 0:
            return pd.DataFrame()

        rows = []
        for pos in selected_idx:
            option = options_df.iloc[int(pos)]
            row = {
                "rank": candidate_row.get("rank", np.nan),
                "source": candidate_row.get("source", ""),
                "bitstring": bitstring,
                "probability": candidate_row.get("probability", np.nan),
                "Ticker": option.get("Ticker", ""),
                "Company": option.get("Company", option.get("Ticker", "")),
                "Option Label": option.get("Option Label", ""),
                "Shares": option.get("Shares", np.nan),
                "Approx Cost USD": option.get("Approx Cost USD", np.nan),
                "Expected Return Proxy": option.get("Expected Return Proxy", np.nan),
                "Annual Volatility": option.get("Annual Volatility", np.nan),
                "decision_id": option.get("decision_id", ""),
                "qubo_value": candidate_row.get("qubo_value", np.nan),
                "return_term": candidate_row.get("return_term", np.nan),
                "risk_term": candidate_row.get("risk_term", np.nan),
                "budget_term": candidate_row.get("budget_term", np.nan),
                "qubo_reconstructed": candidate_row.get("qubo_reconstructed", np.nan),
                "selected_usd": candidate_row.get("selected_usd", np.nan),
                "budget_gap": candidate_row.get("budget_gap", np.nan),
                "abs_budget_gap": candidate_row.get("abs_budget_gap", np.nan),
                "num_options": candidate_row.get("num_options", np.nan),
                "num_distinct_assets": candidate_row.get("num_distinct_assets", np.nan),
                "portfolio_return": candidate_row.get("portfolio_return", np.nan),
                "portfolio_vol": candidate_row.get("portfolio_vol", np.nan),
                "sharpe_like": candidate_row.get("sharpe_like", np.nan),
                "cash_weight": candidate_row.get("cash_weight", np.nan),
                "portfolio_return_budget_normalized": candidate_row.get("portfolio_return_budget_normalized", np.nan),
                "portfolio_vol_budget_normalized": candidate_row.get("portfolio_vol_budget_normalized", np.nan),
                "sharpe_like_budget_normalized": candidate_row.get("sharpe_like_budget_normalized", np.nan),
                "max_position_usd": candidate_row.get("max_position_usd", np.nan),
            }
            rows.append(row)

        return pd.DataFrame(rows)

    def _portfolio_contents_text(self, optimizer) -> str:
        holdings_df = self._build_portfolio_contents_df(optimizer)
        if holdings_df is None or len(holdings_df) == 0:
            return "No portfolio contents available."

        blocks = []
        grouped = holdings_df.groupby(["portfolio_group", "bitstring"], sort=False)
        for (portfolio_group, bitstring), group in grouped:
            first = group.iloc[0]
            prob = first.get("probability", np.nan)
            prob_str = "n/a" if pd.isna(prob) else f"{float(prob):.4f}"
            blocks.append(f"{portfolio_group} Portfolio")
            blocks.append(
                f"rank={int(first.get('rank', 0))}  "
                f"source={first.get('source', 'n/a')}  "
                f"bitstring={bitstring}  "
                f"probability={prob_str}"
            )
            blocks.append(
                f"QUBO={float(first.get('qubo_value', np.nan)):.6f}  "
                f"invested={float(first.get('selected_usd', np.nan)):,.2f} USD  "
                f"return={float(first.get('portfolio_return', np.nan)):.6f}  "
                f"volatility={float(first.get('portfolio_vol', np.nan)):.6f}  "
                f"Sharpe={float(first.get('sharpe_like', np.nan)):.6f}"
            )
            if "portfolio_return_budget_normalized" in group.columns:
                blocks.append(
                    f"cash weight={100.0 * float(first.get('cash_weight', np.nan)):.2f}%  "
                    f"return (budget-norm)={float(first.get('portfolio_return_budget_normalized', np.nan)):.6f}  "
                    f"volatility (budget-norm)={float(first.get('portfolio_vol_budget_normalized', np.nan)):.6f}  "
                    f"Sharpe (budget-norm)={float(first.get('sharpe_like_budget_normalized', np.nan)):.6f}"
                )
            blocks.append("Selected assets:")

            asset_cols = [
                "Ticker",
                "Company",
                "Option Label",
                "Shares",
                "Approx Cost USD",
                "portfolio_weight",
            ]
            asset_df = group[asset_cols].copy()
            asset_df["Shares"] = asset_df["Shares"].apply(lambda v: "" if pd.isna(v) else f"{float(v):.4f}")
            asset_df["Approx Cost USD"] = asset_df["Approx Cost USD"].apply(
                lambda v: "" if pd.isna(v) else f"{float(v):,.2f}"
            )
            asset_df["portfolio_weight"] = asset_df["portfolio_weight"].apply(
                lambda v: "" if pd.isna(v) else f"{100 * float(v):.2f}%"
            )
            blocks.append(asset_df.to_string(index=False))
            blocks.append("")

        return "\n".join(blocks).strip()

    def _collect_circuit_details(self, optimizer) -> dict:
        probe = QAOAOptimizerV51(str(optimizer.xlsx_path), log_callback=None, progress_callback=None)
        probe.load_input()
        probe.build_qubo()

        qaoa_p = int(getattr(optimizer, "qaoa_p", getattr(probe, "qaoa_p", 0)) or 0)
        h, J, offset = probe._qubo_to_ising(probe.Q, probe.constant)
        nonzero_h = int(np.count_nonzero(np.abs(h) > 1e-12))
        nonzero_zz = int(sum(1 for value in J.values() if abs(value) > 1e-12))
        exact_threshold = int(getattr(probe, "qaoa_exact_probability_max_qubits", 0))
        exact_mode = probe.n <= exact_threshold

        hadamard_total = int(probe.n)
        cost_rz_total = int(qaoa_p * nonzero_h)
        zz_rz_total = int(qaoa_p * nonzero_zz)
        mixer_rx_total = int(qaoa_p * probe.n)
        cnot_total = int(qaoa_p * 2 * nonzero_zz)
        total_single_qubit = hadamard_total + cost_rz_total + zz_rz_total + mixer_rx_total
        total_two_qubit = cnot_total
        total_gate_applications = total_single_qubit + total_two_qubit

        initial_hadamard_depth = 1 if probe.n > 0 else 0
        cost_layer_depth = (1 if nonzero_h > 0 else 0) + (3 * nonzero_zz)
        mixer_layer_depth = 1 if probe.n > 0 else 0
        per_qaoa_layer_depth = cost_layer_depth + mixer_layer_depth
        total_serialized_depth = initial_hadamard_depth + qaoa_p * per_qaoa_layer_depth

        cost_layer_two_qubit_depth = 2 * nonzero_zz
        mixer_layer_two_qubit_depth = 0
        total_two_qubit_depth = qaoa_p * cost_layer_two_qubit_depth

        mode_text = "exact_probs" if exact_mode else f"shots_{int(getattr(probe, 'qaoa_shots', 0))}"
        state_space = 2 ** int(probe.n) if probe.n < 63 else None

        return {
            "workbook": Path(optimizer.xlsx_path).name,
            "n": int(probe.n),
            "p": qaoa_p,
            "mode_text": mode_text,
            "exact_threshold": exact_threshold,
            "state_space": state_space,
            "offset": float(offset),
            "nonzero_h": nonzero_h,
            "nonzero_zz": nonzero_zz,
            "hadamard_total": hadamard_total,
            "cost_rz_total": cost_rz_total,
            "zz_rz_total": zz_rz_total,
            "mixer_rx_total": mixer_rx_total,
            "cnot_total": cnot_total,
            "total_single_qubit": total_single_qubit,
            "total_two_qubit": total_two_qubit,
            "total_gate_applications": total_gate_applications,
            "initial_hadamard_depth": initial_hadamard_depth,
            "cost_layer_depth": cost_layer_depth,
            "mixer_layer_depth": mixer_layer_depth,
            "per_qaoa_layer_depth": per_qaoa_layer_depth,
            "total_serialized_depth": total_serialized_depth,
            "cost_layer_two_qubit_depth": cost_layer_two_qubit_depth,
            "mixer_layer_two_qubit_depth": mixer_layer_two_qubit_depth,
            "total_two_qubit_depth": total_two_qubit_depth,
        }

    def _build_circuit_details_text(self, details: dict) -> str:
        try:
            lines = [
                "QAOA Circuit Details",
                "",
                f"Workbook: {details['workbook']}",
                f"QAOA layers (p): {details['p']}",
                f"QAOA mode: {details['mode_text']}",
                f"Total qubits used: {details['n']}",
                f"Exact-mode threshold: {details['exact_threshold']} qubits",
            ]
            if details["state_space"] is not None:
                lines.append(f"Hilbert-space size: {details['state_space']:,} basis states")
            lines += [
                "",
                "Cost Hamiltonian structure",
                f"Nonzero single-qubit Z terms: {details['nonzero_h']}",
                f"Nonzero two-qubit ZZ terms: {details['nonzero_zz']}",
                f"QUBO constant / Ising offset: {details['offset']:.6f}",
                "",
                "Gate counts",
                f"Hadamard gates: {details['hadamard_total']}",
                f"Cost RZ gates from h terms: {details['cost_rz_total']}",
                f"Cost ZZ-decomposition RZ gates: {details['zz_rz_total']}",
                f"Mixer RX gates: {details['mixer_rx_total']}",
                f"CNOT gates: {details['cnot_total']}",
                f"Total single-qubit gates: {details['total_single_qubit']}",
                f"Total two-qubit gates: {details['total_two_qubit']}",
                f"Total gate applications: {details['total_gate_applications']}",
                "",
                "Sequential depth estimates",
                f"Initial Hadamard depth: {details['initial_hadamard_depth']}",
                f"Cost-layer sequential depth: {details['cost_layer_depth']}",
                f"Mixer-layer sequential depth: {details['mixer_layer_depth']}",
                f"Per-QAOA-layer sequential depth: {details['per_qaoa_layer_depth']}",
                f"Total sequential depth: {details['total_serialized_depth']}",
                "",
                "Sequential two-qubit depth",
                f"Cost-layer sequential 2-qubit gates: {details['cost_layer_two_qubit_depth']}",
                f"Mixer-layer sequential 2-qubit gates: {details['mixer_layer_two_qubit_depth']}",
                f"Total sequential 2-qubit gates: {details['total_two_qubit_depth']}",
                "",
                "Interpretation",
                "The depth estimate above is the sequential depth of the circuit as currently constructed in code.",
                "It is not a compiler-optimized hardware depth; commuting or parallelizable gates are not merged here.",
                "The circuit sketch below is a layer schematic that highlights init, cost, and mixer blocks in order.",
            ]
            return "\n".join(lines)
        except Exception as exc:
            return f"Could not build circuit details.\n\n{exc}"

    def _build_circuit_schematic_image(self, details: dict):
        n = int(details["n"])
        p = int(details["p"])
        block_labels = [("init", "Init\nH")]
        for layer_idx in range(1, p + 1):
            block_labels.append(("cost", f"C{layer_idx}\nZ:{details['nonzero_h']}\nZZ:{details['nonzero_zz']}"))
            block_labels.append(("mixer", f"M{layer_idx}\nRX:{n}"))

        block_width = 1.35
        gap = 0.20
        width_units = 0.8 + len(block_labels) * (block_width + gap)
        fig_w = max(8.0, min(36.0, width_units * 0.95))
        fig_h = min(4.2, max(2.6, 0.20 * n + 1.4))

        fig, ax = plt.subplots(figsize=(fig_w, fig_h))
        fig.patch.set_facecolor(self.colors["bg"])
        ax.set_facecolor(self.colors["panel"])

        y_positions = np.arange(n, 0, -1)
        x_start = 0.45
        total_width = x_start + len(block_labels) * (block_width + gap)

        for y in y_positions:
            ax.hlines(y, x_start - 0.2, total_width, color="#3A4C73", linewidth=1.0, alpha=0.85)

        palette = {
            "init": ("#7C4DFF", "#F0E7FF"),
            "cost": ("#F59E0B", "#FFF1D6"),
            "mixer": ("#06B6D4", "#DDF9FF"),
        }

        for idx, (kind, label) in enumerate(block_labels):
            x0 = x_start + idx * (block_width + gap)
            face, text_color = palette[kind]
            rect = plt.Rectangle(
                (x0, 0.55),
                block_width,
                n - 0.1,
                facecolor=face,
                edgecolor="#D6E3FF",
                linewidth=1.0,
                alpha=0.88,
            )
            ax.add_patch(rect)
            ax.text(
                x0 + block_width / 2,
                n / 2 + 0.5,
                label,
                ha="center",
                va="center",
                fontsize=8,
                color=text_color,
                fontweight="bold",
            )

        for wire_idx, y in enumerate(y_positions):
            ax.text(
                x_start - 0.28,
                y,
                f"q{wire_idx}",
                ha="right",
                va="center",
                fontsize=8,
                color=self.colors["muted"],
            )

        ax.set_xlim(0, total_width + 0.2)
        ax.set_ylim(0.3, n + 0.8)
        ax.set_xticks([])
        ax.set_yticks([])
        for spine in ax.spines.values():
            spine.set_visible(False)
        ax.set_title(
            "Circuit Layer Schematic",
            color=self.colors["accent"],
            fontsize=13,
            fontweight="bold",
            pad=10,
        )
        fig.tight_layout(pad=0.8)

        buffer = io.BytesIO()
        fig.savefig(buffer, format="png", dpi=180, facecolor=fig.get_facecolor(), bbox_inches="tight")
        plt.close(fig)
        buffer.seek(0)
        return Image.open(buffer).convert("RGBA")

    def _add_circuit_details_tab(self, notebook, optimizer):
        details = self._collect_circuit_details(optimizer)
        frame = ttk.Frame(notebook, style="App.TFrame")
        notebook.add(frame, text="Circuit Details")
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(1, weight=1)

        text = scrolledtext.ScrolledText(
            frame,
            height=16,
            bg=self.colors["entry_bg"],
            fg=self.colors["text"],
            insertbackground=self.colors["text"],
            selectbackground=self.colors["accent_alt"],
            font=("Menlo", 10),
            relief=tk.FLAT,
            highlightthickness=1,
            highlightbackground=self.colors["border"],
            highlightcolor=self.colors["accent"],
            wrap=tk.WORD,
        )
        text.grid(row=0, column=0, sticky="ew", padx=8, pady=(8, 4))
        text.insert("1.0", self._build_circuit_details_text(details))
        text.configure(state=tk.DISABLED)

        canvas_frame = ttk.Frame(frame, style="App.TFrame")
        canvas_frame.grid(row=1, column=0, sticky="nsew", padx=8, pady=(4, 8))
        canvas_frame.columnconfigure(0, weight=1)
        canvas_frame.rowconfigure(0, weight=1)

        canvas = tk.Canvas(
            canvas_frame,
            bg=self.colors["bg"],
            highlightthickness=1,
            highlightbackground=self.colors["border"],
            highlightcolor=self.colors["accent"],
        )
        canvas.grid(row=0, column=0, sticky="nsew")

        x_scroll = ttk.Scrollbar(canvas_frame, orient=tk.HORIZONTAL, command=canvas.xview)
        x_scroll.grid(row=1, column=0, sticky="ew")
        y_scroll = ttk.Scrollbar(canvas_frame, orient=tk.VERTICAL, command=canvas.yview)
        y_scroll.grid(row=0, column=1, sticky="ns")
        canvas.configure(xscrollcommand=x_scroll.set, yscrollcommand=y_scroll.set)

        try:
            image = self._build_circuit_schematic_image(details)
            tk_image = ImageTk.PhotoImage(image)
            self.result_images.append(tk_image)
            canvas.create_image(0, 0, anchor="nw", image=tk_image)
            canvas.configure(scrollregion=(0, 0, image.width, image.height))
        except Exception as exc:
            canvas.create_text(
                20,
                20,
                anchor="nw",
                text=f"Could not render circuit schematic.\n\n{exc}",
                fill=self.colors["text"],
                font=("Helvetica", 11),
            )
            canvas.configure(scrollregion=(0, 0, 900, 240))

    def _load_results_from_workbook(self, xlsx_path: str):
        path = Path(xlsx_path)
        xls = pd.ExcelFile(path)
        sheets = set(xls.sheet_names)

        def read_sheet(name: str):
            if name not in sheets:
                return pd.DataFrame()
            return pd.read_excel(xls, sheet_name=name)

        results_summary = read_sheet("Results_Summary")
        summary_map = {}
        if len(results_summary.columns) >= 2:
            key_col = results_summary.columns[0]
            val_col = results_summary.columns[1]
            summary_map = dict(zip(results_summary[key_col], results_summary[val_col]))

        decision_vars = int(summary_map.get("Decision variables", 0)) if pd.notna(summary_map.get("Decision variables", 0)) else 0

        def normalize_bitstrings(df: pd.DataFrame) -> pd.DataFrame:
            if df is None or len(df) == 0 or "bitstring" not in df.columns:
                return df
            df = df.copy()

            def norm(value):
                if pd.isna(value):
                    return value
                text = str(value).strip()
                if text.endswith(".0"):
                    text = text[:-2]
                if decision_vars > 0 and set(text).issubset({"0", "1"}):
                    return text.zfill(decision_vars)
                return text

            df["bitstring"] = df["bitstring"].apply(norm)
            return df

        return SimpleNamespace(
            xlsx_path=path,
            classical_results=normalize_bitstrings(read_sheet("Classical_Candidates")),
            samples_df=normalize_bitstrings(read_sheet("QAOA_Samples")),
            overview_df=normalize_bitstrings(read_sheet("Results_Overview")),
            portfolios_df=normalize_bitstrings(read_sheet("Results_Portfolios")),
            solver_comparison_df=normalize_bitstrings(read_sheet("Solver_Comparison")),
            history_df=read_sheet("Optimization_History"),
            qaoa_mode=summary_map.get("QAOA mode", "unknown"),
            qaoa_p=int(summary_map.get("QAOA p", 0)) if pd.notna(summary_map.get("QAOA p", 0)) else 0,
            qaoa_total_states_considered=int(summary_map.get("QAOA total states considered", 0)) if pd.notna(summary_map.get("QAOA total states considered", 0)) else 0,
            qaoa_max_state_probability=float(summary_map.get("QAOA max state probability", np.nan)) if pd.notna(summary_map.get("QAOA max state probability", np.nan)) else np.nan,
            qaoa_top10_probability_mass=float(summary_map.get("QAOA top10 probability mass", np.nan)) if pd.notna(summary_map.get("QAOA top10 probability mass", np.nan)) else np.nan,
            qaoa_feasibility_budget_tolerance_usd=float(summary_map.get("QAOA feasibility budget tolerance USD", np.nan)) if pd.notna(summary_map.get("QAOA feasibility budget tolerance USD", np.nan)) else np.nan,
            lambda_budget=float(summary_map.get("Budget lambda", summary_map.get("lambda_budget", np.nan))) if pd.notna(summary_map.get("Budget lambda", summary_map.get("lambda_budget", np.nan))) else np.nan,
            lambda_variance=float(summary_map.get("Risk lambda", summary_map.get("lambda_variance", np.nan))) if pd.notna(summary_map.get("Risk lambda", summary_map.get("lambda_variance", np.nan))) else np.nan,
            risk_free=float(summary_map.get("Risk-free rate", summary_map.get("risk_free_rate_annual", np.nan))) if pd.notna(summary_map.get("Risk-free rate", summary_map.get("risk_free_rate_annual", np.nan))) else np.nan,
        )

    def _show_results_window(self, optimizer: QAOAOptimizerV51):
        chart_dir = self._results_output_dir(str(optimizer.xlsx_path))
        chart_paths = [
            ("Risk / Return - Sharpe (Quantum by Probability)", self._save_risk_return_sharpe_chart(optimizer, chart_dir)),
            ("Risk / Return - QUBO (Quantum by Probability)", self._save_risk_return_qubo_chart(optimizer, chart_dir)),
            ("Risk / Return - QUBO (Quantum by QUBO)", self._save_risk_return_qubo_quantum_by_qubo_chart(optimizer, chart_dir)),
            ("Risk / Return - QUBO (Budget-Normalized)", self._save_risk_return_qubo_budget_normalized_chart(optimizer, chart_dir)),
            ("QUBO Breakdown", self._save_qubo_breakdown_chart(optimizer, chart_dir)),
            ("Solver Comparison", self._save_solver_comparison_chart(optimizer, chart_dir)),
        ]
        qaoa_history_path = self._save_qaoa_history_chart(optimizer, chart_dir)
        if qaoa_history_path is not None:
            chart_paths.append(("QAOA History", qaoa_history_path))

        if self.results_window is not None and self.results_window.winfo_exists():
            self.results_window.destroy()

        self.results_window = tk.Toplevel(self.root)
        self.results_window.title("Optimization Results")
        self.results_window.geometry("1220x860")
        self.results_window.configure(bg=self.colors["bg"])

        notebook = ttk.Notebook(self.results_window)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        self.result_images = []

        summary_frame = ttk.Frame(notebook, style="App.TFrame")
        notebook.add(summary_frame, text="Summary")
        summary_frame.columnconfigure(0, weight=1)
        summary_frame.rowconfigure(0, weight=1)
        summary_text = scrolledtext.ScrolledText(
            summary_frame,
            bg=self.colors["entry_bg"],
            fg=self.colors["text"],
            insertbackground=self.colors["text"],
            selectbackground=self.colors["accent_alt"],
            font=("Menlo", 10),
            relief=tk.FLAT,
            highlightthickness=1,
            highlightbackground=self.colors["border"],
            highlightcolor=self.colors["accent"],
            wrap=tk.WORD,
        )
        summary_text.grid(row=0, column=0, sticky="nsew", padx=8, pady=8)
        summary_text.insert("1.0", self._build_summary_text(optimizer, chart_dir))
        summary_text.configure(state=tk.DISABLED)

        self._add_circuit_details_tab(notebook, optimizer)

        classical_cols = self._display_columns(
            optimizer.classical_results,
            [
                "bitstring",
                "qubo_value",
                "selected_usd",
                "budget_gap",
                "cash_weight",
                "portfolio_return",
                "portfolio_vol",
                "sharpe_like",
                "portfolio_return_budget_normalized",
                "portfolio_vol_budget_normalized",
                "sharpe_like_budget_normalized",
            ],
        )
        quantum_cols = self._display_columns(
            optimizer.samples_df,
            [
                "bitstring",
                "probability",
                "qubo_value",
                "selected_usd",
                "budget_gap",
                "cash_weight",
                "portfolio_return",
                "portfolio_vol",
                "sharpe_like",
                "portfolio_return_budget_normalized",
                "portfolio_vol_budget_normalized",
                "sharpe_like_budget_normalized",
            ],
        )
        holdings_df = self._build_portfolio_contents_df(optimizer)
        holdings_cols = self._display_columns(
            holdings_df,
            [
                "portfolio_group",
                "rank",
                "source",
                "bitstring",
                "probability",
                "Ticker",
                "Company",
                "Option Label",
                "Shares",
                "Approx Cost USD",
                "portfolio_weight",
                "selected_usd",
                "portfolio_return",
                "portfolio_vol",
                "sharpe_like",
                "cash_weight",
                "portfolio_return_budget_normalized",
                "portfolio_vol_budget_normalized",
                "sharpe_like_budget_normalized",
            ],
        )
        formatters = {
            "qubo_value": lambda v: f"{float(v):.6f}",
            "probability": lambda v: f"{float(v):.4f}",
            "selected_usd": lambda v: f"{float(v):,.2f}",
            "Approx Cost USD": lambda v: f"{float(v):,.2f}",
            "cash_weight": lambda v: f"{100.0 * float(v):.2f}%",
            "portfolio_return": lambda v: f"{float(v):.6f}",
            "portfolio_vol": lambda v: f"{float(v):.6f}",
            "sharpe_like": lambda v: f"{float(v):.6f}",
            "portfolio_return_budget_normalized": lambda v: f"{float(v):.6f}",
            "portfolio_vol_budget_normalized": lambda v: f"{float(v):.6f}",
            "sharpe_like_budget_normalized": lambda v: f"{float(v):.6f}",
            "budget_gap": lambda v: f"{float(v):,.2f}",
            "portfolio_weight": lambda v: f"{100*float(v):.2f}%",
            "Shares": lambda v: f"{float(v):.4f}",
        }
        classical_text = self._render_dataframe_text(optimizer.classical_results.head(20), classical_cols, formatters)
        quantum_text = self._render_dataframe_text(
            optimizer.samples_df.sort_values("probability", ascending=False).head(20) if len(optimizer.samples_df) else optimizer.samples_df,
            quantum_cols,
            formatters,
        )
        portfolio_contents_text = self._portfolio_contents_text(optimizer)

        self._add_text_tab(notebook, "Top Classical", classical_text)
        self._add_text_tab(notebook, "Top Quantum", quantum_text)
        self._add_text_tab(notebook, "Portfolio Contents", portfolio_contents_text)

        for tab_name, image_path in chart_paths:
            frame = ttk.Frame(notebook, style="App.TFrame")
            notebook.add(frame, text=tab_name)
            frame.columnconfigure(0, weight=1)
            frame.rowconfigure(0, weight=1)

            if image_path is not None and Path(image_path).exists():
                image = Image.open(image_path).convert("RGBA")
                max_w, max_h = 1120, 760
                image.thumbnail((max_w, max_h), Image.Resampling.LANCZOS)
                tk_image = ImageTk.PhotoImage(image)
                self.result_images.append(tk_image)
                label = ttk.Label(frame, image=tk_image, background=self.colors["bg"])
                label.grid(row=0, column=0, sticky="nsew", padx=8, pady=8)
            else:
                ttk.Label(frame, text="No chart available", style="Subtitle.TLabel").grid(row=0, column=0, padx=20, pady=20)

    def _show_existing_results(self):
        xlsx_path = self.file_path_var.get().strip()
        if not self._workbook_has_results(xlsx_path):
            messagebox.showerror("No Results", "This workbook does not contain completed optimization result sheets yet.")
            return
        try:
            loaded = self._load_results_from_workbook(xlsx_path)
            self._show_results_window(loaded)
        except Exception as exc:
            messagebox.showerror("Error", f"Could not load existing results:\n\n{exc}")

    def _build_optimizer(self, xlsx_path):
        if self.use_workbook_settings_var.get():
            return QAOAOptimizerV51(
                xlsx_path=xlsx_path,
                progress_callback=self._update_status,
                log_callback=self._log_message,
                stop_check=lambda: not self.is_running,
            )

        return QAOAOptimizerV51(
            xlsx_path=xlsx_path,
            refresh_override=self.refresh_data_var.get(),
            enable_qaoa_override=self.enable_qaoa_var.get(),
            enable_classical_override=self.enable_classical_var.get(),
            qaoa_p_override=int(self.qaoa_p_var.get()),
            qaoa_maxiter_override=int(self.qaoa_maxiter_var.get()),
            qaoa_shots_override=self._current_qaoa_shots_value(),
            qaoa_multistart_restarts_override=int(self.qaoa_multistart_restarts_var.get()),
            qaoa_layerwise_warm_start_override=self.qaoa_layerwise_warm_start_var.get(),
            qaoa_restart_perturbation_override=float(self.qaoa_restart_perturbation_var.get()),
            lambda_budget_override=float(self.lambda_budget_var.get()),
            lambda_variance_override=float(self.lambda_variance_var.get()),
            risk_free_rate_override=float(self.risk_free_rate_var.get()),
            rng_seed_override=int(self.rng_seed_var.get()),
            progress_callback=self._update_status,
            log_callback=self._log_message,
            stop_check=lambda: not self.is_running,
        )

    def _run_optimization(self):
        xlsx_path = self.file_path_var.get().strip()
        if not xlsx_path or not Path(xlsx_path).exists():
            messagebox.showerror("Error", "Please select a valid Excel file.")
            return

        try:
            self._write_gui_settings_to_workbook(xlsx_path)
        except Exception as exc:
            messagebox.showerror("Error", f"Could not update workbook settings:\n\n{exc}")
            return

        self.log_text.configure(state=tk.NORMAL)
        self.log_text.delete("1.0", tk.END)
        self.log_text.configure(state=tk.DISABLED)

        self.progress_var.set(0)
        self.run_btn.configure(state=tk.DISABLED)
        self.stop_btn.configure(state=tk.NORMAL)
        self.show_results_btn.configure(state=tk.DISABLED)
        self.is_running = True

        self.optimizer_thread = threading.Thread(
            target=self._optimization_worker,
            args=(xlsx_path,),
            daemon=True,
        )
        self.optimizer_thread.start()

    def _stop_optimization(self):
        self.is_running = False
        self._update_status("Stopping... waiting for current optimization step to finish", self.progress_var.get())
        self._log_message("Stop requested. Cancelling at the next safe checkpoint...")

    def _optimization_worker(self, xlsx_path):
        try:
            self._log_message("=" * 70)
            self._log_message("Starting QAOA RQP by qubit-lab.ch")
            self._log_message("=" * 70)
            self.optimizer = self._build_optimizer(xlsx_path)
            self.optimizer.run_all()
            self.root.after(0, lambda: self._show_results_window(self.optimizer))
            self._log_message("=" * 70)
            self._log_message("Results written to Excel successfully.")
            self._log_message("=" * 70)
            self.root.after(0, lambda: messagebox.showinfo("Success", "Optimization complete. Results written to Excel."))
        except OptimizationCancelled:
            self._log_message("Optimization cancelled by user.")
            self._update_status("Stopped", self.progress_var.get())
        except OptimizationError as exc:
            self._log_message(f"Error: {exc}")
            self.root.after(0, lambda exc=exc: messagebox.showerror("Error", str(exc)))
        except Exception as exc:
            self._log_message("Unexpected error:")
            self._log_message(str(exc))
            self._log_message(traceback.format_exc())
            self.root.after(0, lambda exc=exc: messagebox.showerror("Error", f"An unexpected error occurred:\n\n{exc}"))
        finally:
            def finalize():
                self.run_btn.configure(state=tk.NORMAL)
                self.stop_btn.configure(state=tk.DISABLED)
                self.is_running = False
                self._refresh_results_button_state()

            self.root.after(0, finalize)


OptimizerGUIV5 = OptimizerGUIV51


def main():
    root = tk.Tk()
    app = OptimizerGUIV51(root)
    root.mainloop()
    return app


if __name__ == "__main__":
    main()
