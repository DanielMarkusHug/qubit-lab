from __future__ import annotations

import importlib
import sys
from itertools import product
from pathlib import Path
from types import SimpleNamespace

import numpy as np
import pandas as pd
import pytest
from openpyxl import Workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
VERSION_DIR = REPO_ROOT / "Version 9"

sys.path.insert(0, str(VERSION_DIR))

from app.qubo_builder import build_qubo_from_workbook  # noqa: E402


ASSETS = [
    {
        "ticker": "FIX",
        "role": "fixed",
        "cost": 25.0,
        "return": 0.08,
        "volatility": 0.20,
    },
    {
        "ticker": "AAA",
        "role": "variable",
        "cost": 40.0,
        "return": 0.12,
        "volatility": 0.30,
    },
    {
        "ticker": "BBB",
        "role": "variable",
        "cost": 60.0,
        "return": 0.04,
        "volatility": 0.40,
    },
]

COVARIANCE = np.array(
    [
        [0.040, 0.010, -0.005],
        [0.010, 0.090, 0.020],
        [-0.005, 0.020, 0.160],
    ],
    dtype=float,
)
BUDGET_USD = 100.0
RISK_FREE_RATE = 0.02
LAMBDA_BUDGET = 7.0
LAMBDA_VARIANCE = 3.0


def test_qubo_build_matches_direct_objective_for_fixed_and_variable_assets(tmp_path):
    optimizer = _build_audit_optimizer(tmp_path)

    assert optimizer.n == 2
    assert optimizer.fixed_tickers == ["FIX"]
    assert optimizer.opt_tickers == ["AAA", "BBB"]

    expected_by_bits = {
        tuple(bits): _direct_budget_first_objective(bits)
        for bits in _all_bit_vectors(optimizer.n)
    }
    expected_q = np.array(
        [
            [
                expected_by_bits[(1, 0)] - expected_by_bits[(0, 0)],
                expected_by_bits[(1, 1)]
                - expected_by_bits[(1, 0)]
                - expected_by_bits[(0, 1)]
                + expected_by_bits[(0, 0)],
            ],
            [0.0, expected_by_bits[(0, 1)] - expected_by_bits[(0, 0)]],
        ],
        dtype=float,
    )

    np.testing.assert_allclose(optimizer.Q, expected_q, atol=1e-10)
    assert optimizer.constant == pytest.approx(expected_by_bits[(0, 0)], abs=1e-10)

    for bits, expected_value in expected_by_bits.items():
        bitvec = np.asarray(bits, dtype=int)
        assert optimizer.qubo_value(bitvec) == pytest.approx(expected_value, abs=1e-10)

        breakdown = optimizer.qubo_term_breakdown(bitvec)
        assert breakdown["qubo_reconstructed"] == pytest.approx(expected_value, abs=1e-10)
        assert (
            breakdown["return_term"] + breakdown["risk_term"] + breakdown["budget_term"]
        ) == pytest.approx(expected_value, abs=1e-10)


def test_qubo_to_ising_preserves_energy_for_every_bitstring(tmp_path):
    optimizer = _build_audit_optimizer(tmp_path)
    h, j_terms, offset = optimizer._qubo_to_ising(optimizer.Q, optimizer.constant)

    for bits in _all_bit_vectors(optimizer.n):
        bitvec = np.asarray(bits, dtype=int)
        z = 1.0 - 2.0 * bitvec.astype(float)
        ising_energy = float(offset + np.dot(h, z))
        for (i, j), coeff in j_terms.items():
            ising_energy += float(coeff * z[i] * z[j])

        assert ising_energy == pytest.approx(optimizer.qubo_value(bitvec), abs=1e-10)


def test_exact_qaoa_circuit_probabilities_match_independent_dense_simulation(
    tmp_path,
    monkeypatch,
):
    pytest.importorskip("pennylane")
    optimizer = _build_audit_optimizer(tmp_path)
    legacy_core = importlib.import_module("qaoa_optimizer_V6_1_core")

    fixed_params = np.array([0.37, -0.21], dtype=float)

    def fake_minimize(objective, x0, method=None, options=None):
        x0 = np.asarray(x0, dtype=float)
        return SimpleNamespace(x=x0, fun=float(objective(x0)))

    monkeypatch.setattr(legacy_core, "minimize", fake_minimize)

    optimizer.enable_qaoa = True
    optimizer.qaoa_p = 1
    optimizer.qaoa_maxiter = 1
    optimizer.qaoa_multistart_restarts = 1
    optimizer.qaoa_layerwise_warm_start = False
    optimizer.qaoa_exact_probability_max_qubits = 8
    optimizer.qaoa_exact_qubo_diagnostic_rows = 4
    optimizer.qaoa_max_export_rows = 4
    optimizer.qaoa_export_mode = "top_k"
    optimizer.qaoa_export_sort_by = "probability"
    optimizer.qaoa_export_feasible_only = False
    optimizer.qaoa_min_probability_to_export = 0.0
    optimizer.qaoa_exact_p1_presearch = False
    optimizer.qaoa_concentration_polish_enabled = False
    optimizer.qaoa_sampled_concentration_polish_enabled = False
    optimizer._qaoa_initial_candidates = lambda *_args, **_kwargs: [("audit_fixed", fixed_params.copy())]

    optimizer.run_qaoa()

    expected_probs = _dense_qaoa_probabilities(optimizer, fixed_params)
    exported_probs = {
        str(row["bitstring"]): float(row["probability"])
        for row in optimizer.samples_df.to_dict("records")
    }

    assert set(exported_probs) == {"00", "01", "10", "11"}
    for idx, expected_prob in enumerate(expected_probs):
        bitstring = format(idx, f"0{optimizer.n}b")
        assert exported_probs[bitstring] == pytest.approx(expected_prob, abs=1e-9)

    expected_qubo = _expected_qubo_from_probabilities(optimizer, expected_probs)
    assert optimizer.qaoa_expected_qubo == pytest.approx(expected_qubo, abs=1e-9)
    assert optimizer.history_df.iloc[0]["expected_qubo"] == pytest.approx(expected_qubo, abs=1e-9)

    best_exact_row = optimizer.qaoa_exact_best_qubo_df.iloc[0]
    brute_best_bits = min(
        _all_bit_vectors(optimizer.n),
        key=lambda bits: optimizer.qubo_value(np.asarray(bits, dtype=int)),
    )
    assert best_exact_row["bitstring"] == "".join(map(str, brute_best_bits))


def test_candidate_ranking_matches_bruteforce_qubo_order(tmp_path):
    optimizer = _build_audit_optimizer(tmp_path)

    rows = []
    for bits in _all_bit_vectors(optimizer.n):
        bitvec = np.asarray(bits, dtype=int)
        stats = optimizer.portfolio_stats(bitvec)
        rows.append(
            {
                "bitstring": "".join(map(str, bits)),
                "qubo_value": optimizer.qubo_value(bitvec),
                **stats,
            }
        )

    sorted_rows = optimizer.sort_candidates(pd.DataFrame(rows))
    expected_rows = sorted(
        rows,
        key=lambda row: (
            row["qubo_value"],
            row["abs_budget_gap"],
            -row["portfolio_return"],
            row["bitstring"],
        ),
    )

    assert sorted_rows["bitstring"].tolist() == [row["bitstring"] for row in expected_rows]
    assert sorted_rows.iloc[0]["qubo_value"] == pytest.approx(expected_rows[0]["qubo_value"])


def _build_audit_optimizer(tmp_path):
    workbook_path = _audit_workbook(tmp_path)
    return build_qubo_from_workbook(workbook_path, lambda _message: None)


def _audit_workbook(tmp_path) -> Path:
    workbook_path = tmp_path / "algorithm_audit.xlsx"
    wb = Workbook()

    assets = wb.active
    assets.title = "Assets"
    asset_headers = [
        "Ticker",
        "Company",
        "Allowed",
        "Decision Role",
        "Indicative Market Cost USD",
        "Expected Return Proxy",
        "Annual Volatility",
    ]
    for col_idx, header in enumerate(asset_headers, start=1):
        assets.cell(2, col_idx).value = header
    for row_idx, asset in enumerate(ASSETS, start=3):
        assets.cell(row_idx, 1).value = asset["ticker"]
        assets.cell(row_idx, 2).value = asset["ticker"]
        assets.cell(row_idx, 3).value = 1
        assets.cell(row_idx, 4).value = asset["role"]
        assets.cell(row_idx, 5).value = asset["cost"]
        assets.cell(row_idx, 6).value = asset["return"]
        assets.cell(row_idx, 7).value = asset["volatility"]

    settings = wb.create_sheet("Settings")
    settings.cell(2, 1).value = "Key"
    settings.cell(2, 2).value = "Value"
    for row_idx, (key, value) in enumerate(
        [
            ("refresh_with_yfinance", 0),
            ("enable_classical_search", 1),
            ("enable_qaoa", 0),
            ("budget_usd", BUDGET_USD),
            ("risk_free_rate_annual", RISK_FREE_RATE),
            ("lambda_budget", LAMBDA_BUDGET),
            ("lambda_variance", LAMBDA_VARIANCE),
            ("rng_seed", 123),
            ("qaoa_exact_qubo_diagnostic_rows", 4),
            ("qaoa_max_export_rows", 4),
        ],
        start=3,
    ):
        settings.cell(row_idx, 1).value = key
        settings.cell(row_idx, 2).value = value

    annual_cov = wb.create_sheet("AnnualizedCovariance")
    tickers = [asset["ticker"] for asset in ASSETS]
    annual_cov.cell(2, 1).value = "Ticker"
    for col_idx, ticker in enumerate(tickers, start=2):
        annual_cov.cell(2, col_idx).value = ticker
    for row_offset, ticker in enumerate(tickers, start=3):
        annual_cov.cell(row_offset, 1).value = ticker
        for col_offset, value in enumerate(COVARIANCE[row_offset - 3], start=2):
            annual_cov.cell(row_offset, col_offset).value = float(value)

    wb.save(workbook_path)
    return workbook_path


def _all_bit_vectors(n: int):
    return [tuple(int(bit) for bit in bits) for bits in product((0, 1), repeat=n)]


def _direct_budget_first_objective(bits) -> float:
    bits = np.asarray(bits, dtype=float)
    costs = np.asarray([asset["cost"] for asset in ASSETS], dtype=float)
    returns = np.asarray([asset["return"] for asset in ASSETS], dtype=float)
    selected = np.concatenate([[1.0], bits])
    scaled_selected_cost = selected * costs / BUDGET_USD

    ret_scale = float(np.max(np.abs(returns - RISK_FREE_RATE)) + 1e-12)
    sigma_scale = float(np.max(np.abs(COVARIANCE)) + 1e-12)
    return_term = float(-np.dot((returns - RISK_FREE_RATE) / ret_scale, scaled_selected_cost))
    risk_term = float(
        LAMBDA_VARIANCE
        * (scaled_selected_cost @ (COVARIANCE / sigma_scale) @ scaled_selected_cost)
    )
    budget_term = float(LAMBDA_BUDGET * (scaled_selected_cost.sum() - 1.0) ** 2)
    return return_term + risk_term + budget_term


def _dense_qaoa_probabilities(optimizer, params) -> np.ndarray:
    gamma = float(params[0])
    beta = float(params[1])
    n_qubits = int(optimizer.n)
    h, j_terms, _offset = optimizer._qubo_to_ising(optimizer.Q, optimizer.constant)

    state = np.full(2**n_qubits, 1.0 / np.sqrt(2**n_qubits), dtype=complex)
    for state_idx in range(2**n_qubits):
        bits = np.asarray(list(map(int, format(state_idx, f"0{n_qubits}b"))), dtype=float)
        z = 1.0 - 2.0 * bits
        cost_energy = float(np.dot(h, z))
        for (i, j), coeff in j_terms.items():
            cost_energy += float(coeff * z[i] * z[j])
        state[state_idx] *= np.exp(-1j * gamma * cost_energy)

    rx = np.array(
        [
            [np.cos(beta), -1j * np.sin(beta)],
            [-1j * np.sin(beta), np.cos(beta)],
        ],
        dtype=complex,
    )
    identity = np.eye(2, dtype=complex)
    for wire in range(n_qubits):
        op = np.array([[1.0 + 0j]])
        for idx in range(n_qubits):
            op = np.kron(op, rx if idx == wire else identity)
        state = op @ state

    probs = np.square(np.abs(state)).astype(float)
    return probs / float(np.sum(probs))


def _expected_qubo_from_probabilities(optimizer, probs: np.ndarray) -> float:
    expected = 0.0
    for state_idx, probability in enumerate(probs):
        bits = np.asarray(list(map(int, format(state_idx, f"0{optimizer.n}b"))), dtype=int)
        expected += float(probability) * optimizer.qubo_value(bits)
    return float(expected)
