"""Workbook cost-column normalization for the Version 9 adapter layer."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook


INDICATIVE_COST_COLUMN = "Indicative Market Cost USD"
LEGACY_COST_COLUMN = "Approx Cost USD"


def _normalize_header(value: Any) -> str:
    return "".join(ch for ch in str(value or "").lower() if ch.isalnum())


_INDICATIVE_ALIASES = {
    _normalize_header(INDICATIVE_COST_COLUMN),
    _normalize_header("Indicative Market Cost (USD)"),
    _normalize_header("Indicative Market Value USD"),
}
_LEGACY_ALIASES = {_normalize_header(LEGACY_COST_COLUMN)}


def normalize_cost_column_for_legacy(workbook_path: Path, log_callback: Callable[[str], None] | None = None) -> dict[str, Any]:
    """Ensure the V6.1 core receives the canonical Version 9 cost column."""

    info: dict[str, Any] = {
        "input_cost_column": None,
        "internal_cost_column": INDICATIVE_COST_COLUMN,
        "normalized": False,
        "both_columns_present": False,
        "conflicting_row_count": 0,
        "warnings": [],
    }
    try:
        wb = load_workbook(workbook_path, data_only=True)
    except Exception:
        return info

    try:
        if "Assets" not in wb.sheetnames:
            return info
        ws = wb["Assets"]
        header_row = 2
        header_cells = list(ws[header_row])
        indicative_col = _find_header_column(header_cells, _INDICATIVE_ALIASES)
        legacy_col = _find_header_column(header_cells, _LEGACY_ALIASES)

        if indicative_col is None:
            if legacy_col is not None:
                info["input_cost_column"] = LEGACY_COST_COLUMN
                indicative_col = ws.max_column + 1
                ws.cell(header_row, indicative_col).value = INDICATIVE_COST_COLUMN
                for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
                    ws.cell(row_idx, indicative_col).value = ws.cell(row_idx, legacy_col).value
                info["normalized"] = True
                info["warnings"].append(
                    "Legacy Approx Cost USD input was mapped to Indicative Market Cost USD for compatibility."
                )
                wb.save(workbook_path)
                if log_callback is not None:
                    log_callback("Cost column used: Indicative Market Cost USD.")
            return info

        info["input_cost_column"] = INDICATIVE_COST_COLUMN
        if legacy_col is not None:
            info["both_columns_present"] = True

        conflict_count = 0
        if info["both_columns_present"]:
            for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
                indicative_value = ws.cell(row_idx, indicative_col).value
                legacy_value = ws.cell(row_idx, legacy_col).value
                if (
                    indicative_value not in (None, "")
                    and legacy_value not in (None, "")
                    and not _values_match(indicative_value, legacy_value)
                ):
                    conflict_count += 1

        info["conflicting_row_count"] = conflict_count
        if conflict_count:
            info["warnings"].append(
                "Indicative Market Cost USD and legacy Approx Cost USD differ; "
                "Indicative Market Cost USD was used for optimizer costs."
            )

        if log_callback is not None:
            log_callback("Cost column used: Indicative Market Cost USD.")
        return info
    finally:
        wb.close()


def annotate_optimizer_cost_columns(optimizer, info: dict[str, Any] | None) -> None:
    info = dict(info or {})
    warnings = list(info.get("warnings") or [])
    setattr(optimizer, "input_cost_column", info.get("input_cost_column") or INDICATIVE_COST_COLUMN)
    setattr(optimizer, "user_cost_column", INDICATIVE_COST_COLUMN)
    setattr(optimizer, "internal_cost_column", INDICATIVE_COST_COLUMN)
    setattr(optimizer, "cost_column_normalized", bool(info.get("normalized", False)))
    setattr(optimizer, "cost_column_both_present", bool(info.get("both_columns_present", False)))
    setattr(optimizer, "cost_column_conflicting_row_count", int(info.get("conflicting_row_count") or 0))
    setattr(optimizer, "cost_column_warnings", warnings)

    for attr_name in ("assets_df", "options_df", "fixed_options_df", "variable_options_df", "portfolios_df"):
        df = getattr(optimizer, attr_name, None)
        if df is not None and hasattr(df, "columns") and LEGACY_COST_COLUMN in df.columns:
            try:
                if INDICATIVE_COST_COLUMN not in df.columns:
                    df[INDICATIVE_COST_COLUMN] = df[LEGACY_COST_COLUMN]
            except Exception:
                pass


def add_indicative_cost_alias_to_frame(df):
    if df is None or not hasattr(df, "columns"):
        return df
    copy = df.copy()
    if LEGACY_COST_COLUMN in copy.columns and INDICATIVE_COST_COLUMN not in copy.columns:
        copy[INDICATIVE_COST_COLUMN] = copy[LEGACY_COST_COLUMN]
    return copy


def _find_header_column(header_cells, normalized_names: set[str]) -> int | None:
    for cell in header_cells:
        if _normalize_header(cell.value) in normalized_names:
            return int(cell.column)
    return None


def _values_match(left: Any, right: Any) -> bool:
    try:
        return abs(float(left) - float(right)) <= 1e-6
    except (TypeError, ValueError):
        return str(left).strip() == str(right).strip()
