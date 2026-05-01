#!/usr/bin/env python3
"""
QAOA Optimizer Core - Version 4.
Fixes v3 convergence bugs: COBYLA rhobeg, maxiter, and gamma init scaling.
"""

import random
import time
from pathlib import Path
from typing import Callable, Optional

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from scipy.optimize import minimize

try:
    import yfinance as yf
except ImportError:
    yf = None

try:
    import pennylane as qml
except ImportError:
    qml = None


class OptimizationError(Exception):
    """Raised when the optimizer cannot complete successfully."""


class OptimizationCancelled(OptimizationError):
    """Raised when a running optimization is cancelled."""


class QAOAOptimizerV3:
    """Application-friendly wrapper around the Version 4 notebook logic."""

    candidate_cols = [
        "bitstring",
        "source",
        "probability",
        "qubo_value",
        "return_term",
        "risk_term",
        "budget_term",
        "qubo_reconstructed",
        "avg_return_term_per_option",
        "avg_risk_term_per_option",
        "avg_budget_term_per_option",
        "selected_usd",
        "budget_gap",
        "abs_budget_gap",
        "num_options",
        "num_distinct_assets",
        "portfolio_return",
        "portfolio_vol",
        "sharpe_like",
        "cash_weight",
        "portfolio_return_budget_normalized",
        "portfolio_vol_budget_normalized",
        "sharpe_like_budget_normalized",
        "max_position_usd",
    ]

    def __init__(
        self,
        xlsx_path: str,
        refresh_override: Optional[bool] = None,
        enable_qaoa_override: Optional[bool] = None,
        enable_classical_override: Optional[bool] = None,
        qaoa_p_override: Optional[int] = None,
        qaoa_maxiter_override: Optional[int] = None,
        qaoa_shots_override: Optional[int] = None,
        qaoa_multistart_restarts_override: Optional[int] = None,
        qaoa_layerwise_warm_start_override: Optional[bool] = None,
        qaoa_restart_perturbation_override: Optional[float] = None,
        lambda_budget_override: Optional[float] = None,
        lambda_variance_override: Optional[float] = None,
        risk_free_rate_override: Optional[float] = None,
        rng_seed_override: Optional[int] = None,
        progress_callback: Optional[Callable[[str, Optional[float]], None]] = None,
        log_callback: Optional[Callable[[str], None]] = None,
        stop_check: Optional[Callable[[], bool]] = None,
    ):
        self.xlsx_path = Path(xlsx_path)

        self.refresh_override = refresh_override
        self.enable_qaoa_override = enable_qaoa_override
        self.enable_classical_override = enable_classical_override
        self.qaoa_p_override = qaoa_p_override
        self.qaoa_maxiter_override = qaoa_maxiter_override
        self.qaoa_shots_override = qaoa_shots_override
        self.qaoa_multistart_restarts_override = qaoa_multistart_restarts_override
        self.qaoa_layerwise_warm_start_override = qaoa_layerwise_warm_start_override
        self.qaoa_restart_perturbation_override = qaoa_restart_perturbation_override
        self.lambda_budget_override = lambda_budget_override
        self.lambda_variance_override = lambda_variance_override
        self.risk_free_rate_override = risk_free_rate_override
        self.rng_seed_override = rng_seed_override

        self.progress_callback = progress_callback or (lambda message, progress=None: None)
        self.log_callback = log_callback or (lambda message: None)
        self.stop_check = stop_check or (lambda: False)

        self.assets_df = pd.DataFrame()
        self.settings = {}
        self.options_df = pd.DataFrame()
        self.annual_cov_df = pd.DataFrame()

        self.refresh_with_yfinance = False
        self.top_n_export = 20
        self.enable_qaoa = True
        self.enable_classical = True
        self.qaoa_p = 1
        self.qaoa_maxiter = 60
        self.qaoa_shots = 4096
        self.qaoa_exact_probability_max_qubits = 20
        self.qaoa_max_qubits_allowed = 24
        self.qaoa_export_mode = "top_k"
        self.qaoa_min_probability_to_export = 1e-12
        self.qaoa_max_export_rows = 5000
        self.qaoa_export_feasible_only = False
        self.qaoa_feasibility_budget_tolerance_usd = 2500.0
        self.qaoa_export_sort_by = "probability"
        self.qaoa_multistart_restarts = 2
        self.qaoa_layerwise_warm_start = True
        self.qaoa_restart_perturbation = 0.05
        self.qaoa_eta_layer_growth_per_layer = 0.40
        self.qaoa_exact_p1_presearch = True
        self.qaoa_exact_p1_presearch_max_qubits = 16
        self.qaoa_exact_p1_presearch_gamma_min = -1.0
        self.qaoa_exact_p1_presearch_gamma_max = 1.0
        self.qaoa_exact_p1_presearch_gamma_points = 41
        self.qaoa_exact_p1_presearch_beta_min = 0.0
        self.qaoa_exact_p1_presearch_beta_max = np.pi
        self.qaoa_exact_p1_presearch_beta_points = 41
        self.classical_max_qubits_allowed = 30
        self.random_search_samples = 8000
        self.local_search_starts = 40
        self.classical_max_neighbor_evals = 200000
        self.overview_classical_pool = 300
        self.overview_qaoa_pool = 500
        self.result_candidate_limit_per_solver = 500
        self.rng_seed = 42
        self.budget_usd = 1_000_000.0
        self.risk_free = 0.04
        self.lambda_budget = 50.0
        self.lambda_variance = 6.0

        self.asset_universe = []
        self.decision_ids = []
        self.opt_tickers = []
        self.opt_cost = np.array([])
        self.opt_ret = np.array([])
        self.Sigma_opt = np.array([])
        self.Q = np.array([])
        self.constant = 0.0
        self.qubo_meta = {}
        self.n = 0

        self.classical_results = pd.DataFrame(columns=self.candidate_cols)
        self.samples_df = pd.DataFrame(columns=self.candidate_cols)
        self.overview_df = pd.DataFrame()
        self.portfolios_df = pd.DataFrame()
        self.solver_comparison_df = pd.DataFrame()
        self.history_df = pd.DataFrame()

        self.best_gammas = np.array([])
        self.best_betas = np.array([])
        self.qaoa_mode = "disabled"
        self.qaoa_total_nonzero_states = 0
        self.qaoa_total_states_considered = 0
        self.qaoa_final_energy = np.nan
        self.qaoa_expected_qubo = np.nan
        self.qaoa_max_state_probability = np.nan
        self.qaoa_top10_probability_mass = np.nan
        self.qaoa_top_probability_state_qubo = np.nan
        self.qaoa_best_exported_qubo = np.nan
        self.qaoa_best_exported_probability = np.nan
        self.qaoa_p1_presearch_used = False
        self.qaoa_p1_presearch_gamma = np.nan
        self.qaoa_p1_presearch_beta = np.nan
        self.qaoa_p1_presearch_energy = np.nan
        self.qaoa_p1_presearch_expected_qubo = np.nan

    def _log(self, message: str):
        self.log_callback(message)

    def _progress(self, message: str, progress: Optional[float] = None):
        self.progress_callback(message, progress)

    def _should_stop(self) -> bool:
        return self.stop_check()

    def _setting_value(self, key, default=None):
        return self.settings[key] if key in self.settings and pd.notna(self.settings[key]) else default

    def _setting_bool(self, key, default=False):
        val = self._setting_value(key, default)
        if isinstance(val, str):
            lowered = val.strip().lower()
            if lowered in {"1", "true", "yes", "y", "on"}:
                return True
            if lowered in {"0", "false", "no", "n", "off"}:
                return False
        try:
            return bool(int(val))
        except Exception:
            return bool(val)

    def _setting_int(self, key, default=0):
        val = self._setting_value(key, default)
        try:
            return int(float(val))
        except Exception:
            return int(default)

    def _setting_float(self, key, default=0.0):
        val = self._setting_value(key, default)
        try:
            return float(val)
        except Exception:
            return float(default)

    def _ensure_not_stopped(self):
        if self._should_stop():
            raise OptimizationCancelled("Optimization cancelled by user.")

    def _update_progress_fraction(self, phase_start: float, phase_end: float, completed: int, total: int, message: str):
        total = max(int(total), 1)
        completed = max(0, min(int(completed), total))
        fraction = completed / total
        progress = phase_start + (phase_end - phase_start) * fraction
        self._progress(message, progress)

    @staticmethod
    def _format_eta(seconds_remaining: float) -> str:
        seconds_remaining = max(0, int(round(seconds_remaining)))
        minutes, seconds = divmod(seconds_remaining, 60)
        hours, minutes = divmod(minutes, 60)
        if hours:
            return f"{hours}h {minutes}m {seconds}s"
        if minutes:
            return f"{minutes}m {seconds}s"
        return f"{seconds}s"

    def _qaoa_eta_seconds_remaining(self, progress_rows: list[dict], stage_restart_plan: list[tuple[int, int]]) -> float:
        if not progress_rows:
            return 0.0

        observed_by_layer: dict[int, list[float]] = {}
        for row in progress_rows:
            try:
                layer = int(row.get("layer_stage", 0))
                eval_sec = float(row.get("eval_sec", np.nan))
            except (TypeError, ValueError):
                continue
            if layer <= 0 or not np.isfinite(eval_sec) or eval_sec <= 0.0:
                continue
            observed_by_layer.setdefault(layer, []).append(eval_sec)

        layer_seconds = {
            layer: float(np.median(values))
            for layer, values in observed_by_layer.items()
            if values
        }
        if not layer_seconds:
            return 0.0

        growth = max(0.0, float(getattr(self, "qaoa_eta_layer_growth_per_layer", 0.40)))
        fit_coeffs = None
        if len(layer_seconds) >= 2:
            layers = np.array(sorted(layer_seconds), dtype=float)
            seconds = np.array([layer_seconds[int(layer)] for layer in layers], dtype=float)
            try:
                fit_coeffs = np.polyfit(layers, seconds, 1)
            except Exception:
                fit_coeffs = None

        def estimate_layer_seconds(layer: int) -> float:
            layer = int(layer)
            if layer in layer_seconds:
                return layer_seconds[layer]
            if fit_coeffs is not None:
                predicted = float(fit_coeffs[0] * layer + fit_coeffs[1])
                if np.isfinite(predicted) and predicted > 0.0:
                    max_seen = max(layer_seconds.values())
                    min_seen = min(layer_seconds.values())
                    return float(np.clip(predicted, 0.5 * min_seen, 3.0 * max_seen))

            nearest_layer = min(layer_seconds, key=lambda seen_layer: abs(seen_layer - layer))
            base = layer_seconds[nearest_layer]
            if layer > nearest_layer:
                return base * (1.0 + growth * (layer - nearest_layer))
            if layer < nearest_layer:
                return base / (1.0 + growth * (nearest_layer - layer))
            return base

        completed_by_layer: dict[int, int] = {}
        for row in progress_rows:
            try:
                layer = int(row.get("layer_stage", 0))
            except (TypeError, ValueError):
                continue
            completed_by_layer[layer] = completed_by_layer.get(layer, 0) + 1

        remaining_seconds = 0.0
        for active_layers, restart_count in stage_restart_plan:
            planned = int(self.qaoa_maxiter) * int(restart_count)
            remaining = max(planned - completed_by_layer.get(int(active_layers), 0), 0)
            if remaining:
                remaining_seconds += remaining * estimate_layer_seconds(int(active_layers))
        return float(max(0.0, remaining_seconds))

    def _progress_plan(self):
        if self.enable_qaoa:
            return {
                "load": 1.0,
                "refresh": 2.5 if self.refresh_with_yfinance else 1.5,
                "build": 4.0,
                "classical_start": 4.0,
                "classical_end": 8.0 if self.enable_classical else 4.0,
                "qaoa_start": 8.0,
                "qaoa_opt_end": 93.0,
                "qaoa_extract_end": 97.0,
                "generate": 98.5,
                "write": 99.5,
            }
        return {
            "load": 10.0,
            "refresh": 18.0 if self.refresh_with_yfinance else 12.0,
            "build": 28.0,
            "classical_start": 28.0,
            "classical_end": 92.0 if self.enable_classical else 28.0,
            "qaoa_start": 92.0,
            "qaoa_opt_end": 92.0,
            "qaoa_extract_end": 92.0,
            "generate": 97.0,
            "write": 99.0,
        }

    def _apply_runtime_settings(self):
        self.refresh_with_yfinance = self._setting_bool("refresh_market_data", False)
        if self.refresh_override is not None:
            self.refresh_with_yfinance = bool(self.refresh_override)

        self.top_n_export = self._setting_int("top_n_export", 20)
        self.enable_qaoa = self._setting_bool("enable_qaoa", True)
        if self.enable_qaoa_override is not None:
            self.enable_qaoa = bool(self.enable_qaoa_override)

        self.qaoa_p = self._setting_int("qaoa_p", 1)
        if self.qaoa_p_override is not None:
            self.qaoa_p = int(self.qaoa_p_override)

        self.qaoa_maxiter = self._setting_int("qaoa_maxiter", 60)
        if self.qaoa_maxiter_override is not None:
            self.qaoa_maxiter = int(self.qaoa_maxiter_override)

        self.qaoa_shots = self._setting_int("qaoa_shots", 4096)
        if self.qaoa_shots_override is not None:
            self.qaoa_shots = int(self.qaoa_shots_override)

        self.qaoa_exact_probability_max_qubits = self._setting_int("qaoa_exact_probability_max_qubits", 20)
        self.qaoa_max_qubits_allowed = self._setting_int("qaoa_max_qubits_allowed", 24)

        self.qaoa_export_mode = str(self._setting_value("qaoa_export_mode", "top_k")).strip().lower()
        if self.qaoa_export_mode not in {"top_k", "all_filtered"}:
            self.qaoa_export_mode = "top_k"

        self.qaoa_min_probability_to_export = self._setting_float("qaoa_min_probability_to_export", 1e-12)
        self.qaoa_max_export_rows = max(1, self._setting_int("qaoa_max_export_rows", 5000))
        self.qaoa_export_feasible_only = self._setting_bool("qaoa_export_feasible_only", False)
        self.qaoa_feasibility_budget_tolerance_usd = self._setting_float(
            "qaoa_feasibility_budget_tolerance_usd", 2500.0
        )
        self.qaoa_export_sort_by = str(self._setting_value("qaoa_export_sort_by", "probability")).strip().lower()
        if self.qaoa_export_sort_by not in {
            "probability",
            "qubo_value",
            "sharpe_like",
            "portfolio_return",
            "abs_budget_gap",
        }:
            self.qaoa_export_sort_by = "probability"
        self.qaoa_multistart_restarts = max(1, self._setting_int("qaoa_multistart_restarts", 2))
        if self.qaoa_multistart_restarts_override is not None:
            self.qaoa_multistart_restarts = max(1, int(self.qaoa_multistart_restarts_override))
        self.qaoa_layerwise_warm_start = self._setting_bool("qaoa_layerwise_warm_start", True)
        if self.qaoa_layerwise_warm_start_override is not None:
            self.qaoa_layerwise_warm_start = bool(self.qaoa_layerwise_warm_start_override)
        self.qaoa_restart_perturbation = max(0.0, self._setting_float("qaoa_restart_perturbation", 0.05))
        if self.qaoa_restart_perturbation_override is not None:
            self.qaoa_restart_perturbation = max(0.0, float(self.qaoa_restart_perturbation_override))
        self.qaoa_eta_layer_growth_per_layer = max(
            0.0,
            self._setting_float(
                "qaoa_eta_layer_growth_per_layer",
                self.qaoa_eta_layer_growth_per_layer,
            ),
        )
        self.qaoa_exact_p1_presearch = self._setting_bool("qaoa_exact_p1_presearch", True)
        self.qaoa_exact_p1_presearch_max_qubits = max(
            1, self._setting_int("qaoa_exact_p1_presearch_max_qubits", 16)
        )
        self.qaoa_exact_p1_presearch_gamma_min = self._setting_float(
            "qaoa_exact_p1_presearch_gamma_min", -1.0
        )
        self.qaoa_exact_p1_presearch_gamma_max = self._setting_float(
            "qaoa_exact_p1_presearch_gamma_max", 1.0
        )
        self.qaoa_exact_p1_presearch_gamma_points = max(
            3, self._setting_int("qaoa_exact_p1_presearch_gamma_points", 41)
        )
        self.qaoa_exact_p1_presearch_beta_min = self._setting_float(
            "qaoa_exact_p1_presearch_beta_min", 0.0
        )
        self.qaoa_exact_p1_presearch_beta_max = self._setting_float(
            "qaoa_exact_p1_presearch_beta_max", float(np.pi)
        )
        self.qaoa_exact_p1_presearch_beta_points = max(
            3, self._setting_int("qaoa_exact_p1_presearch_beta_points", 41)
        )

        self.enable_classical = self._setting_bool("enable_classical_search", True)
        if self.enable_classical_override is not None:
            self.enable_classical = bool(self.enable_classical_override)

        self.classical_max_qubits_allowed = self._setting_int("classical_max_qubits_allowed", 30)
        self.random_search_samples = self._setting_int("classical_random_search_samples", 8000)
        self.local_search_starts = self._setting_int("classical_local_search_starts", 40)
        self.classical_max_neighbor_evals = self._setting_int("classical_max_neighbor_evals", 200000)
        self.overview_classical_pool = self._setting_int("overview_classical_pool", 300)
        self.overview_qaoa_pool = self._setting_int("overview_qaoa_pool", 500)
        self.result_candidate_limit_per_solver = self._setting_int("result_candidate_limit_per_solver", 500)
        self.rng_seed = self._setting_int("rng_seed", 42)
        if self.rng_seed_override is not None:
            self.rng_seed = int(self.rng_seed_override)

        random.seed(self.rng_seed)
        np.random.seed(self.rng_seed)

        self.budget_usd = self._setting_float("budget_usd", 1_000_000.0)
        self.risk_free = self._setting_float("risk_free_rate_annual", 0.04)
        if self.risk_free_rate_override is not None:
            self.risk_free = float(self.risk_free_rate_override)
        self.lambda_budget = self._setting_float("lambda_budget", 50.0)
        if self.lambda_budget_override is not None:
            self.lambda_budget = float(self.lambda_budget_override)
        self.lambda_variance = self._setting_float("lambda_variance", 6.0)
        if self.lambda_variance_override is not None:
            self.lambda_variance = float(self.lambda_variance_override)

    def load_input(self):
        if not self.xlsx_path.exists():
            raise OptimizationError(f"Excel file not found: {self.xlsx_path}")

        self._log(f"Loading workbook: {self.xlsx_path.name}")
        xls = pd.ExcelFile(self.xlsx_path)
        self.assets_df = pd.read_excel(xls, sheet_name="Assets", header=1)
        settings_df = pd.read_excel(xls, sheet_name="Settings", header=1)
        self.settings = dict(zip(settings_df["Key"], settings_df["Value"]))
        self._apply_runtime_settings()

        self._log(f"Loaded {len(self.assets_df)} asset rows")
        self._log(f"Loaded {len(self.settings)} settings")
        self._log(f"Refresh with yfinance: {self.refresh_with_yfinance}")
        self._log(f"Enable QAOA: {self.enable_qaoa}")
        self._log(f"Enable classical search: {self.enable_classical}")
        self._log(f"QAOA export mode: {self.qaoa_export_mode}")
        self._log(f"QAOA export sort by: {self.qaoa_export_sort_by}")

    def refresh_market_data(self):
        if not self.refresh_with_yfinance:
            self._log("Skipping market data refresh because workbook settings do not request it.")
            return

        if yf is None:
            raise OptimizationError("yfinance is required when refresh_market_data = 1.")

        self._ensure_not_stopped()
        self._log("Refreshing market data from yfinance...")

        option_rows = self.assets_df.loc[self.assets_df["Ticker"].notna()].copy()
        if "Allowed" in option_rows.columns:
            option_rows = option_rows.loc[option_rows["Allowed"].fillna(1).astype(int) == 1].copy()

        tickers = list(dict.fromkeys(option_rows["Ticker"].astype(str).tolist()))
        if not tickers:
            raise OptimizationError("No tickers available to refresh.")

        self._log(f"Requested ticker refresh for {len(tickers)} symbols.")

        prices = yf.download(
            tickers=tickers,
            period="12mo",
            interval="1d",
            auto_adjust=True,
            progress=False,
        )["Close"]

        if isinstance(prices, pd.Series):
            prices = prices.to_frame()

        prices = prices.sort_index()
        available_tickers = []
        missing_tickers = []
        for ticker in tickers:
            if ticker in prices.columns and prices[ticker].notna().any():
                available_tickers.append(ticker)
            else:
                missing_tickers.append(ticker)

        if not available_tickers:
            raise OptimizationError(
                "Market data refresh failed: Yahoo Finance did not return usable data for any requested ticker."
            )

        if missing_tickers:
            self._log(
                "Tickers not found or without usable data: " + ", ".join(missing_tickers)
            )
            self._log(
                "Behavior: those tickers are skipped during refresh, their existing workbook values stay unchanged, "
                "and only found tickers contribute to refreshed return/covariance data."
            )

        self._log("Tickers refreshed successfully: " + ", ".join(available_tickers))

        prices = prices[available_tickers].dropna(how="all").ffill()

        valid_tickers = []
        insufficient_history_tickers = []
        for ticker in available_tickers:
            if prices[ticker].dropna().shape[0] >= 2:
                valid_tickers.append(ticker)
            else:
                insufficient_history_tickers.append(ticker)

        if insufficient_history_tickers:
            self._log(
                "Tickers with insufficient price history for refresh: "
                + ", ".join(insufficient_history_tickers)
            )
            self._log(
                "Behavior: those tickers are skipped for returns, volatility, and covariance updates, "
                "and their existing workbook values stay unchanged."
            )

        if not valid_tickers:
            raise OptimizationError(
                "Market data refresh failed: downloaded data did not contain at least two valid prices for any ticker."
            )

        prices = prices[valid_tickers]
        rets = prices.pct_change(fill_method=None)

        latest_price = prices.ffill().iloc[-1]
        total_return_12m = pd.Series(
            {
                ticker: float(series.dropna().iloc[-1] / series.dropna().iloc[0] - 1)
                for ticker, series in prices.items()
            }
        )
        ann_vol = rets.std(skipna=True) * np.sqrt(252)
        mean_daily = rets.mean(skipna=True)
        std_daily = rets.std(skipna=True)
        daily_cov = rets.cov(min_periods=2)
        annual_cov = daily_cov * 252

        if daily_cov.isna().all().all():
            raise OptimizationError(
                "Market data refresh failed: covariance matrix could not be computed from the downloaded price history."
            )

        wb = load_workbook(self.xlsx_path)
        sh_assets = wb["Assets"]
        sh_returns = wb["Returns"]
        sh_cov = wb["Covariance"]
        sh_acov = wb["AnnualizedCovariance"]
        sh_ph = wb["PriceHistory"]

        asset_headers = [sh_assets.cell(2, col).value for col in range(1, sh_assets.max_column + 1)]
        col_idx = {header: idx + 1 for idx, header in enumerate(asset_headers)}

        for row in sh_ph.iter_rows(
            min_row=3,
            max_row=max(sh_ph.max_row, 5000),
            min_col=1,
            max_col=1 + len(valid_tickers),
        ):
            for cell in row:
                cell.value = None

        for row_idx, dt in enumerate(prices.index, start=3):
            sh_ph.cell(row_idx, 1, dt.to_pydatetime())
            for col_idx_price, ticker in enumerate(valid_tickers, start=2):
                value = prices.loc[dt, ticker]
                sh_ph.cell(row_idx, col_idx_price, None if pd.isna(value) else float(value))

        for row_idx in range(3, sh_assets.max_row + 1):
            ticker = sh_assets.cell(row_idx, col_idx["Ticker"]).value
            if not ticker or ticker not in latest_price.index:
                continue

            sh_assets.cell(row_idx, col_idx["Current Price (USD)"], float(latest_price[ticker]))
            sh_assets.cell(row_idx, col_idx["Expected Return Proxy"], float(total_return_12m[ticker]))
            sh_assets.cell(row_idx, col_idx["Annual Volatility"], float(ann_vol[ticker]))

            if "Mean Daily Return" in col_idx:
                sh_assets.cell(row_idx, col_idx["Mean Daily Return"], float(mean_daily[ticker]))
            if "Std Daily Return" in col_idx:
                sh_assets.cell(row_idx, col_idx["Std Daily Return"], float(std_daily[ticker]))
            if "Price Source Status" in col_idx:
                sh_assets.cell(row_idx, col_idx["Price Source Status"], "Refreshed with yfinance")
            if "Source URL" in col_idx:
                sh_assets.cell(row_idx, col_idx["Source URL"], "https://pypi.org/project/yfinance/")

            if "Shares" in col_idx and "Approx Cost USD" in col_idx:
                shares = sh_assets.cell(row_idx, col_idx["Shares"]).value
                try:
                    if shares is not None and not pd.isna(shares):
                        sh_assets.cell(row_idx, col_idx["Approx Cost USD"], float(shares) * float(latest_price[ticker]))
                except Exception:
                    pass

        for row in sh_returns.iter_rows(min_row=3, max_row=max(sh_returns.max_row, 5000), min_col=1, max_col=6):
            for cell in row:
                cell.value = None

        for row_idx, ticker in enumerate(valid_tickers, start=3):
            sh_returns.cell(row_idx, 1, ticker)
            sh_returns.cell(row_idx, 2, float(total_return_12m[ticker]))
            sh_returns.cell(row_idx, 3, float(ann_vol[ticker]))
            sh_returns.cell(row_idx, 4, float(mean_daily[ticker]))
            sh_returns.cell(row_idx, 5, float(std_daily[ticker]))
            sh_returns.cell(row_idx, 6, "Refreshed with yfinance")

        for i, row_ticker in enumerate(valid_tickers, start=3):
            sh_cov.cell(i, 1, row_ticker)
            sh_acov.cell(i, 1, row_ticker)
            for j, col_ticker in enumerate(valid_tickers, start=2):
                sh_cov.cell(2, j, col_ticker)
                sh_acov.cell(2, j, col_ticker)
                cov_val = daily_cov.loc[row_ticker, col_ticker]
                acov_val = annual_cov.loc[row_ticker, col_ticker]
                sh_cov.cell(i, j, None if pd.isna(cov_val) else float(cov_val))
                sh_acov.cell(i, j, None if pd.isna(acov_val) else float(acov_val))

        wb.save(self.xlsx_path)
        skipped_count = len(missing_tickers) + len(insufficient_history_tickers)
        if skipped_count:
            self._log(
                "Refresh summary: "
                f"{len(valid_tickers)} refreshed, {skipped_count} skipped. "
                "A later optimization step may still fail if skipped tickers are still required in AnnualizedCovariance."
            )
        else:
            self._log(f"Refresh summary: all {len(valid_tickers)} tickers refreshed successfully.")
        self._log(f"Workbook refreshed and saved to {self.xlsx_path.resolve()}")
        self.load_input()

    def _build_qubo_budget_first(self):
        n = len(self.opt_ret)
        Q = np.zeros((n, n), dtype=float)

        scaled_cost = self.opt_cost / float(self.budget_usd)

        ret_excess = self.opt_ret - self.risk_free
        ret_scale = np.max(np.abs(ret_excess)) + 1e-12
        ret_scaled = ret_excess / ret_scale

        sigma_scale = np.max(np.abs(self.Sigma_opt)) + 1e-12
        Sigma_scaled = self.Sigma_opt / sigma_scale

        reward = ret_scaled * scaled_cost
        for i in range(n):
            Q[i, i] += -reward[i]

        for i in range(n):
            Q[i, i] += self.lambda_variance * (scaled_cost[i] ** 2) * Sigma_scaled[i, i]
        for i in range(n):
            for j in range(i + 1, n):
                Q[i, j] += 2.0 * self.lambda_variance * scaled_cost[i] * scaled_cost[j] * Sigma_scaled[i, j]

        for i in range(n):
            Q[i, i] += self.lambda_budget * (scaled_cost[i] ** 2 - 2.0 * scaled_cost[i])
        for i in range(n):
            for j in range(i + 1, n):
                Q[i, j] += 2.0 * self.lambda_budget * scaled_cost[i] * scaled_cost[j]

        self.Q = Q
        self.constant = float(self.lambda_budget)
        self.qubo_meta = {
            "scaled_cost": scaled_cost,
            "ret_scaled": ret_scaled,
            "Sigma_scaled": Sigma_scaled,
            "lambda_budget": float(self.lambda_budget),
            "lambda_variance": float(self.lambda_variance),
            "opt_tickers": list(self.opt_tickers),
            "budget_usd": float(self.budget_usd),
            "ret_scale": float(ret_scale),
            "sigma_scale": float(sigma_scale),
        }

    def build_qubo(self):
        self._ensure_not_stopped()
        self._log("Building Version 4 QUBO problem...")

        xls = pd.ExcelFile(self.xlsx_path)
        options_df = self.assets_df.loc[self.assets_df["Ticker"].notna()].copy()
        options_df["Ticker"] = options_df["Ticker"].astype(str)

        if "Allowed" in options_df.columns:
            options_df = options_df.loc[options_df["Allowed"].fillna(1).astype(int) == 1].copy()

        required_option_cols = ["Ticker", "Approx Cost USD", "Expected Return Proxy", "Annual Volatility"]
        missing_option_cols = [col for col in required_option_cols if col not in options_df.columns]
        if missing_option_cols:
            raise OptimizationError(
                "Assets sheet must define the selectable options directly. "
                f"Missing required columns: {missing_option_cols}"
            )

        for col in ["Approx Cost USD", "Expected Return Proxy", "Annual Volatility"]:
            options_df[col] = pd.to_numeric(options_df[col], errors="coerce")

        if options_df["Approx Cost USD"].isna().any():
            bad = options_df.loc[options_df["Approx Cost USD"].isna(), "Ticker"].tolist()
            raise OptimizationError(f"Approx Cost USD missing for options: {bad}")

        if (options_df["Approx Cost USD"] <= 0).any():
            bad = options_df.loc[options_df["Approx Cost USD"] <= 0, ["Ticker", "Approx Cost USD"]]
            raise OptimizationError(f"Approx Cost USD must be > 0 for all options. Bad rows:\n{bad}")

        if "decision_id" not in options_df.columns:
            options_df["decision_id"] = [
                f"{ticker}_opt{i + 1}" for i, ticker in enumerate(options_df["Ticker"].tolist())
            ]
        else:
            options_df["decision_id"] = options_df["decision_id"].astype(str)

        if "Company" not in options_df.columns:
            options_df["Company"] = options_df["Ticker"]
        if "Option Label" not in options_df.columns:
            options_df["Option Label"] = ""
        if "Shares" not in options_df.columns:
            options_df["Shares"] = np.nan

        options_df = options_df.reset_index(drop=True)
        annual_cov_df = pd.read_excel(xls, sheet_name="AnnualizedCovariance", header=1)
        annual_cov_df = annual_cov_df.rename(columns={annual_cov_df.columns[0]: "Ticker"})
        annual_cov_df = annual_cov_df.loc[annual_cov_df["Ticker"].notna()].copy()
        annual_cov_df["Ticker"] = annual_cov_df["Ticker"].astype(str)
        annual_cov_df = annual_cov_df.set_index("Ticker")

        asset_universe = list(dict.fromkeys(options_df["Ticker"].tolist()))
        missing_cov_assets = [
            ticker for ticker in asset_universe if ticker not in annual_cov_df.index or ticker not in annual_cov_df.columns
        ]
        if missing_cov_assets:
            raise OptimizationError(
                "AnnualizedCovariance sheet is missing these tickers used by the options: "
                + ", ".join(missing_cov_assets)
            )

        Sigma_assets = annual_cov_df.loc[asset_universe, asset_universe].to_numpy(dtype=float)
        ticker_to_idx = {ticker: idx for idx, ticker in enumerate(asset_universe)}

        self.options_df = options_df
        self.annual_cov_df = annual_cov_df
        self.asset_universe = asset_universe
        self.decision_ids = options_df["decision_id"].tolist()
        self.opt_tickers = options_df["Ticker"].tolist()
        self.opt_cost = options_df["Approx Cost USD"].astype(float).to_numpy()
        self.opt_ret = options_df["Expected Return Proxy"].astype(float).to_numpy()

        self.n = len(options_df)
        self.Sigma_opt = np.zeros((self.n, self.n), dtype=float)
        for i in range(self.n):
            ai = ticker_to_idx[self.opt_tickers[i]]
            for j in range(self.n):
                aj = ticker_to_idx[self.opt_tickers[j]]
                self.Sigma_opt[i, j] = Sigma_assets[ai, aj]

        if self.enable_qaoa and self.n > self.qaoa_max_qubits_allowed:
            self._log(
                f"QAOA auto-disabled: n={self.n} exceeds "
                f"qaoa_max_qubits_allowed={self.qaoa_max_qubits_allowed}."
            )
            self.enable_qaoa = False

        if self.enable_classical and self.n > self.classical_max_qubits_allowed:
            self._log(
                f"Classical search auto-disabled: n={self.n} exceeds "
                f"classical_max_qubits_allowed={self.classical_max_qubits_allowed}."
            )
            self.enable_classical = False

        self._log(f"Assets referenced by options: {len(self.asset_universe)}")
        self._log(f"Decision variables from Excel: {self.n}")
        self._log(f"QAOA layers p: {self.qaoa_p}")

        self._build_qubo_budget_first()
        self._log(f"QUBO matrix size: {self.Q.shape}")

    def qubo_value(self, bitvec):
        x = np.asarray(bitvec, dtype=float)
        return float(x @ self.Q @ x + self.constant)

    def qubo_term_breakdown(self, bitvec):
        x = np.asarray(bitvec, dtype=float)
        scaled_cost = np.asarray(self.qubo_meta["scaled_cost"], dtype=float)
        ret_scaled = np.asarray(self.qubo_meta["ret_scaled"], dtype=float)
        Sigma_scaled = np.asarray(self.qubo_meta["Sigma_scaled"], dtype=float)
        lambda_budget = float(self.qubo_meta["lambda_budget"])
        lambda_variance = float(self.qubo_meta["lambda_variance"])

        num_selected = int(x.sum())

        return_term = float(-np.dot(ret_scaled * scaled_cost, x))
        weighted = x * scaled_cost
        risk_term = float(lambda_variance * (weighted @ Sigma_scaled @ weighted))
        budget_term = float(lambda_budget * (weighted.sum() - 1.0) ** 2)
        total = float(return_term + risk_term + budget_term)
        denom = max(num_selected, 1)

        return {
            "return_term": return_term,
            "risk_term": risk_term,
            "budget_term": budget_term,
            "qubo_reconstructed": total,
            "avg_return_term_per_option": float(return_term / denom),
            "avg_risk_term_per_option": float(risk_term / denom),
            "avg_budget_term_per_option": float(budget_term / denom),
        }

    def portfolio_stats(self, bitvec):
        x = np.asarray(bitvec, dtype=float)
        selected_cost = x * self.opt_cost
        total = float(selected_cost.sum())
        num_selected = int(x.sum())
        budget = float(self.budget_usd)

        if total <= 0:
            cash_weight = np.nan if budget <= 0 else 1.0
            port_ret_budget = np.nan if budget <= 0 else float(self.risk_free)
            return {
                "selected_usd": 0.0,
                "budget_gap": -budget,
                "abs_budget_gap": abs(-budget),
                "num_options": 0,
                "num_distinct_assets": 0,
                "portfolio_return": 0.0,
                "portfolio_vol": 0.0,
                "sharpe_like": np.nan,
                "cash_weight": cash_weight,
                "portfolio_return_budget_normalized": port_ret_budget,
                "portfolio_vol_budget_normalized": 0.0 if budget > 0 else np.nan,
                "sharpe_like_budget_normalized": np.nan,
                "max_position_usd": 0.0,
            }

        weights = selected_cost / total
        port_ret = float(np.dot(weights, self.opt_ret))
        port_var = float(weights @ self.Sigma_opt @ weights)
        port_vol = float(np.sqrt(max(port_var, 0.0)))
        sharpe = np.nan if port_vol == 0 else (port_ret - self.risk_free) / port_vol
        if budget > 0:
            budget_weights = selected_cost / budget
            cash_weight = float(1.0 - total / budget)
            port_ret_budget = float(np.dot(budget_weights, self.opt_ret) + cash_weight * self.risk_free)
            port_var_budget = float(budget_weights @ self.Sigma_opt @ budget_weights)
            port_vol_budget = float(np.sqrt(max(port_var_budget, 0.0)))
            sharpe_budget = np.nan if port_vol_budget == 0 else (port_ret_budget - self.risk_free) / port_vol_budget
        else:
            cash_weight = np.nan
            port_ret_budget = np.nan
            port_vol_budget = np.nan
            sharpe_budget = np.nan
        distinct_assets = len({self.opt_tickers[i] for i, bit in enumerate(x) if bit > 0.5})
        max_position = float(selected_cost.max())
        budget_gap = total - budget

        return {
            "selected_usd": total,
            "budget_gap": budget_gap,
            "abs_budget_gap": abs(budget_gap),
            "num_options": num_selected,
            "num_distinct_assets": distinct_assets,
            "portfolio_return": port_ret,
            "portfolio_vol": port_vol,
            "sharpe_like": sharpe,
            "cash_weight": cash_weight,
            "portfolio_return_budget_normalized": port_ret_budget,
            "portfolio_vol_budget_normalized": port_vol_budget,
            "sharpe_like_budget_normalized": sharpe_budget,
            "max_position_usd": max_position,
        }

    def row_is_feasible(self, stats):
        return float(stats.get("abs_budget_gap", np.inf)) <= float(self.qaoa_feasibility_budget_tolerance_usd)

    def qaoa_export_priority(self, row):
        if self.qaoa_export_sort_by == "probability":
            return float(row.get("probability", 0.0))
        if self.qaoa_export_sort_by == "qubo_value":
            return -float(row.get("qubo_value", np.inf))
        if self.qaoa_export_sort_by == "sharpe_like":
            val = row.get("sharpe_like", np.nan)
            return -1e99 if pd.isna(val) else float(val)
        if self.qaoa_export_sort_by == "portfolio_return":
            return float(row.get("portfolio_return", -np.inf))
        if self.qaoa_export_sort_by == "abs_budget_gap":
            return -float(row.get("abs_budget_gap", np.inf))
        return float(row.get("probability", 0.0))

    def sort_qaoa_export(self, df: pd.DataFrame) -> pd.DataFrame:
        if df is None or len(df) == 0:
            return pd.DataFrame(columns=self.candidate_cols)

        df = df.copy()
        if self.qaoa_export_sort_by == "probability":
            primary, ascending_primary = "probability", False
        elif self.qaoa_export_sort_by == "qubo_value":
            primary, ascending_primary = "qubo_value", True
        elif self.qaoa_export_sort_by == "sharpe_like":
            primary, ascending_primary = "sharpe_like", False
        elif self.qaoa_export_sort_by == "portfolio_return":
            primary, ascending_primary = "portfolio_return", False
        elif self.qaoa_export_sort_by == "abs_budget_gap":
            primary, ascending_primary = "abs_budget_gap", True
        else:
            primary, ascending_primary = "probability", False

        sort_cols = []
        ascending = []
        if primary in df.columns:
            sort_cols.append(primary)
            ascending.append(ascending_primary)
        if "qubo_value" in df.columns and primary != "qubo_value":
            sort_cols.append("qubo_value")
            ascending.append(True)
        if "abs_budget_gap" in df.columns and primary != "abs_budget_gap":
            sort_cols.append("abs_budget_gap")
            ascending.append(True)
        if "probability" in df.columns and primary != "probability":
            sort_cols.append("probability")
            ascending.append(False)
        if "portfolio_return" in df.columns and primary != "portfolio_return":
            sort_cols.append("portfolio_return")
            ascending.append(False)
        if "bitstring" in df.columns:
            sort_cols.append("bitstring")
            ascending.append(True)

        return df.sort_values(sort_cols, ascending=ascending).reset_index(drop=True)

    def sort_candidates(self, df: pd.DataFrame) -> pd.DataFrame:
        if df is None or len(df) == 0:
            return pd.DataFrame(columns=self.candidate_cols)

        df = df.copy()
        sort_cols = []
        ascending = []
        if "qubo_value" in df.columns:
            sort_cols.append("qubo_value")
            ascending.append(True)
        if "abs_budget_gap" in df.columns:
            sort_cols.append("abs_budget_gap")
            ascending.append(True)
        if "portfolio_return" in df.columns:
            sort_cols.append("portfolio_return")
            ascending.append(False)
        if "bitstring" in df.columns:
            sort_cols.append("bitstring")
            ascending.append(True)
        return df.sort_values(sort_cols, ascending=ascending).reset_index(drop=True)

    def random_solution(self):
        return np.random.randint(0, 2, size=self.n, dtype=int)

    def greedy_improve_unconstrained(self, x):
        x = x.copy()
        improved = True
        neighbor_evals = 0
        while improved and neighbor_evals < self.classical_max_neighbor_evals:
            self._ensure_not_stopped()
            improved = False
            current_val = self.qubo_value(x)
            for idx in range(len(x)):
                if neighbor_evals >= self.classical_max_neighbor_evals:
                    break
                y = x.copy()
                y[idx] = 1 - y[idx]
                y_val = self.qubo_value(y)
                neighbor_evals += 1
                if y_val < current_val - 1e-9:
                    x = y
                    current_val = y_val
                    improved = True
        return x

    def run_classical_search(self):
        self.classical_results = pd.DataFrame(columns=self.candidate_cols)
        if not self.enable_classical:
            self._log("Classical search disabled.")
            return

        self._log("Running Version 4 classical heuristic...")
        progress_plan = self._progress_plan()
        phase_start = float(progress_plan["classical_start"])
        phase_end = float(progress_plan["classical_end"])
        candidate_dict = {}
        total_steps = max(self.random_search_samples, 0) + min(self.n, max(self.local_search_starts, 0))
        completed_steps = 0

        for _ in range(max(self.random_search_samples, 0)):
            self._ensure_not_stopped()
            x = self.random_solution()
            x = self.greedy_improve_unconstrained(x)
            bitstring = "".join(map(str, x.astype(int)))
            stats = self.portfolio_stats(x)
            term_stats = self.qubo_term_breakdown(x)
            candidate_dict[bitstring] = {
                "bitstring": bitstring,
                "source": "classical_heuristic",
                "probability": np.nan,
                "qubo_value": self.qubo_value(x),
                **term_stats,
                **stats,
            }
            completed_steps += 1
            if completed_steps == 1 or completed_steps % 25 == 0 or completed_steps == total_steps:
                self._update_progress_fraction(
                    phase_start,
                    phase_end,
                    completed_steps,
                    total_steps,
                    f"Running classical search... {completed_steps}/{total_steps}",
                )

        for start_idx in range(min(self.n, max(self.local_search_starts, 0))):
            self._ensure_not_stopped()
            x = np.zeros(self.n, dtype=int)
            x[start_idx] = 1
            x = self.greedy_improve_unconstrained(x)
            bitstring = "".join(map(str, x.astype(int)))
            stats = self.portfolio_stats(x)
            term_stats = self.qubo_term_breakdown(x)
            candidate_dict[bitstring] = {
                "bitstring": bitstring,
                "source": "classical_heuristic",
                "probability": np.nan,
                "qubo_value": self.qubo_value(x),
                **term_stats,
                **stats,
            }
            completed_steps += 1
            if completed_steps == 1 or completed_steps % 10 == 0 or completed_steps == total_steps:
                self._update_progress_fraction(
                    phase_start,
                    phase_end,
                    completed_steps,
                    total_steps,
                    f"Running classical search... {completed_steps}/{total_steps}",
                )

        self.classical_results = self.sort_candidates(pd.DataFrame(candidate_dict.values()))
        self._log(f"Classical candidate count: {len(self.classical_results)}")

    @staticmethod
    def _qubo_to_ising(Q, constant=0.0):
        n = Q.shape[0]
        h = np.zeros(n, dtype=float)
        J = {}
        offset = float(constant)
        for i in range(n):
            qii = Q[i, i]
            offset += qii / 2
            h[i] += -qii / 2
        for i in range(n):
            for j in range(i + 1, n):
                qij = Q[i, j]
                if abs(qij) > 1e-12:
                    offset += qij / 4
                    h[i] += -qij / 4
                    h[j] += -qij / 4
                    J[(i, j)] = qij / 4
        return h, J, offset

    def _adaptive_qaoa_gamma_scale(self) -> float:
        if self.Q is None or len(self.Q) == 0:
            return 0.30

        h, J, _ = self._qubo_to_ising(self.Q, self.constant)
        if len(h) == 0:
            return 0.30

        row_strength = np.abs(np.asarray(h, dtype=float))
        for (i, j), value in J.items():
            coupling = abs(float(value))
            row_strength[i] += coupling
            row_strength[j] += coupling

        max_strength = float(np.max(row_strength)) if len(row_strength) else 0.0
        if not np.isfinite(max_strength) or max_strength <= 1e-12:
            return 0.30

        # Keep the initial cost-phase angle within a conservative first-quarter
        # period of the current Ising scale instead of using a fixed gamma range.
        return float(np.pi / (4.0 * max_strength))

    def _base_qaoa_init(self, layers: int) -> np.ndarray:
        return self._deterministic_qaoa_seed_family(layers)[0][1]

    def _deterministic_qaoa_seed_family(self, layers: int) -> list[tuple[str, np.ndarray]]:
        gamma_scale = self._adaptive_qaoa_gamma_scale()
        gamma_mag = np.linspace(0.5 * gamma_scale, gamma_scale, layers)
        beta_low = np.linspace(0.40, 0.20, layers)
        beta_high = np.pi - beta_low
        family = [
            ("linear_neg", np.concatenate([-gamma_mag, beta_low]).astype(float)),
            ("linear_pos_beta_mirror", np.concatenate([gamma_mag, beta_high]).astype(float)),
            ("linear_pos", np.concatenate([gamma_mag, beta_low]).astype(float)),
            ("linear_neg_beta_mirror", np.concatenate([-gamma_mag, beta_high]).astype(float)),
        ]
        return family

    @staticmethod
    def _lift_p1_seed(p1_seed: np.ndarray, target_layers: int) -> np.ndarray:
        lifted = np.zeros(2 * target_layers, dtype=float)
        lifted[0] = float(p1_seed[0])
        lifted[target_layers] = float(p1_seed[1])
        return lifted.astype(float)

    def _expand_qaoa_params(self, prev_params: np.ndarray, target_layers: int) -> np.ndarray:
        if target_layers <= 1 or prev_params is None or len(prev_params) == 0:
            return self._base_qaoa_init(target_layers)

        prev_layers = len(prev_params) // 2
        prev_gammas = np.asarray(prev_params[:prev_layers], dtype=float)
        prev_betas = np.asarray(prev_params[prev_layers:], dtype=float)

        gammas = np.empty(target_layers, dtype=float)
        betas = np.empty(target_layers, dtype=float)
        gammas[:prev_layers] = prev_gammas
        betas[:prev_layers] = prev_betas
        # Preserve the lower-depth solution exactly by padding the new layer(s)
        # with zeros. Repeating the last layer changes the circuit and can make
        # the higher-p warm start substantially worse than the lower-p optimum.
        gammas[prev_layers:] = 0.0
        betas[prev_layers:] = 0.0
        return np.concatenate([gammas, betas]).astype(float)

    def _qaoa_initial_candidates(
        self,
        layers: int,
        prev_best: Optional[np.ndarray],
        rng,
        preferred_seed: Optional[np.ndarray] = None,
    ) -> list[tuple[str, np.ndarray]]:
        candidates = []
        deterministic_seeds = list(self._deterministic_qaoa_seed_family(layers))
        if preferred_seed is not None:
            preferred_entry = ("p1_grid_seed", np.asarray(preferred_seed, dtype=float))
            if layers == 1:
                deterministic_seeds.insert(0, preferred_entry)
            else:
                # A strong p=1 seed does not reliably map to a strong direct
                # p>1 basin, so keep native p-layer seeds ahead of the lifted
                # p=1 candidate when the restart budget is tight.
                deterministic_seeds.append(preferred_entry)
        target_count = max(1, int(self.qaoa_multistart_restarts))

        if prev_best is not None and self.qaoa_layerwise_warm_start:
            warm = self._expand_qaoa_params(prev_best, layers)
            candidates.append(("warm_start", warm))

            # For higher-p stages, spend the default restart budget exploring
            # the neighborhood of the warmed lower-p solution before trying
            # colder deterministic seeds that can land far away.
            for restart_idx in range(max(target_count - 1, 0)):
                perturbed = warm + rng.normal(0.0, self.qaoa_restart_perturbation, size=2 * layers)
                candidates.append((f"warm_perturb_{restart_idx + 1}", perturbed.astype(float)))

            for init_type, params in deterministic_seeds:
                candidates.append((init_type, params))
        else:
            for init_type, params in deterministic_seeds:
                if len(candidates) >= target_count:
                    break
                candidates.append((init_type, params))
            anchor = candidates[0][1] if candidates else self._base_qaoa_init(layers)
            for restart_idx in range(max(target_count - len(candidates), 0)):
                perturbed = anchor + rng.normal(0.0, self.qaoa_restart_perturbation, size=2 * layers)
                candidates.append((f"random_perturb_{restart_idx + 1}", perturbed.astype(float)))

        unique_candidates = []
        seen = set()
        for init_type, params in candidates:
            key = tuple(np.round(np.asarray(params, dtype=float), 12))
            if key in seen:
                continue
            seen.add(key)
            unique_candidates.append((init_type, np.asarray(params, dtype=float)))
            if len(unique_candidates) >= target_count:
                break
        return unique_candidates

    def _qaoa_restart_count(self, has_prev_best: bool) -> int:
        return max(1, int(self.qaoa_multistart_restarts))

    def _run_exact_p1_presearch(self, make_qaoa_energy, offset: float, qaoa_progress_start: float, qaoa_opt_end: float):
        self.qaoa_p1_presearch_used = False
        self.qaoa_p1_presearch_gamma = np.nan
        self.qaoa_p1_presearch_beta = np.nan
        self.qaoa_p1_presearch_energy = np.nan
        self.qaoa_p1_presearch_expected_qubo = np.nan

        gamma_grid = np.linspace(
            self.qaoa_exact_p1_presearch_gamma_min,
            self.qaoa_exact_p1_presearch_gamma_max,
            self.qaoa_exact_p1_presearch_gamma_points,
        )
        beta_grid = np.linspace(
            self.qaoa_exact_p1_presearch_beta_min,
            self.qaoa_exact_p1_presearch_beta_max,
            self.qaoa_exact_p1_presearch_beta_points,
        )

        total = max(len(gamma_grid) * len(beta_grid), 1)
        qaoa_energy_p1 = make_qaoa_energy(1)
        best_energy = np.inf
        best_pair = None
        counter = 0

        self._log(
            "Running exact p=1 grid presearch for QAOA initialization "
            f"({len(gamma_grid)}x{len(beta_grid)} points)..."
        )
        for beta in beta_grid:
            for gamma in gamma_grid:
                self._ensure_not_stopped()
                energy = float(qaoa_energy_p1(np.array([gamma, beta], dtype=float)))
                counter += 1
                if energy < best_energy:
                    best_energy = energy
                    best_pair = (float(gamma), float(beta))
                if counter % max(1, total // 20) == 0 or counter == total:
                    self._update_progress_fraction(
                        qaoa_progress_start,
                        qaoa_opt_end,
                        counter,
                        total,
                        f"Running exact p=1 grid presearch... {counter}/{total}",
                    )

        if best_pair is None:
            return None, 0

        self.qaoa_p1_presearch_used = True
        self.qaoa_p1_presearch_gamma = float(best_pair[0])
        self.qaoa_p1_presearch_beta = float(best_pair[1])
        self.qaoa_p1_presearch_energy = float(best_energy)
        self.qaoa_p1_presearch_expected_qubo = float(best_energy + offset)
        self._log(
            "Exact p=1 presearch best seed: "
            f"gamma={self.qaoa_p1_presearch_gamma:.6f}  "
            f"beta={self.qaoa_p1_presearch_beta:.6f}  "
            f"energy={self.qaoa_p1_presearch_energy:.6f}  "
            f"expected_qubo={self.qaoa_p1_presearch_expected_qubo:.6f}"
        )
        return np.array([self.qaoa_p1_presearch_gamma, self.qaoa_p1_presearch_beta], dtype=float), total

    def run_qaoa(self):
        self.samples_df = pd.DataFrame(columns=self.candidate_cols)

        history_columns = ["iteration", "layer_stage", "restart", "init_type"]
        history_columns += [f"gamma_{i + 1}" for i in range(self.qaoa_p)]
        history_columns += [f"beta_{i + 1}" for i in range(self.qaoa_p)]
        history_columns += ["energy", "expected_qubo", "best_energy", "best_expected_qubo", "elapsed_sec", "eval_sec"]
        self.history_df = pd.DataFrame(columns=history_columns)

        self.best_gammas = np.full(self.qaoa_p, np.nan, dtype=float)
        self.best_betas = np.full(self.qaoa_p, np.nan, dtype=float)
        self.qaoa_mode = "disabled"
        self.qaoa_total_nonzero_states = 0
        self.qaoa_total_states_considered = 0
        self.qaoa_final_energy = np.nan
        self.qaoa_expected_qubo = np.nan
        self.qaoa_max_state_probability = np.nan
        self.qaoa_top10_probability_mass = np.nan
        self.qaoa_top_probability_state_qubo = np.nan
        self.qaoa_best_exported_qubo = np.nan
        self.qaoa_best_exported_probability = np.nan
        self.qaoa_p1_presearch_used = False
        self.qaoa_p1_presearch_gamma = np.nan
        self.qaoa_p1_presearch_beta = np.nan
        self.qaoa_p1_presearch_energy = np.nan
        self.qaoa_p1_presearch_expected_qubo = np.nan

        if not self.enable_qaoa:
            self._log("QAOA disabled.")
            return

        if qml is None:
            raise OptimizationError("PennyLane is required when QAOA is enabled.")

        self._log(f"Running Version 4 QAOA (p={self.qaoa_p})...")
        h, J, offset = self._qubo_to_ising(self.Q, self.constant)

        coeffs = []
        ops = []
        for i, coeff in enumerate(h):
            if abs(coeff) > 1e-12:
                coeffs.append(float(coeff))
                ops.append(qml.PauliZ(i))
        for (i, j), coeff in J.items():
            if abs(coeff) > 1e-12:
                coeffs.append(float(coeff))
                ops.append(qml.PauliZ(i) @ qml.PauliZ(j))
        if not coeffs:
            coeffs = [0.0]
            ops = [qml.Identity(0)]

        H = qml.Hamiltonian(coeffs, ops)
        exact_mode = self.n <= self.qaoa_exact_probability_max_qubits
        self.qaoa_mode = "exact_probs" if exact_mode else f"shots_{self.qaoa_shots}"
        dev = qml.device("lightning.qubit", wires=self.n, shots=None if exact_mode else self.qaoa_shots)

        def apply_qaoa_layer(gamma, beta):
            for i, coeff in enumerate(h):
                if abs(coeff) > 1e-12:
                    qml.RZ(2 * gamma * coeff, wires=i)
            for (i, j), coeff in J.items():
                if abs(coeff) > 1e-12:
                    qml.CNOT(wires=[i, j])
                    qml.RZ(2 * gamma * coeff, wires=j)
                    qml.CNOT(wires=[i, j])
            for wire in range(self.n):
                qml.RX(2 * beta, wires=wire)

        def make_qaoa_energy(active_layers: int):
            @qml.qnode(dev, interface="autograd")
            def qaoa_energy(params):
                gammas = params[:active_layers]
                betas = params[active_layers:]
                for wire in range(self.n):
                    qml.Hadamard(wires=wire)
                for layer in range(active_layers):
                    apply_qaoa_layer(gammas[layer], betas[layer])
                return qml.expval(H)

            return qaoa_energy

        if exact_mode:
            @qml.qnode(dev, interface="autograd")
            def qaoa_probs(params):
                gammas = params[: self.qaoa_p]
                betas = params[self.qaoa_p :]
                for wire in range(self.n):
                    qml.Hadamard(wires=wire)
                for layer in range(self.qaoa_p):
                    apply_qaoa_layer(gammas[layer], betas[layer])
                return qml.probs(wires=range(self.n))

        else:

            @qml.qnode(dev, interface="autograd")
            def qaoa_samples(params):
                gammas = params[: self.qaoa_p]
                betas = params[self.qaoa_p :]
                for wire in range(self.n):
                    qml.Hadamard(wires=wire)
                for layer in range(self.qaoa_p):
                    apply_qaoa_layer(gammas[layer], betas[layer])
                return qml.sample(wires=range(self.n))

        progress_rows = []
        start_time = time.time()
        rng = np.random.default_rng(self.rng_seed)
        eval_counter = 0
        stage_sequence = list(range(1, self.qaoa_p + 1)) if self.qaoa_layerwise_warm_start else [self.qaoa_p]
        presearch_seed = None
        planned_presearch_evals = 0
        if exact_mode and self.qaoa_exact_p1_presearch and self.n <= self.qaoa_exact_p1_presearch_max_qubits:
            planned_presearch_evals = (
                self.qaoa_exact_p1_presearch_gamma_points * self.qaoa_exact_p1_presearch_beta_points
            )
        stage_restart_plan = []
        for stage_index, active_layers in enumerate(stage_sequence):
            has_prev_best = stage_index > 0 and self.qaoa_layerwise_warm_start
            restart_count = self._qaoa_restart_count(has_prev_best)
            stage_restart_plan.append((active_layers, restart_count))
        total_planned_restarts = sum(restart_count for _, restart_count in stage_restart_plan)
        estimated_evals = max(planned_presearch_evals + self.qaoa_maxiter * total_planned_restarts, 1)
        progress_plan = self._progress_plan()
        qaoa_progress_start = float(progress_plan["qaoa_start"])
        qaoa_opt_end = float(progress_plan["qaoa_opt_end"])
        qaoa_extract_end = float(progress_plan["qaoa_extract_end"])

        def make_objective_logged(active_layers: int, restart_idx: int, restart_total: int, stage_index: int, stage_total: int, init_type: str):
            qaoa_energy = make_qaoa_energy(active_layers)

            def objective_logged(x):
                nonlocal eval_counter
                self._ensure_not_stopped()
                eval_start = time.time()
                energy = float(qaoa_energy(x))
                eval_sec = time.time() - eval_start
                expected_qubo = float(energy + offset)
                elapsed = time.time() - start_time
                global_best = min([energy] + [row["energy"] for row in progress_rows]) if progress_rows else energy
                stage_prior = [
                    row["energy"]
                    for row in progress_rows
                    if int(row.get("layer_stage", -1)) == active_layers and int(row.get("restart", -1)) == restart_idx
                ]
                stage_best = min([energy] + stage_prior) if stage_prior else energy
                global_best_qubo = float(global_best + offset)
                stage_best_qubo = float(stage_best + offset)

                row = {
                    "iteration": len(progress_rows) + 1,
                    "layer_stage": active_layers,
                    "restart": restart_idx,
                    "init_type": init_type,
                }
                for idx in range(self.qaoa_p):
                    row[f"gamma_{idx + 1}"] = float(x[idx]) if idx < active_layers else np.nan
                for idx in range(self.qaoa_p):
                    beta_val = float(x[active_layers + idx]) if idx < active_layers else np.nan
                    row[f"beta_{idx + 1}"] = beta_val
                row["energy"] = energy
                row["expected_qubo"] = expected_qubo
                row["best_energy"] = global_best
                row["best_expected_qubo"] = global_best_qubo
                row["stage_best_energy"] = stage_best
                row["stage_best_expected_qubo"] = stage_best_qubo
                row["elapsed_sec"] = elapsed
                row["eval_sec"] = eval_sec
                progress_rows.append(row)

                gamma_str = "  ".join([f"g{i + 1}={x[i]: .5f}" for i in range(active_layers)])
                beta_str = "  ".join([f"b{i + 1}={x[active_layers + i]: .5f}" for i in range(active_layers)])
                self._log(
                    f"stage={active_layers}/{self.qaoa_p}  restart={restart_idx} ({init_type})  "
                    f"iter={len(progress_rows):>3}  {gamma_str}  {beta_str}  "
                    f"energy={energy: .12f}  exp_qubo={expected_qubo: .12f}  "
                    f"stage_best={stage_best: .12f}  stage_best_qubo={stage_best_qubo: .12f}  "
                    f"global_best={global_best: .12f}  global_best_qubo={global_best_qubo: .12f}  "
                    f"eval={eval_sec: .1f}s  elapsed={elapsed: .1f}s"
                )
                eval_counter += 1
                eta_seconds = self._qaoa_eta_seconds_remaining(progress_rows, stage_restart_plan)
                eta = self._format_eta(eta_seconds)
                self._update_progress_fraction(
                    qaoa_progress_start,
                    qaoa_opt_end,
                    eval_counter,
                    estimated_evals,
                    f"Running QAOA... stage {stage_index}/{stage_total} (p={active_layers}), "
                    f"restart {restart_idx}/{restart_total}, eval {eval_counter}/{estimated_evals} | ETA ~ {eta}",
                )
                return energy

            return objective_logged

        if planned_presearch_evals > 0:
            presearch_seed, actual_presearch_evals = self._run_exact_p1_presearch(
                make_qaoa_energy,
                offset,
                qaoa_progress_start,
                qaoa_opt_end,
            )
            eval_counter += int(actual_presearch_evals)

        best_params = None
        for stage_index, (active_layers, planned_restart_count) in enumerate(stage_restart_plan, start=1):
            self._log(
                f"QAOA stage {stage_index}/{len(stage_restart_plan)} "
                f"(p={active_layers}): optimizing {active_layers} layer(s)..."
            )
            stage_best_res = None
            stage_best_energy = np.inf
            preferred_seed = None
            if presearch_seed is not None:
                if active_layers == 1:
                    preferred_seed = presearch_seed
                elif stage_index == 1 and not self.qaoa_layerwise_warm_start:
                    preferred_seed = self._lift_p1_seed(presearch_seed, active_layers)
            init_candidates = self._qaoa_initial_candidates(active_layers, best_params, rng, preferred_seed)

            for restart_idx, (init_type, x0) in enumerate(init_candidates, start=1):
                self._ensure_not_stopped()
                self._log(
                    f"QAOA stage {stage_index}/{len(stage_restart_plan)} "
                    f"(p={active_layers}): restart {restart_idx}/{len(init_candidates)} "
                    f"using {init_type}."
                )
                objective_logged = make_objective_logged(
                    active_layers,
                    restart_idx,
                    len(init_candidates),
                    stage_index,
                    len(stage_restart_plan),
                    init_type,
                )
                res = minimize(objective_logged, x0=x0, method="COBYLA", options={"maxiter": self.qaoa_maxiter})
                final_energy = float(res.fun)
                if final_energy < stage_best_energy:
                    stage_best_energy = final_energy
                    stage_best_res = res

            if stage_best_res is None:
                raise OptimizationError(f"QAOA optimization failed at stage {active_layers}.")

            best_params = np.array(stage_best_res.x, dtype=float)
            best_gamma_str = "  ".join([f"g{i + 1}={best_params[i]: .5f}" for i in range(active_layers)])
            best_beta_str = "  ".join(
                [f"b{i + 1}={best_params[active_layers + i]: .5f}" for i in range(active_layers)]
            )
            self._log(
                f"Completed QAOA stage {stage_index}/{len(stage_restart_plan)} (p={active_layers}): "
                f"best_energy={stage_best_energy: .6f}  "
                f"best_expected_qubo={stage_best_energy + offset: .6f}  "
                f"{best_gamma_str}  {best_beta_str}"
            )

        self.best_gammas = np.array(best_params[: self.qaoa_p], dtype=float)
        self.best_betas = np.array(best_params[self.qaoa_p :], dtype=float)
        self.history_df = pd.DataFrame(progress_rows)

        q_rows = []
        params_opt = np.concatenate([self.best_gammas, self.best_betas])
        final_energy_fn = make_qaoa_energy(self.qaoa_p)
        self.qaoa_final_energy = float(final_energy_fn(params_opt))
        self.qaoa_expected_qubo = float(self.qaoa_final_energy + offset)

        if exact_mode:
            self._progress("Extracting exact QAOA probabilities...", qaoa_opt_end)
            probs = np.array(qaoa_probs(params_opt))
            self.qaoa_total_states_considered = int(len(probs))
            self.qaoa_total_nonzero_states = int(np.count_nonzero(probs > 0))
            self.qaoa_max_state_probability = float(np.max(probs)) if len(probs) else np.nan
            topk_mass_count = min(10, len(probs))
            if topk_mass_count > 0:
                topk_mass_idx = np.argpartition(probs, -topk_mass_count)[-topk_mass_count:]
                self.qaoa_top10_probability_mass = float(np.sum(probs[topk_mass_idx]))
            self._log(
                f"Exact QAOA diagnostics: expected_qubo={self.qaoa_expected_qubo:.12f}  "
                f"max_state_probability={self.qaoa_max_state_probability:.12f}  "
                f"top10_probability_mass={self.qaoa_top10_probability_mass:.12f}"
            )

            import heapq

            def build_exact_row(state_idx: int, prob: float):
                bitstring = format(state_idx, f"0{self.n}b")
                bits = np.array(list(map(int, bitstring)), dtype=int)
                stats = self.portfolio_stats(bits)
                if self.qaoa_export_feasible_only and not self.row_is_feasible(stats):
                    return None
                term_stats = self.qubo_term_breakdown(bits)
                return {
                    "bitstring": "".join(map(str, bits.astype(int))),
                    "source": f"qaoa_full_pennylane_p{self.qaoa_p}",
                    "probability": float(prob),
                    "qubo_value": self.qubo_value(bits),
                    **term_stats,
                    **stats,
                }

            fast_exact_probability_path = (
                self.qaoa_export_mode == "top_k"
                and self.qaoa_export_sort_by == "probability"
                and not self.qaoa_export_feasible_only
            )

            if fast_exact_probability_path:
                self._log(
                    f"Fast exact export path: selecting top {self.qaoa_max_export_rows} states by probability "
                    f"from {len(probs):,} exact states."
                )
                top_count = min(self.qaoa_max_export_rows, len(probs))
                if top_count > 0:
                    top_idx = np.argpartition(probs, -top_count)[-top_count:]
                    top_idx = top_idx[np.argsort(probs[top_idx])[::-1]]
                    for rank_pos, idx in enumerate(top_idx, start=1):
                        self._ensure_not_stopped()
                        prob = float(probs[idx])
                        if prob < self.qaoa_min_probability_to_export:
                            break
                        row = build_exact_row(int(idx), prob)
                        if row is not None:
                            q_rows.append(row)
                        if rank_pos == 1 or rank_pos % 100 == 0 or rank_pos == top_count:
                            self._update_progress_fraction(
                                qaoa_opt_end,
                                qaoa_extract_end,
                                rank_pos,
                                max(top_count, 1),
                                f"Extracting exact QAOA states... {rank_pos}/{top_count}",
                            )
            else:
                exact_heap = []
                exact_counter = 0
                self._log(
                    f"Full exact export path: scanning {len(probs):,} exact states. "
                    "This can take a while for larger qubit counts."
                )

                for idx, prob in enumerate(probs):
                    if idx % 256 == 0:
                        self._ensure_not_stopped()
                    if idx == 0 or (idx + 1) % 65536 == 0 or (idx + 1) == len(probs):
                        self._update_progress_fraction(
                            qaoa_opt_end,
                            qaoa_extract_end,
                            idx + 1,
                            len(probs),
                            f"Scanning exact QAOA states... {idx + 1:,}/{len(probs):,}",
                        )
                    if prob < self.qaoa_min_probability_to_export:
                        continue

                    row = build_exact_row(int(idx), float(prob))
                    if row is None:
                        continue

                    if self.qaoa_export_mode == "all_filtered":
                        q_rows.append(row)
                    else:
                        priority = self.qaoa_export_priority(row)
                        payload = (priority, exact_counter, row)
                        if len(exact_heap) < self.qaoa_max_export_rows:
                            heapq.heappush(exact_heap, payload)
                        elif priority > exact_heap[0][0]:
                            heapq.heapreplace(exact_heap, payload)
                        exact_counter += 1

                if self.qaoa_export_mode == "top_k":
                    q_rows = [item[2] for item in exact_heap]

        else:
            self._progress("Extracting sampled QAOA solutions...", qaoa_opt_end)
            raw_samples = np.asarray(qaoa_samples(params_opt))
            if raw_samples.ndim == 1:
                raw_samples = raw_samples.reshape(1, -1)

            sample_counts = {}
            for sample in raw_samples:
                bits = np.asarray(sample, dtype=int).reshape(-1)
                bitstring = "".join(map(str, bits.astype(int)))
                sample_counts[bitstring] = sample_counts.get(bitstring, 0) + 1

            total_shots = sum(sample_counts.values())
            self.qaoa_total_states_considered = int(total_shots)
            self.qaoa_total_nonzero_states = int(len(sample_counts))
            total_unique_samples = max(len(sample_counts), 1)

            for idx, (bitstring, count) in enumerate(sample_counts.items(), start=1):
                if idx % 32 == 0:
                    self._ensure_not_stopped()
                bits = np.array(list(map(int, bitstring)), dtype=int)
                stats = self.portfolio_stats(bits)
                if self.qaoa_export_feasible_only and not self.row_is_feasible(stats):
                    continue
                term_stats = self.qubo_term_breakdown(bits)
                q_rows.append(
                    {
                        "bitstring": bitstring,
                        "source": f"qaoa_shot_pennylane_p{self.qaoa_p}",
                        "probability": float(count / total_shots),
                        "qubo_value": self.qubo_value(bits),
                        **term_stats,
                        **stats,
                    }
                )
                if idx == 1 or idx % 100 == 0 or idx == total_unique_samples:
                    self._update_progress_fraction(
                        qaoa_opt_end,
                        qaoa_extract_end,
                        idx,
                        total_unique_samples,
                        f"Extracting sampled QAOA solutions... {idx}/{total_unique_samples}",
                    )

        self.samples_df = self.sort_qaoa_export(pd.DataFrame(q_rows))
        if len(self.samples_df) > self.qaoa_max_export_rows:
            self.samples_df = self.samples_df.head(self.qaoa_max_export_rows).copy()
        if len(self.samples_df):
            top_prob_row = (
                self.samples_df.sort_values("probability", ascending=False).iloc[0]
                if "probability" in self.samples_df.columns
                else None
            )
            best_qubo_row = self.samples_df.sort_values("qubo_value").iloc[0]
            self.qaoa_best_exported_qubo = float(best_qubo_row.get("qubo_value", np.nan))
            self.qaoa_best_exported_probability = float(best_qubo_row.get("probability", np.nan))
            if top_prob_row is not None:
                self.qaoa_top_probability_state_qubo = float(top_prob_row.get("qubo_value", np.nan))
            self._log(
                f"QAOA exported-state diagnostics: best_exported_qubo={self.qaoa_best_exported_qubo:.12f}  "
                f"best_exported_probability={self.qaoa_best_exported_probability:.12f}  "
                f"top_probability_state_qubo={self.qaoa_top_probability_state_qubo:.12f}"
            )

        self._log(f"QAOA candidate count: {len(self.samples_df)}")
        self._log(f"QAOA total states considered: {self.qaoa_total_states_considered}")
        self._log(f"QAOA nonzero states encountered: {self.qaoa_total_nonzero_states}")

    def exploded_portfolio_rows(self, rank, source, bitstring, bitvec):
        term_stats = self.qubo_term_breakdown(bitvec)
        stats = self.portfolio_stats(bitvec)
        rows = []
        for idx, flag in enumerate(bitvec):
            if flag != 1:
                continue
            rows.append(
                {
                    "rank": rank,
                    "source": source,
                    "bitstring": bitstring,
                    "Ticker": self.options_df.loc[idx, "Ticker"],
                    "Company": self.options_df.loc[idx, "Company"]
                    if "Company" in self.options_df.columns
                    else self.options_df.loc[idx, "Ticker"],
                    "Option Label": self.options_df.loc[idx, "Option Label"]
                    if "Option Label" in self.options_df.columns
                    else "",
                    "Shares": self.options_df.loc[idx, "Shares"] if "Shares" in self.options_df.columns else np.nan,
                    "Approx Cost USD": float(self.options_df.loc[idx, "Approx Cost USD"]),
                    "Expected Return Proxy": float(self.options_df.loc[idx, "Expected Return Proxy"]),
                    "Annual Volatility": float(self.options_df.loc[idx, "Annual Volatility"]),
                    "decision_id": self.options_df.loc[idx, "decision_id"],
                    **term_stats,
                    **stats,
                    "qubo_value": self.qubo_value(bitvec),
                }
            )
        return rows

    def generate_results(self):
        self._ensure_not_stopped()
        self._log("Generating Version 4 result tables...")

        classical_pool = (
            self.classical_results.head(self.overview_classical_pool).copy()
            if len(self.classical_results)
            else self.classical_results.copy()
        )
        qaoa_pool = (
            self.sort_candidates(self.samples_df).head(self.overview_qaoa_pool).copy()
            if len(self.samples_df)
            else self.samples_df.copy()
        )

        combined_frames = []
        if len(classical_pool):
            combined_frames.append(classical_pool[self.candidate_cols])
        if len(qaoa_pool):
            combined_frames.append(qaoa_pool[self.candidate_cols])

        combined = (
            pd.concat(combined_frames, ignore_index=True)
            if combined_frames
            else pd.DataFrame(columns=self.candidate_cols)
        )
        combined = self.sort_candidates(combined).drop_duplicates(subset=["bitstring"], keep="first").reset_index(drop=True)

        self.overview_df = self.sort_candidates(combined).head(self.top_n_export).copy()
        if len(self.overview_df):
            self.overview_df.insert(0, "rank", range(1, len(self.overview_df) + 1))
        else:
            self.overview_df = pd.DataFrame(columns=["rank"] + self.candidate_cols)

        exploded_rows = []
        for _, row in self.overview_df.iterrows():
            bitstring = row["bitstring"]
            source = row["source"]
            rank = int(row["rank"])
            bitvec = np.array(list(map(int, bitstring)), dtype=int)
            exploded_rows.extend(self.exploded_portfolio_rows(rank, source, bitstring, bitvec))
        self.portfolios_df = pd.DataFrame(exploded_rows)

        solver_frames = []
        if len(self.classical_results):
            best_classical = self.sort_candidates(self.classical_results).head(1).copy()
            solver_frames.append(best_classical.assign(solver="Classical Heuristic"))
        if len(self.samples_df):
            best_qaoa = self.sort_candidates(self.samples_df).head(1).copy()
            solver_frames.append(best_qaoa.assign(solver=f"QAOA PennyLane p={self.qaoa_p} ({self.qaoa_mode})"))

        self.solver_comparison_df = pd.concat(solver_frames, ignore_index=True) if solver_frames else pd.DataFrame()
        if len(self.solver_comparison_df):
            self.solver_comparison_df = self.solver_comparison_df[
                [
                    "solver",
                    "bitstring",
                    "qubo_value",
                    "return_term",
                    "risk_term",
                    "budget_term",
                    "qubo_reconstructed",
                    "avg_return_term_per_option",
                    "avg_risk_term_per_option",
                    "avg_budget_term_per_option",
                    "selected_usd",
                    "budget_gap",
                    "abs_budget_gap",
                    "num_options",
                    "num_distinct_assets",
                    "portfolio_return",
                    "portfolio_vol",
                    "sharpe_like",
                    "cash_weight",
                    "portfolio_return_budget_normalized",
                    "portfolio_vol_budget_normalized",
                    "sharpe_like_budget_normalized",
                    "max_position_usd",
                ]
            ]

    @staticmethod
    def _write_df(ws, df, dark_fill, white_bold):
        for j, col in enumerate(df.columns, start=1):
            ws.cell(1, j, col)
            ws.cell(1, j).fill = dark_fill
            ws.cell(1, j).font = white_bold
        for i, row in enumerate(df.itertuples(index=False), start=2):
            for j, val in enumerate(row, start=1):
                ws.cell(i, j, val)

    def write_results(self):
        self._ensure_not_stopped()
        self._log("Writing Version 4 results back to Excel...")

        wb = load_workbook(self.xlsx_path)
        for name in [
            "Results_Summary",
            "Results_Overview",
            "Results_Portfolios",
            "QAOA_Samples",
            "Classical_Candidates",
            "Solver_Comparison",
            "Optimization_History",
        ]:
            if name in wb.sheetnames:
                wb.remove(wb[name])

        dark = PatternFill("solid", fgColor="1F4E78")
        white_bold = Font(color="FFFFFF", bold=True)

        summary_ws = wb.create_sheet("Results_Summary")
        summary_ws["A1"] = "Results summary"
        summary_ws["A1"].fill = dark
        summary_ws["A1"].font = white_bold

        summary_items = [
            ("Top-N exported", int(len(self.overview_df))),
            ("Classical candidate count", int(len(self.classical_results))),
            ("QAOA candidate count", int(len(self.samples_df))),
            ("QAOA enabled", int(self.enable_qaoa)),
            ("QAOA p", int(self.qaoa_p)),
            ("QAOA mode", self.qaoa_mode),
            ("Budget lambda", float(self.lambda_budget)),
            ("Risk lambda", float(self.lambda_variance)),
            ("Risk-free rate", float(self.risk_free)),
            ("QAOA export mode", self.qaoa_export_mode),
            ("QAOA export sort by", self.qaoa_export_sort_by),
            ("QAOA multistart restarts", int(self.qaoa_multistart_restarts)),
            ("QAOA layerwise warm start", int(self.qaoa_layerwise_warm_start)),
            ("QAOA restart perturbation", float(self.qaoa_restart_perturbation)),
            ("QAOA ETA layer growth per layer", float(self.qaoa_eta_layer_growth_per_layer)),
            ("QAOA exact p1 presearch", int(self.qaoa_exact_p1_presearch)),
            ("QAOA exact p1 presearch max qubits", int(self.qaoa_exact_p1_presearch_max_qubits)),
            ("QAOA exact p1 presearch gamma min", float(self.qaoa_exact_p1_presearch_gamma_min)),
            ("QAOA exact p1 presearch gamma max", float(self.qaoa_exact_p1_presearch_gamma_max)),
            ("QAOA exact p1 presearch gamma points", int(self.qaoa_exact_p1_presearch_gamma_points)),
            ("QAOA exact p1 presearch beta min", float(self.qaoa_exact_p1_presearch_beta_min)),
            ("QAOA exact p1 presearch beta max", float(self.qaoa_exact_p1_presearch_beta_max)),
            ("QAOA exact p1 presearch beta points", int(self.qaoa_exact_p1_presearch_beta_points)),
            ("QAOA min probability to export", float(self.qaoa_min_probability_to_export)),
            ("QAOA max export rows", int(self.qaoa_max_export_rows)),
            ("QAOA feasible only", int(self.qaoa_export_feasible_only)),
            (
                "QAOA feasibility budget tolerance USD",
                float(self.qaoa_feasibility_budget_tolerance_usd),
            ),
            ("Decision variables", int(self.n)),
            ("Unique tickers", int(len(self.asset_universe))),
            ("QAOA total states considered", int(self.qaoa_total_states_considered)),
            ("QAOA nonzero states encountered", int(self.qaoa_total_nonzero_states)),
            ("QAOA final energy", float(self.qaoa_final_energy) if pd.notna(self.qaoa_final_energy) else None),
            ("QAOA expected QUBO", float(self.qaoa_expected_qubo) if pd.notna(self.qaoa_expected_qubo) else None),
            ("QAOA max state probability", float(self.qaoa_max_state_probability) if pd.notna(self.qaoa_max_state_probability) else None),
            ("QAOA top10 probability mass", float(self.qaoa_top10_probability_mass) if pd.notna(self.qaoa_top10_probability_mass) else None),
            ("QAOA top-probability state QUBO", float(self.qaoa_top_probability_state_qubo) if pd.notna(self.qaoa_top_probability_state_qubo) else None),
            ("QAOA best exported QUBO", float(self.qaoa_best_exported_qubo) if pd.notna(self.qaoa_best_exported_qubo) else None),
            ("QAOA best exported probability", float(self.qaoa_best_exported_probability) if pd.notna(self.qaoa_best_exported_probability) else None),
            ("QAOA p1 presearch used", int(self.qaoa_p1_presearch_used)),
            ("QAOA p1 presearch gamma", float(self.qaoa_p1_presearch_gamma) if pd.notna(self.qaoa_p1_presearch_gamma) else None),
            ("QAOA p1 presearch beta", float(self.qaoa_p1_presearch_beta) if pd.notna(self.qaoa_p1_presearch_beta) else None),
            ("QAOA p1 presearch energy", float(self.qaoa_p1_presearch_energy) if pd.notna(self.qaoa_p1_presearch_energy) else None),
            ("QAOA p1 presearch expected QUBO", float(self.qaoa_p1_presearch_expected_qubo) if pd.notna(self.qaoa_p1_presearch_expected_qubo) else None),
        ]
        summary_items += [
            (
                f"QAOA gamma_{idx + 1}",
                float(self.best_gammas[idx]) if idx < len(self.best_gammas) and np.isfinite(self.best_gammas[idx]) else None,
            )
            for idx in range(self.qaoa_p)
        ]
        summary_items += [
            (
                f"QAOA beta_{idx + 1}",
                float(self.best_betas[idx]) if idx < len(self.best_betas) and np.isfinite(self.best_betas[idx]) else None,
            )
            for idx in range(self.qaoa_p)
        ]
        summary_items += [
            ("Best overview Sharpe-like", float(self.overview_df["sharpe_like"].max()) if len(self.overview_df) else None),
            ("Best overview invested USD", float(self.overview_df.iloc[0]["selected_usd"]) if len(self.overview_df) else None),
            ("Best overview abs budget gap", float(self.overview_df.iloc[0]["abs_budget_gap"]) if len(self.overview_df) else None),
            ("Best overview return term", float(self.overview_df.iloc[0]["return_term"]) if len(self.overview_df) else None),
            ("Best overview risk term", float(self.overview_df.iloc[0]["risk_term"]) if len(self.overview_df) else None),
            ("Best overview budget term", float(self.overview_df.iloc[0]["budget_term"]) if len(self.overview_df) else None),
            ("Best overview cash weight", float(self.overview_df.iloc[0]["cash_weight"]) if len(self.overview_df) else None),
            (
                "Best overview return budget-normalized",
                float(self.overview_df.iloc[0]["portfolio_return_budget_normalized"]) if len(self.overview_df) else None,
            ),
            (
                "Best overview volatility budget-normalized",
                float(self.overview_df.iloc[0]["portfolio_vol_budget_normalized"]) if len(self.overview_df) else None,
            ),
            (
                "Best overview Sharpe budget-normalized",
                float(self.overview_df.iloc[0]["sharpe_like_budget_normalized"]) if len(self.overview_df) else None,
            ),
            ("Optimization iterations", int(len(self.history_df))),
        ]
        for idx, (label, value) in enumerate(summary_items, start=2):
            summary_ws.cell(idx, 1, label)
            summary_ws.cell(idx, 2, value)

        overview_ws = wb.create_sheet("Results_Overview")
        self._write_df(overview_ws, self.overview_df, dark, white_bold)

        portfolios_ws = wb.create_sheet("Results_Portfolios")
        self._write_df(portfolios_ws, self.portfolios_df, dark, white_bold)

        qaoa_ws = wb.create_sheet("QAOA_Samples")
        self._write_df(qaoa_ws, self.samples_df.head(self.result_candidate_limit_per_solver), dark, white_bold)

        classical_ws = wb.create_sheet("Classical_Candidates")
        self._write_df(
            classical_ws,
            self.classical_results.head(self.result_candidate_limit_per_solver),
            dark,
            white_bold,
        )

        solver_ws = wb.create_sheet("Solver_Comparison")
        self._write_df(solver_ws, self.solver_comparison_df, dark, white_bold)

        history_ws = wb.create_sheet("Optimization_History")
        self._write_df(history_ws, self.history_df, dark, white_bold)

        wb.save(self.xlsx_path)
        self._log(f"Results written to {self.xlsx_path.resolve()}")

    def run_all(self):
        self._progress("Loading workbook...", 1)
        self.load_input()
        self._ensure_not_stopped()
        progress_plan = self._progress_plan()

        self._progress("Refreshing market data...", progress_plan["refresh"])
        self.refresh_market_data()
        self._ensure_not_stopped()

        self._progress("Building optimization problem...", progress_plan["build"])
        self.build_qubo()
        self._ensure_not_stopped()

        self._progress("Running classical search...", progress_plan["classical_start"])
        self.run_classical_search()
        self._ensure_not_stopped()

        self._progress("Running QAOA...", progress_plan["qaoa_start"])
        self.run_qaoa()
        self._ensure_not_stopped()

        self._progress("Generating results...", progress_plan["generate"])
        self.generate_results()
        self._ensure_not_stopped()

        self._progress("Writing results...", progress_plan["write"])
        self.write_results()
        self._progress("Optimization complete!", 100)
        return self


QAOAOptimizerV4 = QAOAOptimizerV3
